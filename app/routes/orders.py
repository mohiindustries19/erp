"""
Order Management Routes - Complete CRUD
"""
from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, current_app
from flask_login import login_required
from app import db
from app.models import Order, OrderItem, Distributor, Product, Warehouse
from app.services.accounting_utils import (
    create_accounting_entry,
    delete_posting,
    resolve_receivable_account,
    resolve_sales_account,
    resolve_output_cgst,
    resolve_output_sgst,
    resolve_output_igst,
)
from datetime import datetime, date
from decimal import Decimal

bp = Blueprint('orders', __name__, url_prefix='/orders')


def _order_has_payments(order: Order) -> bool:
    try:
        return bool(order.payments) and len(order.payments) > 0
    except Exception:
        # Be conservative if relationship loading behaves unexpectedly
        return True

@bp.route('/')
@login_required
def list_orders():
    """List all orders"""
    orders = Order.query.order_by(Order.created_at.desc()).all()
    return render_template('orders/list.html', orders=orders)

@bp.route('/add', methods=['GET', 'POST'])
@login_required
def add_order():
    """Create new order"""
    if request.method == 'POST':
        try:
            # Get form data
            items_data = request.form.getlist('items')  # Will be sent as JSON
            
            # If no items, check if it's JSON request
            if not items_data:
                # Try to get from JSON body
                import json
                items_json = request.form.get('items_json')
                if items_json:
                    items_data = json.loads(items_json)
                else:
                    flash('Please add at least one item to the order!', 'error')
                    return redirect(url_for('orders.add_order'))
            
            # Generate order number
            today = date.today()
            order_prefix = f'ORD{today.strftime("%Y%m%d")}'
            last_order = Order.query.filter(Order.order_number.like(f'{order_prefix}%')).order_by(Order.id.desc()).first()
            
            if last_order:
                last_num = int(last_order.order_number[-4:])
                order_number = f'{order_prefix}{last_num+1:04d}'
            else:
                order_number = f'{order_prefix}0001'
            
            # Create order
            order = Order(
                order_number=order_number,
                distributor_id=request.form.get('distributor_id'),
                order_date=datetime.strptime(request.form.get('order_date'), '%Y-%m-%d').date(),
                order_type=request.form.get('order_type', 'regular'),
                payment_terms=request.form.get('payment_terms'),
                remarks=request.form.get('remarks'),
                status='confirmed',
                payment_status='pending',
                paid_amount=0.0
            )
            
            db.session.add(order)
            db.session.flush()  # Get order ID
            
            # Add order items
            added_items = 0
            for item_data in items_data:
                if isinstance(item_data, str):
                    import json
                    item_data = json.loads(item_data)
                
                product = Product.query.get(item_data['product_id'])
                if not product:
                    continue
                
                quantity = int(item_data['quantity'])
                discount_percent = float(item_data.get('discount', 0))
                client_unit_price = float(item_data.get('unit_price', product.base_price))
                server_unit_price = float(product.base_price)
                
                if quantity <= 0:
                    flash('Quantity must be greater than zero for all items.', 'error')
                    db.session.rollback()
                    return redirect(url_for('orders.add_order'))
                
                if discount_percent < 0 or discount_percent > 100:
                    flash('Discount must be between 0 and 100.', 'error')
                    db.session.rollback()
                    return redirect(url_for('orders.add_order'))
                
                # Enforce server-side pricing
                if abs(client_unit_price - server_unit_price) > 0.01:
                    flash(f'Price mismatch for {product.name}. Please refresh and try again.', 'error')
                    db.session.rollback()
                    return redirect(url_for('orders.add_order'))
                
                order_item = OrderItem(
                    order_id=order.id,
                    product_id=product.id,
                    quantity=quantity,
                    unit_price=server_unit_price,
                    discount_percent=discount_percent,
                    hsn_code=product.hsn_code,
                    gst_rate=product.gst_rate
                )
                
                order_item.calculate_line_total()
                db.session.add(order_item)
                added_items += 1
            
            if added_items == 0:
                flash('Please add at least one valid item to the order!', 'error')
                db.session.rollback()
                return redirect(url_for('orders.add_order'))
            
            # Calculate order totals
            order.calculate_totals()
            
            min_order_value = float(current_app.config.get('MIN_ORDER_VALUE', 0))
            if order.total_amount < min_order_value:
                flash(f'Minimum order value is ₹{min_order_value:,.2f}.', 'error')
                db.session.rollback()
                return redirect(url_for('orders.add_order'))
            
            # Post accounting entries for the sale (simple ledger)
            delete_posting('order', order.id)
            ar = resolve_receivable_account()
            sales = resolve_sales_account()
            out_cgst = resolve_output_cgst()
            out_sgst = resolve_output_sgst()
            out_igst = resolve_output_igst()

            customer = order.distributor.business_name if order.distributor else 'Customer'
            desc = f'Sale {order.order_number} to {customer}'

            create_accounting_entry(
                entry_date=order.order_date,
                reference_type='order',
                reference_id=order.id,
                account=ar,
                debit=Decimal(str(order.total_amount or 0)),
                credit=Decimal('0'),
                description=desc,
            )
            create_accounting_entry(
                entry_date=order.order_date,
                reference_type='order',
                reference_id=order.id,
                account=sales,
                debit=Decimal('0'),
                credit=Decimal(str(order.taxable_amount or 0)),
                description=desc,
            )
            if (order.cgst_amount or 0) > 0:
                create_accounting_entry(
                    entry_date=order.order_date,
                    reference_type='order',
                    reference_id=order.id,
                    account=out_cgst,
                    debit=Decimal('0'),
                    credit=Decimal(str(order.cgst_amount or 0)),
                    description=f'{desc} - CGST',
                )
            if (order.sgst_amount or 0) > 0:
                create_accounting_entry(
                    entry_date=order.order_date,
                    reference_type='order',
                    reference_id=order.id,
                    account=out_sgst,
                    debit=Decimal('0'),
                    credit=Decimal(str(order.sgst_amount or 0)),
                    description=f'{desc} - SGST',
                )
            if (order.igst_amount or 0) > 0:
                create_accounting_entry(
                    entry_date=order.order_date,
                    reference_type='order',
                    reference_id=order.id,
                    account=out_igst,
                    debit=Decimal('0'),
                    credit=Decimal(str(order.igst_amount or 0)),
                    description=f'{desc} - IGST',
                )

            db.session.commit()
            
            flash(f'Order {order_number} created successfully!', 'success')
            return redirect(url_for('orders.view_order', id=order.id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error creating order: {str(e)}', 'error')
            return redirect(url_for('orders.add_order'))
    
    # GET request
    distributors = Distributor.query.filter_by(status='active').all()
    products = Product.query.filter_by(is_active=True).all()
    today = date.today().strftime('%Y-%m-%d')
    
    return render_template('orders/add.html', 
                         distributors=distributors, 
                         products=products,
                         today=today)

@bp.route('/<int:id>')
@login_required
def view_order(id):
    """View order details"""
    order = Order.query.get_or_404(id)
    return render_template('orders/view.html', order=order)

@bp.route('/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def edit_order(id):
    """Edit existing order"""
    order = Order.query.get_or_404(id)
    
    if request.method == 'POST':
        try:
            order.order_date = datetime.strptime(request.form.get('order_date'), '%Y-%m-%d').date()
            order.order_type = request.form.get('order_type')
            order.payment_terms = request.form.get('payment_terms')
            order.remarks = request.form.get('remarks')
            order.status = request.form.get('status')
            
            db.session.commit()
            flash(f'Order {order.order_number} updated successfully!', 'success')
            return redirect(url_for('orders.view_order', id=order.id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating order: {str(e)}', 'error')
    
    distributors = Distributor.query.filter_by(status='active').all()
    products = Product.query.filter_by(is_active=True).all()
    
    return render_template('orders/edit.html', 
                         order=order,
                         distributors=distributors,
                         products=products)

@bp.route('/<int:id>/delete', methods=['POST'])
@login_required
def delete_order(id):
    """Delete order"""
    order = Order.query.get_or_404(id)
    
    try:
        if _order_has_payments(order):
            flash('Cannot delete an order that has payments. Please delete/cancel payments first.', 'error')
            return redirect(url_for('orders.view_order', id=order.id))

        if order.status not in ('draft', 'cancelled'):
            flash('Only Draft/Cancelled orders can be deleted. Cancel the order first if needed.', 'error')
            return redirect(url_for('orders.view_order', id=order.id))

        order_number = order.order_number
        delete_posting('order', order.id)
        db.session.delete(order)
        db.session.commit()
        flash(f'Order {order_number} deleted successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting order: {str(e)}', 'error')
    
    return redirect(url_for('orders.list_orders'))


@bp.route('/<int:id>/cancel', methods=['POST'])
@login_required
def cancel_order(id):
    """Cancel an order (keeps record, removes sales postings)."""
    order = Order.query.get_or_404(id)

    try:
        if _order_has_payments(order):
            flash('Cannot cancel an order that has payments. Please reverse/delete payments first.', 'error')
            return redirect(url_for('orders.view_order', id=order.id))

        if order.status == 'delivered':
            flash('Delivered orders cannot be cancelled.', 'error')
            return redirect(url_for('orders.view_order', id=order.id))

        order.status = 'cancelled'
        # Remove postings so cancelled orders do not affect ledgers.
        delete_posting('order', order.id)
        db.session.commit()
        flash(f'Order {order.order_number} cancelled.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error cancelling order: {str(e)}', 'error')

    return redirect(url_for('orders.view_order', id=order.id))

@bp.route('/<int:id>/invoice')
@login_required
def print_invoice(id):
    """Print invoice for order"""
    order = Order.query.get_or_404(id)
    return render_template('orders/invoice.html', order=order)

@bp.route('/<int:id>/export-pdf')
@login_required
def export_pdf(id):
    """Export invoice as PDF - uses browser print-to-PDF"""
    # Redirect to print page with auto-print
    order = Order.query.get_or_404(id)
    return render_template('orders/invoice_download.html', order=order)

@bp.route('/<int:id>/export-excel')
@login_required
def export_excel(id):
    """Export invoice as Excel"""
    from flask import make_response
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO
    
    order = Order.query.get_or_404(id)
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Invoice {order.order_number}"
    
    # Styles
    header_fill = PatternFill(start_color="D00000", end_color="D00000", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=16, color="D00000")
    bold_font = Font(bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Company Header
    ws.merge_cells('A1:G1')
    ws['A1'] = 'MOHI INDUSTRIES'
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:G2')
    ws['A2'] = '4-1, Plot No G-2, Industrial Area Road, Hajipur Industrial Area, Hajipur, Bihar 844102'
    ws['A2'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A3:G3')
    ws['A3'] = 'Phone: +91 9262650010 | Email: info@mohiindustries.in | GSTIN: 10GANPS5418H1ZJ | FSSAI: 10423110000282'
    ws['A3'].alignment = Alignment(horizontal='center')
    
    # Invoice Title
    ws.merge_cells('A5:G5')
    ws['A5'] = 'TAX INVOICE'
    ws['A5'].font = Font(bold=True, size=14)
    ws['A5'].alignment = Alignment(horizontal='center')
    
    # Invoice Details
    row = 7
    ws[f'A{row}'] = 'Invoice No:'
    ws[f'A{row}'].font = bold_font
    ws[f'B{row}'] = order.order_number
    ws[f'E{row}'] = 'Date:'
    ws[f'E{row}'].font = bold_font
    ws[f'F{row}'] = order.order_date.strftime('%d-%m-%Y')
    
    row += 1
    ws[f'A{row}'] = 'Status:'
    ws[f'A{row}'].font = bold_font
    ws[f'B{row}'] = order.status.upper()
    ws[f'E{row}'] = 'Payment Status:'
    ws[f'E{row}'].font = bold_font
    ws[f'F{row}'] = order.payment_status.upper()
    
    # Customer Details
    row += 2
    ws[f'A{row}'] = 'BILL TO:'
    ws[f'A{row}'].font = Font(bold=True, color="D00000")
    row += 1
    ws[f'A{row}'] = order.distributor.business_name
    ws[f'A{row}'].font = bold_font
    row += 1
    ws[f'A{row}'] = order.distributor.address_line1
    if order.distributor.address_line2:
        row += 1
        ws[f'A{row}'] = order.distributor.address_line2
    row += 1
    ws[f'A{row}'] = f"{order.distributor.city}, {order.distributor.state} - {order.distributor.pincode}"
    row += 1
    ws[f'A{row}'] = f"GSTIN: {order.distributor.gstin or 'Unregistered'}"
    row += 1
    ws[f'A{row}'] = f"Phone: {order.distributor.phone}"
    
    # Items Table Header
    row += 2
    headers = ['#', 'Product', 'HSN', 'Qty', 'Rate', 'Disc%', 'Taxable', 'GST%', 'Amount']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # Items
    row += 1
    for idx, item in enumerate(order.items, start=1):
        ws.cell(row=row, column=1).value = idx
        ws.cell(row=row, column=2).value = item.product.name
        ws.cell(row=row, column=3).value = item.hsn_code
        ws.cell(row=row, column=4).value = item.quantity
        ws.cell(row=row, column=5).value = f"₹{item.unit_price:.2f}"
        ws.cell(row=row, column=6).value = f"{item.discount_percent}%"
        ws.cell(row=row, column=7).value = f"₹{item.line_total:.2f}"
        ws.cell(row=row, column=8).value = f"{item.gst_rate}%"
        ws.cell(row=row, column=9).value = f"₹{item.line_total * (1 + item.gst_rate/100):.2f}"
        
        for col in range(1, 10):
            ws.cell(row=row, column=col).border = border
        
        row += 1
    
    # Totals
    row += 1
    ws.cell(row=row, column=7).value = 'Subtotal:'
    ws.cell(row=row, column=7).font = bold_font
    ws.cell(row=row, column=9).value = f"₹{order.subtotal:.2f}"
    
    if order.discount_amount > 0:
        row += 1
        ws.cell(row=row, column=7).value = 'Discount:'
        ws.cell(row=row, column=7).font = bold_font
        ws.cell(row=row, column=9).value = f"-₹{order.discount_amount:.2f}"
    
    row += 1
    ws.cell(row=row, column=7).value = 'Taxable Amount:'
    ws.cell(row=row, column=7).font = bold_font
    ws.cell(row=row, column=9).value = f"₹{order.taxable_amount:.2f}"
    
    if order.cgst_amount > 0:
        row += 1
        ws.cell(row=row, column=7).value = 'CGST:'
        ws.cell(row=row, column=7).font = bold_font
        ws.cell(row=row, column=9).value = f"₹{order.cgst_amount:.2f}"
        
        row += 1
        ws.cell(row=row, column=7).value = 'SGST:'
        ws.cell(row=row, column=7).font = bold_font
        ws.cell(row=row, column=9).value = f"₹{order.sgst_amount:.2f}"
    
    if order.igst_amount > 0:
        row += 1
        ws.cell(row=row, column=7).value = 'IGST:'
        ws.cell(row=row, column=7).font = bold_font
        ws.cell(row=row, column=9).value = f"₹{order.igst_amount:.2f}"
    
    row += 1
    ws.cell(row=row, column=7).value = 'TOTAL AMOUNT:'
    ws.cell(row=row, column=7).font = Font(bold=True, size=12, color="D00000")
    ws.cell(row=row, column=9).value = f"₹{order.total_amount:.2f}"
    ws.cell(row=row, column=9).font = Font(bold=True, size=12, color="D00000")
    
    # Payment Status
    row += 2
    outstanding = order.total_amount - order.paid_amount
    ws.cell(row=row, column=7).value = 'Amount Paid:'
    ws.cell(row=row, column=7).font = bold_font
    ws.cell(row=row, column=9).value = f"₹{order.paid_amount:.2f}"
    
    row += 1
    ws.cell(row=row, column=7).value = 'Balance Due:'
    ws.cell(row=row, column=7).font = Font(bold=True, color="D00000")
    ws.cell(row=row, column=9).value = f"₹{outstanding:.2f}"
    ws.cell(row=row, column=9).font = Font(bold=True, color="D00000")
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 8
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 8
    ws.column_dimensions['I'].width = 12
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Create response
    response = make_response(output.read())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=Invoice_{order.order_number}.xlsx'
    
    return response

@bp.route('/<int:id>/send-email', methods=['POST'])
@login_required
def send_email(id):
    """Send invoice via email"""
    from flask_mail import Message, Mail
    from flask import current_app
    
    order = Order.query.get_or_404(id)
    
    # Get email addresses
    to_email = request.form.get('to_email') or order.distributor.email
    cc_email = request.form.get('cc_email')  # Optional CC
    
    if not to_email:
        flash('Distributor email not found. Please update distributor profile.', 'error')
        return redirect(url_for('orders.view_order', id=order.id))
    
    try:
        # Generate PDF (optional). On Windows, WeasyPrint may require GTK runtime.
        pdf_bytes = None
        pdf_error = None
        try:
            from weasyprint import HTML  # type: ignore

            html_string = render_template('orders/invoice.html', order=order)
            pdf_bytes = HTML(string=html_string, base_url=request.url_root).write_pdf()
        except Exception as e:
            pdf_error = str(e)
        
        # Create email
        mail = Mail(current_app)
        msg = Message(
            subject=f'Invoice {order.order_number} - Mohi Industries',
            sender=current_app.config.get('MAIL_DEFAULT_SENDER'),
            recipients=[to_email]
        )
        
        if cc_email:
            msg.cc = [cc_email]
        
        # Email body
        outstanding = order.total_amount - order.paid_amount
        invoice_link = url_for('orders.print_invoice', id=order.id, _external=True)
        msg.body = f"""Dear {order.distributor.business_name},

Please find attached invoice for your order.

Invoice Details:
- Invoice No: {order.order_number}
- Date: {order.order_date.strftime('%d-%m-%Y')}
- Total Amount: ₹{order.total_amount:,.2f}
- Amount Paid: ₹{order.paid_amount:,.2f}
- Balance Due: ₹{outstanding:,.2f}

Payment Terms: {order.payment_terms or 'As per agreement'}

Invoice link (view/print): {invoice_link}

Bank Details:
Bank Name: State Bank of India
Account No: 1234567890
IFSC Code: SBIN0001234
Branch: Hajipur

Thank you for your business!

Best regards,
Mohi Industries
Phone: +91 9262650010
Email: info@mohiindustries.in
"""

        # Attach PDF if available, otherwise send without attachment.
        if pdf_bytes:
            msg.attach(
                f"Invoice_{order.order_number}.pdf",
                "application/pdf",
                pdf_bytes,
            )
        
        # Send email
        mail.send(msg)

        if pdf_bytes:
            flash(f'Invoice sent successfully to {to_email}!', 'success')
        else:
            flash(
                'Invoice email sent, but PDF attachment was skipped. '
                f'PDF generation error: {pdf_error}',
                'warning',
            )
        
    except Exception as e:
        flash(f'Error sending email: {str(e)}', 'error')
    
    return redirect(url_for('orders.view_order', id=order.id))

