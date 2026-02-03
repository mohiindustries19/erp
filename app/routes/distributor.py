"""
Distributor Management Routes - Complete CRUD
"""
from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file
from flask_login import login_required
from app import db
from app.models import Distributor
from datetime import datetime, date
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO

bp = Blueprint('distributor', __name__, url_prefix='/distributors')

@bp.route('/')
@login_required
def list_distributors():
    """List all distributors"""
    distributors = Distributor.query.order_by(Distributor.created_at.desc()).all()
    return render_template('distributors/list.html', distributors=distributors)

@bp.route('/export-excel')
@login_required
def export_excel():
    """Export all distributors to Excel"""
    distributors = Distributor.query.order_by(Distributor.code).all()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Distributors"
    
    # Header styling
    header_fill = PatternFill(start_color="D00000", end_color="D00000", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = [
        'Code', 'Business Name', 'Contact Person', 'Phone', 'Email',
        'GSTIN', 'PAN', 'City', 'State', 'Pincode',
        'Territory', 'Margin %', 'Credit Limit', 'Credit Days',
        'Payment Terms', 'Status', 'Onboarding Date'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Data rows
    for row_idx, dist in enumerate(distributors, 2):
        ws.cell(row=row_idx, column=1, value=dist.code)
        ws.cell(row=row_idx, column=2, value=dist.business_name)
        ws.cell(row=row_idx, column=3, value=dist.contact_person)
        ws.cell(row=row_idx, column=4, value=dist.phone)
        ws.cell(row=row_idx, column=5, value=dist.email)
        ws.cell(row=row_idx, column=6, value=dist.gstin)
        ws.cell(row=row_idx, column=7, value=dist.pan)
        ws.cell(row=row_idx, column=8, value=dist.city)
        ws.cell(row=row_idx, column=9, value=dist.state)
        ws.cell(row=row_idx, column=10, value=dist.pincode)
        ws.cell(row=row_idx, column=11, value=dist.territory)
        ws.cell(row=row_idx, column=12, value=dist.margin_percentage)
        ws.cell(row=row_idx, column=13, value=dist.credit_limit)
        ws.cell(row=row_idx, column=14, value=dist.credit_days)
        ws.cell(row=row_idx, column=15, value=dist.payment_terms)
        ws.cell(row=row_idx, column=16, value=dist.status)
        ws.cell(row=row_idx, column=17, value=dist.onboarding_date.strftime('%d-%m-%Y') if dist.onboarding_date else '')
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 10
    ws.column_dimensions['K'].width = 15
    ws.column_dimensions['L'].width = 10
    ws.column_dimensions['M'].width = 15
    ws.column_dimensions['N'].width = 12
    ws.column_dimensions['O'].width = 15
    ws.column_dimensions['P'].width = 10
    ws.column_dimensions['Q'].width = 15
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f'distributors_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@bp.route('/add', methods=['GET', 'POST'])
@login_required
def add_distributor():
    """Create new distributor"""
    if request.method == 'POST':
        try:
            # Generate distributor code
            last_dist = Distributor.query.order_by(Distributor.id.desc()).first()
            next_num = (last_dist.id + 1) if last_dist else 1
            code = f'DIST{next_num:04d}'
            
            distributor = Distributor(
                code=code,
                business_name=request.form.get('business_name'),
                contact_person=request.form.get('contact_person'),
                phone=request.form.get('phone'),
                email=request.form.get('email'),
                gstin=request.form.get('gstin'),
                pan=request.form.get('pan'),
                address_line1=request.form.get('address_line1'),
                address_line2=request.form.get('address_line2'),
                city=request.form.get('city'),
                state=request.form.get('state'),
                state_code=request.form.get('state_code'),
                pincode=request.form.get('pincode'),
                territory=request.form.get('territory'),
                margin_percentage=float(request.form.get('margin_percentage', 12)),
                credit_limit=float(request.form.get('credit_limit', 0)),
                credit_days=int(request.form.get('credit_days', 0)),
                payment_terms=request.form.get('payment_terms', 'advance'),
                status='active',
                onboarding_date=date.today()
            )
            
            db.session.add(distributor)
            db.session.commit()
            
            flash(f'Distributor {code} - {distributor.business_name} added successfully!', 'success')
            return redirect(url_for('distributor.view_distributor', id=distributor.id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error adding distributor: {str(e)}', 'error')
    
    return render_template('distributors/add.html')

@bp.route('/<int:id>')
@login_required
def view_distributor(id):
    """View distributor details"""
    distributor = Distributor.query.get_or_404(id)
    return render_template('distributors/view.html', distributor=distributor)

@bp.route('/<int:id>/print')
@login_required
def print_distributor(id):
    """Print distributor profile"""
    distributor = Distributor.query.get_or_404(id)
    return render_template('distributors/distributor_print.html', distributor=distributor)

@bp.route('/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def edit_distributor(id):
    """Edit existing distributor"""
    distributor = Distributor.query.get_or_404(id)
    
    if request.method == 'POST':
        try:
            distributor.business_name = request.form.get('business_name')
            distributor.contact_person = request.form.get('contact_person')
            distributor.phone = request.form.get('phone')
            distributor.email = request.form.get('email')
            distributor.gstin = request.form.get('gstin')
            distributor.pan = request.form.get('pan')
            distributor.address_line1 = request.form.get('address_line1')
            distributor.address_line2 = request.form.get('address_line2')
            distributor.city = request.form.get('city')
            distributor.state = request.form.get('state')
            distributor.state_code = request.form.get('state_code')
            distributor.pincode = request.form.get('pincode')
            distributor.territory = request.form.get('territory')
            distributor.margin_percentage = float(request.form.get('margin_percentage', 12))
            distributor.credit_limit = float(request.form.get('credit_limit', 0))
            distributor.credit_days = int(request.form.get('credit_days', 0))
            distributor.payment_terms = request.form.get('payment_terms')
            distributor.status = request.form.get('status')
            
            db.session.commit()
            flash(f'Distributor {distributor.code} updated successfully!', 'success')
            return redirect(url_for('distributor.view_distributor', id=distributor.id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating distributor: {str(e)}', 'error')
    
    return render_template('distributors/edit.html', distributor=distributor)

@bp.route('/<int:id>/delete', methods=['POST'])
@login_required
def delete_distributor(id):
    """Delete distributor"""
    distributor = Distributor.query.get_or_404(id)
    
    try:
        # Check if distributor has orders
        if distributor.orders.count() > 0:
            flash(f'Cannot delete distributor {distributor.code} - has existing orders. Set status to inactive instead.', 'error')
            return redirect(url_for('distributor.view_distributor', id=id))
        
        code = distributor.code
        db.session.delete(distributor)
        db.session.commit()
        flash(f'Distributor {code} deleted successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting distributor: {str(e)}', 'error')
    
    return redirect(url_for('distributor.list_distributors'))

@bp.route('/api/search')
@login_required
def search_distributors():
    """Search distributors API"""
    query = request.args.get('q', '')
    distributors = Distributor.query.filter(
        (Distributor.business_name.ilike(f'%{query}%')) |
        (Distributor.code.ilike(f'%{query}%'))
    ).limit(10).all()
    
    return jsonify([{
        'id': d.id,
        'code': d.code,
        'business_name': d.business_name,
        'phone': d.phone
    } for d in distributors])
