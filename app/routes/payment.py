"""
Payment Management Routes - Payment Tracking and Accounting
"""
from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file
from flask_login import login_required, current_user
from app import db
from app.models import Payment, Order
from app.models.accounting import AccountingEntry
from app.services.accounting_utils import (
    create_accounting_entry,
    delete_posting,
    resolve_payment_account,
    resolve_receivable_account,
)
from datetime import datetime, date
from decimal import Decimal
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO

bp = Blueprint('payment', __name__, url_prefix='/payments')

@bp.route('/')
@login_required
def list_payments():
    """List all payments"""
    payments = Payment.query.order_by(Payment.payment_date.desc()).all()
    return render_template('payments/list.html', payments=payments)

@bp.route('/export-excel')
@login_required
def export_excel():
    """Export all payments to Excel"""
    payments = Payment.query.order_by(Payment.payment_date.desc()).all()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Payments"
    
    # Header styling
    header_fill = PatternFill(start_color="D00000", end_color="D00000", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = [
        'Payment #', 'Date', 'Order #', 'Customer', 'Amount',
        'Payment Mode', 'Reference #', 'Bank', 'Status',
        'Clearance Date', 'Remarks'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Data rows
    for row_idx, payment in enumerate(payments, 2):
        ws.cell(row=row_idx, column=1, value=payment.payment_number)
        ws.cell(row=row_idx, column=2, value=payment.payment_date.strftime('%d-%m-%Y'))
        ws.cell(row=row_idx, column=3, value=payment.order.order_number)
        ws.cell(row=row_idx, column=4, value=payment.order.distributor.business_name)
        ws.cell(row=row_idx, column=5, value=payment.amount)
        ws.cell(row=row_idx, column=6, value=payment.payment_mode)
        ws.cell(row=row_idx, column=7, value=payment.reference_number or '')
        ws.cell(row=row_idx, column=8, value=payment.bank_name or '')
        ws.cell(row=row_idx, column=9, value=payment.status)
        ws.cell(row=row_idx, column=10, value=payment.clearance_date.strftime('%d-%m-%Y') if payment.clearance_date else '')
        ws.cell(row=row_idx, column=11, value=payment.remarks or '')
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 30
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f'payments_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@bp.route('/<int:id>/print')
@login_required
def print_payment(id):
    """Print payment receipt"""
    payment = Payment.query.get_or_404(id)
    return render_template('payments/payment_receipt_print.html', payment=payment)

@bp.route('/pending')
@login_required
def pending_payments():
    """List pending payments (orders with unpaid balance)"""
    orders = Order.query.filter(
        Order.payment_status.in_(['pending', 'partial'])
    ).order_by(Order.order_date.desc()).all()
    
    return render_template('payments/pending.html', orders=orders)

@bp.route('/add/<int:order_id>', methods=['GET', 'POST'])
@login_required
def add_payment(order_id):
    """Record payment for an order"""
    order = Order.query.get_or_404(order_id)
    
    if request.method == 'POST':
        try:
            # Calculate outstanding amount
            total_paid = sum(p.amount for p in order.payments if p.status == 'cleared')
            outstanding = order.total_amount - total_paid
            
            payment_amount = float(request.form.get('amount'))
            if payment_amount <= 0:
                flash('Payment amount must be greater than zero.', 'error')
                return redirect(url_for('payment.add_payment', order_id=order_id))
            
            if payment_amount > outstanding:
                flash(f'Payment amount cannot exceed outstanding balance of ₹{outstanding:.2f}', 'error')
                return redirect(url_for('payment.add_payment', order_id=order_id))
            
            # Generate payment number
            today = date.today()
            payment_prefix = f'PAY{today.strftime("%Y%m%d")}'
            last_payment = Payment.query.filter(
                Payment.payment_number.like(f'{payment_prefix}%')
            ).order_by(Payment.id.desc()).first()
            
            if last_payment:
                last_num = int(last_payment.payment_number[-4:])
                payment_number = f'{payment_prefix}{last_num+1:04d}'
            else:
                payment_number = f'{payment_prefix}0001'
            
            # Create payment
            payment = Payment(
                payment_number=payment_number,
                order_id=order_id,
                payment_date=datetime.strptime(request.form.get('payment_date'), '%Y-%m-%d').date(),
                amount=payment_amount,
                payment_mode=request.form.get('payment_mode'),
                reference_number=request.form.get('reference_number'),
                bank_name=request.form.get('bank_name'),
                status=request.form.get('status', 'cleared'),
                remarks=request.form.get('remarks'),
                recorded_by=current_user.id
            )
            
            # Handle cheque date
            if request.form.get('cheque_date'):
                payment.cheque_date = datetime.strptime(request.form.get('cheque_date'), '%Y-%m-%d').date()
            
            # Set clearance date if cleared
            if payment.status == 'cleared':
                payment.clearance_date = payment.payment_date
            
            db.session.add(payment)
            
            # Update order payment status
            order.paid_amount = total_paid + payment_amount
            
            if order.paid_amount >= order.total_amount:
                order.payment_status = 'paid'
            elif order.paid_amount > 0:
                order.payment_status = 'partial'
            else:
                order.payment_status = 'pending'
            
            # Create accounting entries
            if payment.status == 'cleared':
                debit_account = resolve_payment_account(payment.payment_mode)
                ar_account = resolve_receivable_account()
                desc = f'Payment received from {order.distributor.business_name} for {order.order_number}'

                create_accounting_entry(
                    entry_date=payment.payment_date,
                    reference_type='payment',
                    reference_id=payment.id,
                    account=debit_account,
                    debit=Decimal(str(payment_amount)),
                    credit=Decimal('0'),
                    description=desc,
                    created_by=current_user.id,
                )
                create_accounting_entry(
                    entry_date=payment.payment_date,
                    reference_type='payment',
                    reference_id=payment.id,
                    account=ar_account,
                    debit=Decimal('0'),
                    credit=Decimal(str(payment_amount)),
                    description=desc,
                    created_by=current_user.id,
                )
            
            db.session.commit()
            
            flash(f'Payment {payment_number} recorded successfully!', 'success')
            return redirect(url_for('orders.view_order', id=order_id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error recording payment: {str(e)}', 'error')
    
    # Calculate outstanding
    total_paid = sum(p.amount for p in order.payments if p.status == 'cleared')
    outstanding = order.total_amount - total_paid
    
    today = date.today().strftime('%Y-%m-%d')
    return render_template('payments/add.html', order=order, outstanding=outstanding, today=today)

@bp.route('/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def edit_payment(id):
    """Edit payment details"""
    payment = Payment.query.get_or_404(id)
    
    if request.method == 'POST':
        try:
            old_amount = payment.amount
            old_status = payment.status
            
            new_amount = float(request.form.get('amount'))
            new_status = request.form.get('status')
            if new_amount <= 0:
                flash('Payment amount must be greater than zero.', 'error')
                return redirect(url_for('payment.edit_payment', id=id))
            
            # Prevent overpayment when marking/keeping as cleared
            order = payment.order
            cleared_excluding_current = sum(
                p.amount for p in order.payments if p.status == 'cleared' and p.id != payment.id
            )
            outstanding = order.total_amount - cleared_excluding_current
            if new_status == 'cleared' and new_amount > outstanding:
                flash(f'Payment amount cannot exceed outstanding balance of ₹{outstanding:.2f}', 'error')
                return redirect(url_for('payment.edit_payment', id=id))
            
            payment.payment_date = datetime.strptime(request.form.get('payment_date'), '%Y-%m-%d').date()
            payment.amount = new_amount
            payment.payment_mode = request.form.get('payment_mode')
            payment.reference_number = request.form.get('reference_number')
            payment.bank_name = request.form.get('bank_name')
            payment.status = new_status
            payment.remarks = request.form.get('remarks')
            
            if request.form.get('cheque_date'):
                payment.cheque_date = datetime.strptime(request.form.get('cheque_date'), '%Y-%m-%d').date()
            
            # Update clearance date
            if payment.status == 'cleared' and not payment.clearance_date:
                payment.clearance_date = date.today()
            
            # Recalculate order payment status
            total_paid = sum(p.amount for p in order.payments if p.status == 'cleared')
            order.paid_amount = total_paid
            
            if order.paid_amount >= order.total_amount:
                order.payment_status = 'paid'
            elif order.paid_amount > 0:
                order.payment_status = 'partial'
            else:
                order.payment_status = 'pending'
            
            # Keep accounting entries in sync with cleared status/amount/mode/date
            delete_posting('payment', payment.id)
            if payment.status == 'cleared':
                debit_account = resolve_payment_account(payment.payment_mode)
                ar_account = resolve_receivable_account()
                desc = f'Payment received from {order.distributor.business_name} for {order.order_number}'
                create_accounting_entry(
                    entry_date=payment.payment_date,
                    reference_type='payment',
                    reference_id=payment.id,
                    account=debit_account,
                    debit=Decimal(str(payment.amount)),
                    credit=Decimal('0'),
                    description=desc,
                    created_by=current_user.id,
                )
                create_accounting_entry(
                    entry_date=payment.payment_date,
                    reference_type='payment',
                    reference_id=payment.id,
                    account=ar_account,
                    debit=Decimal('0'),
                    credit=Decimal(str(payment.amount)),
                    description=desc,
                    created_by=current_user.id,
                )

            db.session.commit()
            flash(f'Payment {payment.payment_number} updated successfully!', 'success')
            return redirect(url_for('payment.list_payments'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating payment: {str(e)}', 'error')
    
    return render_template('payments/edit.html', payment=payment)

@bp.route('/<int:id>/delete', methods=['POST'])
@login_required
def delete_payment(id):
    """Delete payment"""
    payment = Payment.query.get_or_404(id)
    
    try:
        order = payment.order
        payment_number = payment.payment_number
        
        # Delete accounting entries
        AccountingEntry.query.filter_by(
            reference_type='payment',
            reference_id=payment.id
        ).delete()
        
        db.session.delete(payment)
        
        # Recalculate order payment status
        total_paid = sum(p.amount for p in order.payments if p.status == 'cleared')
        order.paid_amount = total_paid
        
        if order.paid_amount >= order.total_amount:
            order.payment_status = 'paid'
        elif order.paid_amount > 0:
            order.payment_status = 'partial'
        else:
            order.payment_status = 'pending'
        
        db.session.commit()
        flash(f'Payment {payment_number} deleted successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting payment: {str(e)}', 'error')
    
    return redirect(url_for('payment.list_payments'))

@bp.route('/receipt/<int:id>')
@login_required
def print_receipt(id):
    """Print payment receipt"""
    payment = Payment.query.get_or_404(id)
    return render_template('payments/receipt.html', payment=payment)
