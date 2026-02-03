"""
Quality Control Routes - Batch QC Workflow
"""
from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file
from flask_login import login_required, current_user
from app import db
from app.models import Batch, ProductCategory
from app.models.qc import QualityCheckTemplate, QualityCheckItem, BatchQualityCheck
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO

bp = Blueprint('qc', __name__, url_prefix='/qc')


@bp.route('/templates')
@login_required
def templates():
    templates = QualityCheckTemplate.query.order_by(QualityCheckTemplate.name).all()
    categories = ProductCategory.query.all()
    return render_template('qc/templates.html', templates=templates, categories=categories)

@bp.route('/templates/export-excel')
@login_required
def export_templates_excel():
    """Export all QC templates to Excel"""
    templates = QualityCheckTemplate.query.order_by(QualityCheckTemplate.name).all()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "QC Templates"
    
    # Header styling
    header_fill = PatternFill(start_color="D00000", end_color="D00000", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = ['Template Name', 'Category', 'Description', 'Check Items', 'Status']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Data rows
    for row_idx, template in enumerate(templates, 2):
        ws.cell(row=row_idx, column=1, value=template.name)
        ws.cell(row=row_idx, column=2, value=template.category.name if template.category else 'All')
        ws.cell(row=row_idx, column=3, value=template.description or '')
        ws.cell(row=row_idx, column=4, value=len(template.items))
        ws.cell(row=row_idx, column=5, value='Active' if template.is_active else 'Inactive')
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 10
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f'qc_templates_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@bp.route('/templates/add', methods=['GET', 'POST'])
@login_required
def add_template():
    if request.method == 'POST':
        try:
            template = QualityCheckTemplate(
                name=request.form.get('name'),
                category_id=request.form.get('category_id') or None,
                description=request.form.get('description'),
                is_active=request.form.get('is_active') == 'on'
            )
            db.session.add(template)
            db.session.flush()

            items_text = request.form.get('items', '')
            for line in [l.strip() for l in items_text.splitlines() if l.strip()]:
                item = QualityCheckItem(template_id=template.id, step=line)
                db.session.add(item)

            db.session.commit()
            flash('QC template created successfully!', 'success')
            return redirect(url_for('qc.templates'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error creating template: {str(e)}', 'error')

    categories = ProductCategory.query.all()
    return render_template('qc/template_add.html', categories=categories)


@bp.route('/templates/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def edit_template(id):
    template = QualityCheckTemplate.query.get_or_404(id)

    if request.method == 'POST':
        try:
            template.name = request.form.get('name')
            template.category_id = request.form.get('category_id') or None
            template.description = request.form.get('description')
            template.is_active = request.form.get('is_active') == 'on'

            QualityCheckItem.query.filter_by(template_id=template.id).delete()
            items_text = request.form.get('items', '')
            for line in [l.strip() for l in items_text.splitlines() if l.strip()]:
                item = QualityCheckItem(template_id=template.id, step=line)
                db.session.add(item)

            db.session.commit()
            flash('QC template updated successfully!', 'success')
            return redirect(url_for('qc.templates'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating template: {str(e)}', 'error')

    categories = ProductCategory.query.all()
    items_text = '\n'.join([item.step for item in template.items])
    return render_template('qc/template_edit.html', template=template, categories=categories, items_text=items_text)


@bp.route('/batches')
@login_required
def batches():
    batches = Batch.query.order_by(Batch.expiry_date).all()
    return render_template('qc/batches.html', batches=batches)


@bp.route('/batches/export-excel')
@login_required
def export_batches_excel():
    """Export all batches with QC status to Excel"""
    batches = Batch.query.order_by(Batch.expiry_date).all()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Batch QC Report"
    
    # Header styling
    header_fill = PatternFill(start_color="D00000", end_color="D00000", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = ['Batch Number', 'Product', 'Quantity', 'Unit', 'Expiry Date', 'QC Status', 'QC Date', 'Remarks']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Data rows
    for row_idx, batch in enumerate(batches, 2):
        ws.cell(row=row_idx, column=1, value=batch.batch_number)
        ws.cell(row=row_idx, column=2, value=batch.product.name)
        ws.cell(row=row_idx, column=3, value=batch.quantity)
        ws.cell(row=row_idx, column=4, value=batch.product.unit)
        ws.cell(row=row_idx, column=5, value=batch.expiry_date.strftime('%d-%m-%Y'))
        ws.cell(row=row_idx, column=6, value=batch.qc_status.upper() if batch.qc_status else 'PENDING')
        ws.cell(row=row_idx, column=7, value=batch.qc_date.strftime('%d-%m-%Y') if batch.qc_date else '')
        ws.cell(row=row_idx, column=8, value=batch.qc_remarks or '')
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 40
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f'batch_qc_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@bp.route('/batches/<int:id>/check', methods=['GET', 'POST'])
@login_required
def batch_check(id):
    batch = Batch.query.get_or_404(id)
    templates = QualityCheckTemplate.query.filter_by(is_active=True).all()

    if request.method == 'POST':
        try:
            template_id = int(request.form.get('template_id'))
            status = request.form.get('status')
            remarks = request.form.get('remarks')

            qc_record = BatchQualityCheck(
                batch_id=batch.id,
                template_id=template_id,
                status=status,
                remarks=remarks,
                checked_by=current_user.id,
                checked_at=datetime.utcnow()
            )
            db.session.add(qc_record)

            batch.qc_status = status
            batch.qc_date = datetime.utcnow().date()
            batch.qc_remarks = remarks

            db.session.commit()
            flash('QC check saved successfully!', 'success')
            return redirect(url_for('qc.batches'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error saving QC check: {str(e)}', 'error')

    return render_template('qc/batch_check.html', batch=batch, templates=templates)

@bp.route('/batches/<int:id>/print')
@login_required
def print_qc_report(id):
    """Print QC report for batch"""
    batch = Batch.query.get_or_404(id)
    qc_checks = BatchQualityCheck.query.filter_by(batch_id=id).order_by(BatchQualityCheck.checked_at.desc()).all()
    return render_template('qc/qc_report_print.html', batch=batch, qc_checks=qc_checks)


@bp.route('/batches/add', methods=['GET', 'POST'])
@login_required
def add_batch_qc():
    """Add new QC check for a batch"""
    if request.method == 'POST':
        try:
            batch_id = request.form.get('batch_id')
            template_id = request.form.get('template_id')
            status = request.form.get('status')
            remarks = request.form.get('remarks')

            qc_record = BatchQualityCheck(
                batch_id=batch_id,
                template_id=template_id,
                status=status,
                remarks=remarks,
                checked_by=current_user.id,
                checked_at=datetime.utcnow()
            )
            db.session.add(qc_record)

            # Update batch QC status
            batch = Batch.query.get(batch_id)
            if batch:
                batch.qc_status = status
                batch.qc_date = datetime.utcnow().date()
                batch.qc_remarks = remarks

            db.session.commit()
            flash('QC check created successfully!', 'success')
            return redirect(url_for('qc.batches'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error creating QC check: {str(e)}', 'error')

    batches = Batch.query.order_by(Batch.batch_number).all()
    templates = QualityCheckTemplate.query.filter_by(is_active=True).all()
    return render_template('qc/batch_qc_add.html', batches=batches, templates=templates)


@bp.route('/batches/<int:id>/view')
@login_required
def view_batch_qc(id):
    """View QC details for a batch"""
    batch = Batch.query.get_or_404(id)
    qc_checks = BatchQualityCheck.query.filter_by(batch_id=id).order_by(BatchQualityCheck.checked_at.desc()).all()
    return render_template('qc/batch_qc_view.html', batch=batch, qc_checks=qc_checks)


@bp.route('/batches/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def edit_batch_qc(id):
    """Edit QC record for a batch"""
    batch = Batch.query.get_or_404(id)
    qc_record = BatchQualityCheck.query.filter_by(batch_id=id).order_by(BatchQualityCheck.checked_at.desc()).first()

    if request.method == 'POST':
        try:
            if qc_record:
                qc_record.template_id = request.form.get('template_id')
                qc_record.status = request.form.get('status')
                qc_record.remarks = request.form.get('remarks')
                qc_record.checked_by = current_user.id
                qc_record.checked_at = datetime.utcnow()
            else:
                qc_record = BatchQualityCheck(
                    batch_id=batch.id,
                    template_id=request.form.get('template_id'),
                    status=request.form.get('status'),
                    remarks=request.form.get('remarks'),
                    checked_by=current_user.id,
                    checked_at=datetime.utcnow()
                )
                db.session.add(qc_record)

            # Update batch QC status
            batch.qc_status = request.form.get('status')
            batch.qc_date = datetime.utcnow().date()
            batch.qc_remarks = request.form.get('remarks')

            db.session.commit()
            flash('QC check updated successfully!', 'success')
            return redirect(url_for('qc.batches'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating QC check: {str(e)}', 'error')

    templates = QualityCheckTemplate.query.filter_by(is_active=True).all()
    return render_template('qc/batch_qc_edit.html', batch=batch, qc_record=qc_record, templates=templates)


@bp.route('/batches/<int:id>/delete', methods=['POST'])
@login_required
def delete_batch_qc(id):
    """Delete QC record for a batch"""
    try:
        batch = Batch.query.get_or_404(id)
        
        # Delete all QC checks for this batch
        BatchQualityCheck.query.filter_by(batch_id=id).delete()
        
        # Reset batch QC status
        batch.qc_status = 'pending'
        batch.qc_date = None
        batch.qc_remarks = None
        
        db.session.commit()
        flash('QC record deleted successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting QC record: {str(e)}', 'error')
    
    return redirect(url_for('qc.batches'))
