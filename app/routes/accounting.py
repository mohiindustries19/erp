"""Accounting Routes - Complete Financial Management."""

import base64
import os

from flask import Blueprint, current_app, flash, jsonify, redirect, render_template, request, url_for, send_file
from flask_login import login_required, current_user
from app import db
from app.models.accounting import ChartOfAccounts, Expense, ExpenseCategory, JournalEntry, AccountingEntry, FinancialYear
from app.services.accounting_utils import post_opening_balance, resolve_payment_account
from app.services.ledger_rebuild import rebuild_ledger
from app.services.permissions import role_required
from app.models import Order, Payment, Distributor, VendorBill, Vendor
from app.services.email_service import EmailService
from datetime import datetime, date, timedelta
from decimal import Decimal
from sqlalchemy import func, and_, or_
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO

bp = Blueprint('accounting', __name__, url_prefix='/accounting')


def _current_financial_year_start(today: date | None = None) -> date:
    today = today or date.today()
    if today.month >= 4:
        return date(today.year, 4, 1)
    return date(today.year - 1, 4, 1)


def _to_decimal(value) -> Decimal:
    if value is None:
        return Decimal('0')
    if isinstance(value, Decimal):
        return value
    return Decimal(str(value))


def _company_name() -> str:
    return current_app.config.get('COMPANY_NAME') or 'Mohi Industries'


def _logo_data_uri() -> str | None:
    """Embed the logo as data-uri for reliable PDF + email rendering."""
    try:
        logo_path = os.path.join(current_app.static_folder, 'logo.png')
        if not os.path.exists(logo_path):
            return None
        with open(logo_path, 'rb') as f:
            b64 = base64.b64encode(f.read()).decode('ascii')
        return f'data:image/png;base64,{b64}'
    except Exception:
        return None


def _report_recipient_email() -> str | None:
    email = (request.form.get('email') or '').strip()
    if email:
        return email
    fallback = (current_app.config.get('COMPANY_EMAIL') or '').strip()
    return fallback or None


@bp.route('/admin/rebuild-ledger', methods=['GET', 'POST'])
@login_required
@role_required(['admin'])
def rebuild_ledger_admin():
    """Admin tool to rebuild auto-posted ledger entries for a date range."""

    today = date.today()
    default_start = date(today.year, today.month, 1)
    default_end = today

    summary = None

    if request.method == 'POST':
        try:
            start_date_str = request.form.get('start_date')
            end_date_str = request.form.get('end_date')
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date() if start_date_str else default_start
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date() if end_date_str else default_end

            include_orders = request.form.get('include_orders') == 'on'
            include_payments = request.form.get('include_payments') == 'on'
            include_expenses = request.form.get('include_expenses') == 'on'
            include_vendor = request.form.get('include_vendor') == 'on'
            dry_run = request.form.get('dry_run') == 'on'

            if end_date < start_date:
                flash('End date must be on or after start date.', 'error')
                return redirect(url_for('accounting.rebuild_ledger_admin'))

            summary = rebuild_ledger(
                start_date=start_date,
                end_date=end_date,
                include_orders=include_orders,
                include_payments=include_payments,
                include_expenses=include_expenses,
                include_vendor=include_vendor,
                dry_run=dry_run,
                run_as_user_id=current_user.id,
            )

            if dry_run:
                flash('Dry run completed (no changes saved).', 'info')
            else:
                flash('Ledger rebuild completed successfully.', 'success')

            return render_template(
                'accounting/rebuild_ledger.html',
                start_date=start_date,
                end_date=end_date,
                include_orders=include_orders,
                include_payments=include_payments,
                include_expenses=include_expenses,
                include_vendor=include_vendor,
                dry_run=dry_run,
                summary=summary,
            )

        except Exception as e:
            db.session.rollback()
            flash(f'Error rebuilding ledger: {str(e)}', 'error')

    return render_template(
        'accounting/rebuild_ledger.html',
        start_date=default_start,
        end_date=default_end,
        include_orders=True,
        include_payments=True,
        include_expenses=True,
        include_vendor=True,
        dry_run=True,
        summary=summary,
    )

# ==================== OPENING BALANCES ====================

@bp.route('/opening-balances')
@login_required
def opening_balances():
    """View and manage opening balances"""
    accounts = ChartOfAccounts.query.filter_by(is_active=True).order_by(
        ChartOfAccounts.root_type, ChartOfAccounts.account_number
    ).all()
    
    # Group by root type
    grouped_accounts = {
        'asset': [],
        'liability': [],
        'income': [],
        'expense': [],
        'equity': []
    }
    
    for account in accounts:
        key = (account.root_type or '').strip().lower()
        if key in grouped_accounts:
            grouped_accounts[key].append(account)
    
    # Calculate totals
    total_assets = sum(acc.get_balance() for acc in grouped_accounts['asset'])
    total_liabilities = sum(acc.get_balance() for acc in grouped_accounts['liability'])
    total_equity = sum(acc.get_balance() for acc in grouped_accounts['equity'])
    
    is_balanced = abs((total_assets) - (total_liabilities + total_equity)) < 0.01
    
    return render_template('accounting/opening_balances.html', 
                         grouped_accounts=grouped_accounts,
                         total_assets=total_assets,
                         total_liabilities=total_liabilities,
                         total_equity=total_equity,
                         is_balanced=is_balanced)


@bp.route('/opening-balances/export-excel')
@login_required
def export_opening_balances_excel():
    """Export Opening Balances to Excel"""
    from flask import make_response
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from io import BytesIO
    
    accounts = ChartOfAccounts.query.filter_by(is_active=True).order_by(
        ChartOfAccounts.root_type, ChartOfAccounts.account_number
    ).all()
    
    grouped_accounts = {
        'asset': [],
        'liability': [],
        'income': [],
        'expense': [],
        'equity': []
    }
    
    for account in accounts:
        key = (account.root_type or '').strip().lower()
        if key in grouped_accounts:
            grouped_accounts[key].append(account)
    
    total_assets = sum(acc.get_balance() for acc in grouped_accounts['asset'])
    total_liabilities = sum(acc.get_balance() for acc in grouped_accounts['liability'])
    total_equity = sum(acc.get_balance() for acc in grouped_accounts['equity'])
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Opening Balances"
    
    header_fill = PatternFill(start_color="D00000", end_color="D00000", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=16, color="D00000")
    bold_font = Font(bold=True)
    
    ws.merge_cells('A1:D1')
    ws['A1'] = 'MOHI INDUSTRIES - OPENING BALANCES'
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')
    
    row = 3
    for type_name, type_accounts in grouped_accounts.items():
        if not type_accounts:
            continue
        
        ws[f'A{row}'] = type_name.upper()
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        ws.cell(row=row, column=1).value = 'Account Code'
        ws.cell(row=row, column=2).value = 'Account Name'
        ws.cell(row=row, column=3).value = 'Type'
        ws.cell(row=row, column=4).value = 'Balance'
        for col in range(1, 5):
            ws.cell(row=row, column=col).font = bold_font
        
        row += 1
        for account in type_accounts:
            ws.cell(row=row, column=1).value = account.account_number
            ws.cell(row=row, column=2).value = account.name
            ws.cell(row=row, column=3).value = account.account_type or ''
            ws.cell(row=row, column=4).value = account.get_balance()
            row += 1
        
        row += 1
    
    row += 1
    ws.cell(row=row, column=2).value = 'Total Assets:'
    ws.cell(row=row, column=2).font = bold_font
    ws.cell(row=row, column=4).value = total_assets
    ws.cell(row=row, column=4).font = Font(bold=True, color="D00000")
    
    row += 1
    ws.cell(row=row, column=2).value = 'Total Liabilities:'
    ws.cell(row=row, column=2).font = bold_font
    ws.cell(row=row, column=4).value = total_liabilities
    ws.cell(row=row, column=4).font = Font(bold=True, color="D00000")
    
    row += 1
    ws.cell(row=row, column=2).value = 'Total Equity:'
    ws.cell(row=row, column=2).font = bold_font
    ws.cell(row=row, column=4).value = total_equity
    ws.cell(row=row, column=4).font = Font(bold=True, color="D00000")
    
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = make_response(output.read())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = 'attachment; filename=Opening_Balances.xlsx'
    
    return response


@bp.route('/opening-balances/print')
@login_required
def print_opening_balances():
    """Print Opening Balances"""
    accounts = ChartOfAccounts.query.filter_by(is_active=True).order_by(
        ChartOfAccounts.root_type, ChartOfAccounts.account_number
    ).all()
    
    grouped_accounts = {
        'asset': [],
        'liability': [],
        'income': [],
        'expense': [],
        'equity': []
    }
    
    for account in accounts:
        key = (account.root_type or '').strip().lower()
        if key in grouped_accounts:
            grouped_accounts[key].append(account)
    
    total_assets = sum(acc.get_balance() for acc in grouped_accounts['asset'])
    total_liabilities = sum(acc.get_balance() for acc in grouped_accounts['liability'])
    total_equity = sum(acc.get_balance() for acc in grouped_accounts['equity'])
    is_balanced = abs((total_assets) - (total_liabilities + total_equity)) < 0.01
    
    return render_template(
        'accounting/opening_balances_print.html',
        grouped_accounts=grouped_accounts,
        total_assets=total_assets,
        total_liabilities=total_liabilities,
        total_equity=total_equity,
        is_balanced=is_balanced,
        company_name=_company_name(),
        logo_data_uri=_logo_data_uri(),
    )



@bp.route('/opening-balances/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def edit_opening_balance(id):
    """Edit opening balance for an account - uses journal entries"""
    account = ChartOfAccounts.query.get_or_404(id)
    
    if request.method == 'POST':
        try:
            opening_balance = float(request.form.get('opening_balance', 0) or 0)
            as_of = _current_financial_year_start()
            post_opening_balance(
                account=account,
                amount_natural=Decimal(str(opening_balance)),
                as_of_date=as_of,
                created_by=current_user.id,
            )
            db.session.commit()
            flash('Opening balance updated successfully.', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating opening balance: {str(e)}', 'error')
        return redirect(url_for('accounting.opening_balances'))
    
    return render_template('accounting/edit_opening_balance.html', account=account)


@bp.route('/opening-balances/bulk-update', methods=['POST'])
@login_required
def bulk_update_opening_balances():
    """Bulk update opening balances - uses journal entries"""
    try:
        as_of = _current_financial_year_start()
        accounts = ChartOfAccounts.query.filter_by(is_active=True).all()
        updated = 0
        for account in accounts:
            # Backward compatible with older templates posting balance_{id}
            key_new = f'opening_balance_{account.id}'
            key_old = f'balance_{account.id}'
            raw = request.form.get(key_new)
            if raw is None:
                raw = request.form.get(key_old)
            if raw is None or str(raw).strip() == '':
                continue
            amount = Decimal(str(raw))
            post_opening_balance(
                account=account,
                amount_natural=amount,
                as_of_date=as_of,
                created_by=current_user.id,
            )
            updated += 1
        db.session.commit()
        flash(f'Opening balances updated for {updated} account(s).', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error updating opening balances: {str(e)}', 'error')
    return redirect(url_for('accounting.opening_balances'))


# ==================== RECEIPTS (PAYMENTS RECEIVED) ====================

@bp.route('/receipts')
@login_required
def list_receipts():
    """List all payments received from customers"""
    # Get filter parameters
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    status = request.args.get('status')
    
    query = Payment.query
    
    if start_date:
        query = query.filter(Payment.payment_date >= datetime.strptime(start_date, '%Y-%m-%d').date())
    if end_date:
        query = query.filter(Payment.payment_date <= datetime.strptime(end_date, '%Y-%m-%d').date())
    if status:
        query = query.filter(Payment.status == status)
    
    receipts = query.order_by(Payment.payment_date.desc()).all()
    
    # Calculate totals
    total_cleared = sum(r.amount for r in receipts if r.status == 'cleared')
    total_pending = sum(r.amount for r in receipts if r.status == 'pending')
    total_bounced = sum(r.amount for r in receipts if r.status == 'bounced')
    
    return render_template('accounting/receipts.html',
                         receipts=receipts,
                         total_cleared=total_cleared,
                         total_pending=total_pending,
                         total_bounced=total_bounced)


@bp.route('/receipts/<int:id>')
@login_required
def view_receipt(id):
    """View receipt details"""
    receipt = Payment.query.get_or_404(id)
    
    # Get accounting entries for this receipt
    entries = AccountingEntry.query.filter_by(
        reference_type='payment',
        reference_id=receipt.id
    ).all()
    
    return render_template('accounting/view_receipt.html', receipt=receipt, entries=entries)


# ==================== SALES ENTRIES ====================

@bp.route('/sales')
@login_required
def list_sales():
    """List all sales/orders with accounting view"""
    # Get filter parameters
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    query = Order.query
    
    if start_date:
        query = query.filter(Order.order_date >= datetime.strptime(start_date, '%Y-%m-%d').date())
    if end_date:
        query = query.filter(Order.order_date <= datetime.strptime(end_date, '%Y-%m-%d').date())
    
    orders = query.order_by(Order.order_date.desc()).all()
    
    # Calculate totals
    total_sales = sum(o.total_amount for o in orders)
    total_received = sum(o.paid_amount for o in orders)
    total_outstanding = total_sales - total_received
    
    return render_template('accounting/sales.html',
                         orders=orders,
                         total_sales=total_sales,
                         total_received=total_received,
                         total_outstanding=total_outstanding)


# ==================== GENERAL LEDGER ====================

@bp.route('/chart-of-accounts')
@login_required
def chart_of_accounts():
    """View General Ledger (Chart of Accounts)"""
    accounts = ChartOfAccounts.query.filter_by(is_active=True).order_by(
        ChartOfAccounts.root_type, ChartOfAccounts.account_number
    ).all()
    
    # Group by root type
    grouped_accounts = {
        'asset': [],
        'liability': [],
        'income': [],
        'expense': [],
        'equity': []
    }
    
    for account in accounts:
        key = (account.root_type or '').strip().lower()
        if key in grouped_accounts:
            grouped_accounts[key].append(account)
    
    return render_template('accounting/chart_of_accounts.html', grouped_accounts=grouped_accounts)


@bp.route('/chart-of-accounts/add', methods=['GET', 'POST'])
@login_required
def add_account():
    """Add new account"""
    if request.method == 'POST':
        try:
            # Backward compatible field names
            account_number = request.form.get('account_number') or request.form.get('account_code')
            name = request.form.get('name') or request.form.get('account_name')
            root_type = request.form.get('root_type') or request.form.get('account_type')
            description = request.form.get('description') or request.form.get('account_group')
            account_type = request.form.get('account_type_detail') or request.form.get('account_subtype')

            root_map = {
                'asset': 'Asset',
                'liability': 'Liability',
                'equity': 'Equity',
                'income': 'Income',
                'expense': 'Expense',
            }
            root_type = root_map.get((root_type or '').lower(), root_type)

            account = ChartOfAccounts(
                account_number=account_number,
                name=name,
                root_type=root_type,
                account_type=account_type,
                description=description,
            )

            # Optional legacy fields if model/table supports them
            if hasattr(account, 'gst_rate') and request.form.get('gst_rate'):
                account.gst_rate = float(request.form.get('gst_rate'))
            if hasattr(account, 'is_gst_applicable'):
                account.is_gst_applicable = request.form.get('is_gst_applicable') == 'on'
            
            db.session.add(account)
            db.session.commit()
            
            flash(f'Account {account.name} created successfully!', 'success')
            return redirect(url_for('accounting.chart_of_accounts'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error creating account: {str(e)}', 'error')
    
    return render_template('accounting/add_account.html')


@bp.route('/chart-of-accounts/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def edit_account(id):
    """Edit account"""
    account = ChartOfAccounts.query.get_or_404(id)
    
    if request.method == 'POST':
        try:
            name = request.form.get('name') or request.form.get('account_name')
            root_type = request.form.get('root_type') or request.form.get('account_type')
            description = request.form.get('description') or request.form.get('account_group')
            account_type = request.form.get('account_type_detail') or request.form.get('account_subtype')

            root_map = {
                'asset': 'Asset',
                'liability': 'Liability',
                'equity': 'Equity',
                'income': 'Income',
                'expense': 'Expense',
            }
            root_type = root_map.get((root_type or '').lower(), root_type)

            if name:
                account.name = name
            if root_type:
                account.root_type = root_type
            account.account_type = account_type
            account.description = description

            if hasattr(account, 'gst_rate') and request.form.get('gst_rate'):
                account.gst_rate = float(request.form.get('gst_rate'))
            if hasattr(account, 'is_gst_applicable'):
                account.is_gst_applicable = request.form.get('is_gst_applicable') == 'on'
            
            db.session.commit()
            flash(f'Account {account.name} updated successfully!', 'success')
            return redirect(url_for('accounting.chart_of_accounts'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating account: {str(e)}', 'error')
    
    return render_template('accounting/edit_account.html', account=account)


@bp.route('/chart-of-accounts/<int:id>/delete', methods=['POST'])
@login_required
def delete_account(id):
    """Delete account"""
    account = ChartOfAccounts.query.get_or_404(id)
    
    try:
        # Check if account has transactions
        if account.journal_entries.count() > 0:
            flash(f'Cannot delete account {account.name} - it has transactions!', 'error')
            return redirect(url_for('accounting.chart_of_accounts'))
        
        account_name = account.name
        account.is_active = False  # Soft delete
        db.session.commit()
        flash(f'Account {account_name} deactivated successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting account: {str(e)}', 'error')
    
    return redirect(url_for('accounting.chart_of_accounts'))


# ==================== EXPENSE MANAGEMENT ====================

def _suggest_expense_category_from_account_name(name: str) -> str:
    if not name:
        return 'General'
    n = str(name).strip().lower()
    if not n:
        return 'General'

    if 'rent' in n:
        return 'Rent'
    if 'salary' in n or 'wages' in n:
        return 'Salary'
    if 'repair' in n or 'maintenance' in n:
        return 'Repairs & Maintenance'
    if 'electric' in n or 'power' in n:
        return 'Electricity'
    if 'fuel' in n or 'diesel' in n or 'petrol' in n:
        return 'Fuel'
    if 'travel' in n or 'convey' in n:
        return 'Travel'
    if 'office' in n or 'stationery' in n:
        return 'Office'
    if 'misc' in n:
        return 'Miscellaneous'

    # Fallback: if the account name ends with “expense(s)”, use the prefix as category.
    cleaned = str(name).strip()
    for suffix in (' expenses', ' expense'):
        if cleaned.lower().endswith(suffix):
            cleaned = cleaned[: -len(suffix)].strip()
            break
    return cleaned or 'General'


def _normalize_expense_category_name(name: str | None) -> str:
    if not name:
        return ''
    return ' '.join(str(name).strip().split())


def _get_expense_category_by_name(name: str) -> ExpenseCategory | None:
    cleaned = _normalize_expense_category_name(name)
    if not cleaned:
        return None
    return ExpenseCategory.query.filter(func.lower(ExpenseCategory.name) == cleaned.lower()).first()


def _ensure_expense_category(name: str, *, allow_create: bool) -> ExpenseCategory | None:
    cleaned = _normalize_expense_category_name(name)
    if not cleaned:
        return None
    existing = _get_expense_category_by_name(cleaned)
    if existing:
        return existing
    if not allow_create:
        return None

    cat = ExpenseCategory(name=cleaned, is_active=True)
    db.session.add(cat)
    db.session.flush()
    return cat

@bp.route('/expenses')
@login_required
def list_expenses():
    """List all expenses"""
    # Get filter parameters
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    category = request.args.get('category')
    
    query = Expense.query
    
    if start_date:
        query = query.filter(Expense.expense_date >= datetime.strptime(start_date, '%Y-%m-%d').date())
    if end_date:
        query = query.filter(Expense.expense_date <= datetime.strptime(end_date, '%Y-%m-%d').date())
    if category:
        query = query.filter(Expense.expense_category == category)
    
    expenses = query.order_by(Expense.expense_date.desc()).all()
    
    # Calculate totals
    total_amount = sum(e.amount for e in expenses)
    total_gst = sum(e.total_gst for e in expenses)
    total_with_gst = sum(e.total_amount for e in expenses)
    
    # Controlled categories for filter
    categories = ExpenseCategory.query.filter_by(is_active=True).order_by(ExpenseCategory.name).all()
    
    return render_template('accounting/expenses.html', 
                         expenses=expenses,
                         total_amount=total_amount,
                         total_gst=total_gst,
                         total_with_gst=total_with_gst,
                         categories=categories)


@bp.route('/expenses/add', methods=['GET', 'POST'])
@login_required
def add_expense():
    """Add new expense"""
    if request.method == 'POST':
        try:
            # Generate expense number
            today = date.today()
            expense_prefix = f'EXP{today.strftime("%Y%m%d")}'
            last_expense = Expense.query.filter(
                Expense.expense_number.like(f'{expense_prefix}%')
            ).order_by(Expense.id.desc()).first()
            
            if last_expense:
                last_num = int(last_expense.expense_number[-4:])
                expense_number = f'{expense_prefix}{last_num+1:04d}'
            else:
                expense_number = f'{expense_prefix}0001'
            
            # Resolve & validate category (controlled dropdown)
            raw_category = _normalize_expense_category_name(request.form.get('expense_category'))

            # Create expense
            expense = Expense(
                expense_number=expense_number,
                expense_date=datetime.strptime(request.form.get('expense_date'), '%Y-%m-%d').date(),
                account_id=request.form.get('account_id'),
                amount=float(request.form.get('amount')),
                is_gst_applicable=request.form.get('is_gst_applicable') == 'on',
                vendor_name=request.form.get('vendor_name'),
                vendor_gstin=request.form.get('vendor_gstin'),
                invoice_number=request.form.get('invoice_number'),
                payment_mode=request.form.get('payment_mode'),
                payment_status=request.form.get('payment_status', 'paid'),
                reference_number=request.form.get('reference_number'),
                expense_category=raw_category or None,
                description=request.form.get('description'),
                remarks=request.form.get('remarks'),
                recorded_by=current_user.id,
                approval_status='approved'  # Auto-approve for now
            )
            
            # Handle invoice date
            if request.form.get('invoice_date'):
                expense.invoice_date = datetime.strptime(request.form.get('invoice_date'), '%Y-%m-%d').date()
            
            # Calculate GST BEFORE adding to session
            # We need to manually load the account since the relationship isn't established yet
            account = ChartOfAccounts.query.get(expense.account_id)

            # Enforce Option A (server-side): auto-generate category if blank/General
            submitted_category = _normalize_expense_category_name(expense.expense_category)
            if (not submitted_category) or (submitted_category.lower() == 'general'):
                suggested = _suggest_expense_category_from_account_name(getattr(account, 'name', ''))
                expense.expense_category = suggested
                _ensure_expense_category(suggested, allow_create=True)
            else:
                # User-supplied category must exist in master list.
                if not _get_expense_category_by_name(submitted_category):
                    suggested = _suggest_expense_category_from_account_name(getattr(account, 'name', ''))
                    flash(
                        f"Category '{submitted_category}' is not in master list. Using '{suggested}' instead.",
                        'warning',
                    )
                    expense.expense_category = suggested
                    _ensure_expense_category(suggested, allow_create=True)
            
            # Manually calculate total_amount
            gst_rate = float(getattr(account, 'gst_rate', 0) or 0)
            if expense.is_gst_applicable and account and gst_rate > 0:
                expense.total_gst = round(expense.amount * gst_rate / 100, 2)
                
                # Check if inter-state or intra-state based on GSTIN
                if expense.vendor_gstin and len(expense.vendor_gstin) >= 2:
                    vendor_state_code = expense.vendor_gstin[:2]
                    company_state_code = '27'  # Maharashtra
                    
                    if vendor_state_code == company_state_code:
                        # Intra-state: CGST + SGST
                        expense.cgst_amount = round(expense.total_gst / 2, 2)
                        expense.sgst_amount = round(expense.total_gst / 2, 2)
                        expense.igst_amount = 0.0
                    else:
                        # Inter-state: IGST
                        expense.igst_amount = expense.total_gst
                        expense.cgst_amount = 0.0
                        expense.sgst_amount = 0.0
                else:
                    # Default to CGST + SGST
                    expense.cgst_amount = round(expense.total_gst / 2, 2)
                    expense.sgst_amount = round(expense.total_gst / 2, 2)
                    expense.igst_amount = 0.0
            else:
                expense.total_gst = 0.0
                expense.cgst_amount = 0.0
                expense.sgst_amount = 0.0
                expense.igst_amount = 0.0
            
            # Always set total_amount
            expense.total_amount = expense.amount + expense.total_gst
            
            # Now add to session
            db.session.add(expense)
            db.session.flush()
            
            # Create accounting entries
            # Debit: Expense Account
            debit_entry = AccountingEntry(
                entry_date=expense.expense_date,
                reference_type='expense',
                reference_id=expense.id,
                account_id=account.id,
                account_head=account.name,
                debit=expense.total_amount,
                credit=0.0,
                description=f'Expense: {expense.description or expense.vendor_name}',
                created_by=current_user.id
            )
            
            # Credit: Cash/Bank Account
            payment_account = resolve_payment_account(expense.payment_mode)
            credit_entry = AccountingEntry(
                entry_date=expense.expense_date,
                reference_type='expense',
                reference_id=expense.id,
                account_id=payment_account.id,
                account_head=payment_account.name,
                debit=0.0,
                credit=expense.total_amount,
                description=f'Payment for: {expense.description or expense.vendor_name}',
                created_by=current_user.id
            )
            
            db.session.add(debit_entry)
            db.session.add(credit_entry)
            
            # Balance is calculated from journal entries automatically
            
            db.session.commit()
            
            flash(f'Expense {expense_number} recorded successfully!', 'success')
            return redirect(url_for('accounting.list_expenses'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error recording expense: {str(e)}', 'error')
    
    # GET request
    accounts = ChartOfAccounts.query.filter_by(
        root_type='Expense',
        is_active=True
    ).order_by(ChartOfAccounts.name).all()

    expense_categories = ExpenseCategory.query.filter_by(is_active=True).order_by(ExpenseCategory.name).all()

    account_category_map = {
        str(a.id): _suggest_expense_category_from_account_name(a.name)
        for a in accounts
    }
    
    today = date.today().strftime('%Y-%m-%d')
    
    return render_template(
        'accounting/add_expense.html',
        accounts=accounts,
        today=today,
        account_category_map=account_category_map,
        expense_categories=expense_categories,
    )


@bp.route('/expenses/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def edit_expense(id):
    """Edit expense"""
    expense = Expense.query.get_or_404(id)
    
    if request.method == 'POST':
        try:
            old_amount = expense.total_amount
            old_account_id = expense.account_id
            
            expense.expense_date = datetime.strptime(request.form.get('expense_date'), '%Y-%m-%d').date()
            expense.account_id = request.form.get('account_id')
            expense.amount = float(request.form.get('amount'))
            expense.is_gst_applicable = request.form.get('is_gst_applicable') == 'on'
            expense.vendor_name = request.form.get('vendor_name')
            expense.vendor_gstin = request.form.get('vendor_gstin')
            expense.invoice_number = request.form.get('invoice_number')
            expense.payment_mode = request.form.get('payment_mode')
            expense.payment_status = request.form.get('payment_status')
            expense.reference_number = request.form.get('reference_number')
            expense.expense_category = _normalize_expense_category_name(request.form.get('expense_category')) or None
            expense.description = request.form.get('description')
            expense.remarks = request.form.get('remarks')
            
            if request.form.get('invoice_date'):
                expense.invoice_date = datetime.strptime(request.form.get('invoice_date'), '%Y-%m-%d').date()
            
            # Recalculate GST
            account = ChartOfAccounts.query.get(expense.account_id)

            # Enforce Option A (server-side): auto-generate category if blank/General
            submitted_category = _normalize_expense_category_name(expense.expense_category)
            if (not submitted_category) or (submitted_category.lower() == 'general'):
                suggested = _suggest_expense_category_from_account_name(getattr(account, 'name', ''))
                expense.expense_category = suggested
                _ensure_expense_category(suggested, allow_create=True)
            else:
                # User-supplied category must exist in master list.
                if not _get_expense_category_by_name(submitted_category):
                    suggested = _suggest_expense_category_from_account_name(getattr(account, 'name', ''))
                    flash(
                        f"Category '{submitted_category}' is not in master list. Using '{suggested}' instead.",
                        'warning',
                    )
                    expense.expense_category = suggested
                    _ensure_expense_category(suggested, allow_create=True)
            
            gst_rate = float(getattr(account, 'gst_rate', 0) or 0)
            if expense.is_gst_applicable and account and gst_rate > 0:
                expense.total_gst = round(expense.amount * gst_rate / 100, 2)
                
                if expense.vendor_gstin and len(expense.vendor_gstin) >= 2:
                    vendor_state_code = expense.vendor_gstin[:2]
                    company_state_code = '27'
                    
                    if vendor_state_code == company_state_code:
                        expense.cgst_amount = round(expense.total_gst / 2, 2)
                        expense.sgst_amount = round(expense.total_gst / 2, 2)
                        expense.igst_amount = 0.0
                    else:
                        expense.igst_amount = expense.total_gst
                        expense.cgst_amount = 0.0
                        expense.sgst_amount = 0.0
                else:
                    expense.cgst_amount = round(expense.total_gst / 2, 2)
                    expense.sgst_amount = round(expense.total_gst / 2, 2)
                    expense.igst_amount = 0.0
            else:
                expense.total_gst = 0.0
                expense.cgst_amount = 0.0
                expense.sgst_amount = 0.0
                expense.igst_amount = 0.0
            
            expense.total_amount = expense.amount + expense.total_gst
            
            # Update accounting entries
            entries = AccountingEntry.query.filter_by(
                reference_type='expense',
                reference_id=expense.id
            ).all()
            
            for entry in entries:
                if entry.debit > 0:  # Expense account entry
                    entry.account_id = account.id
                    entry.account_head = account.name
                    entry.debit = expense.total_amount
                else:  # Payment account entry
                    payment_account = resolve_payment_account(expense.payment_mode)
                    entry.account_id = payment_account.id
                    entry.account_head = payment_account.name
                    entry.credit = expense.total_amount
            
            # Balance is calculated from journal entries automatically
            
            db.session.commit()
            flash(f'Expense {expense.expense_number} updated successfully!', 'success')
            return redirect(url_for('accounting.list_expenses'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating expense: {str(e)}', 'error')
    
    accounts = ChartOfAccounts.query.filter_by(
        root_type='Expense',
        is_active=True
    ).order_by(ChartOfAccounts.name).all()

    expense_categories = ExpenseCategory.query.filter_by(is_active=True).order_by(ExpenseCategory.name).all()

    account_category_map = {
        str(a.id): _suggest_expense_category_from_account_name(a.name)
        for a in accounts
    }

    return render_template(
        'accounting/edit_expense.html',
        expense=expense,
        accounts=accounts,
        account_category_map=account_category_map,
        expense_categories=expense_categories,
    )


@bp.route('/expenses/categories')
@login_required
@role_required(['admin'])
def expense_categories():
    """Manage master list of expense categories."""
    categories = ExpenseCategory.query.order_by(ExpenseCategory.name).all()
    usage = {
        name: count
        for name, count in db.session.query(
            Expense.expense_category,
            func.count(Expense.id),
        ).group_by(Expense.expense_category)
        .all()
        if name
    }

    return render_template(
        'accounting/expense_categories.html',
        categories=categories,
        usage=usage,
    )


@bp.route('/expenses/categories/add', methods=['POST'])
@login_required
@role_required(['admin'])
def add_expense_category():
    name = _normalize_expense_category_name(request.form.get('name'))
    if not name:
        flash('Category name is required.', 'error')
        return redirect(url_for('accounting.expense_categories'))

    existing = _get_expense_category_by_name(name)
    if existing:
        flash('Category already exists.', 'info')
        return redirect(url_for('accounting.expense_categories'))

    db.session.add(ExpenseCategory(name=name, is_active=True))
    db.session.commit()
    flash('Category added.', 'success')
    return redirect(url_for('accounting.expense_categories'))


@bp.route('/expenses/categories/<int:id>/toggle', methods=['POST'])
@login_required
@role_required(['admin'])
def toggle_expense_category(id: int):
    category = ExpenseCategory.query.get_or_404(id)
    category.is_active = not bool(category.is_active)
    db.session.commit()
    flash('Category updated.', 'success')
    return redirect(url_for('accounting.expense_categories'))


@bp.route('/expenses/<int:id>/delete', methods=['POST'])
@login_required
def delete_expense(id):
    """Delete expense"""
    expense = Expense.query.get_or_404(id)
    
    try:
        expense_number = expense.expense_number
        
        # Delete accounting entries
        AccountingEntry.query.filter_by(
            reference_type='expense',
            reference_id=expense.id
        ).delete()
        
        # Balance is calculated from journal entries automatically
        
        db.session.delete(expense)
        db.session.commit()
        
        flash(f'Expense {expense_number} deleted successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting expense: {str(e)}', 'error')
    
    return redirect(url_for('accounting.list_expenses'))


@bp.route('/expenses/<int:id>')
@login_required
def view_expense(id):
    """View expense details"""
    expense = Expense.query.get_or_404(id)
    return render_template('accounting/view_expense.html', expense=expense)


@bp.route('/expenses/export-excel')
@login_required
def export_expenses_excel():
    """Export expenses to Excel"""
    # Get filter parameters
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    category = request.args.get('category')
    
    query = Expense.query
    
    if start_date:
        query = query.filter(Expense.expense_date >= datetime.strptime(start_date, '%Y-%m-%d').date())
    if end_date:
        query = query.filter(Expense.expense_date <= datetime.strptime(end_date, '%Y-%m-%d').date())
    if category:
        query = query.filter(Expense.expense_category == category)
    
    expenses = query.order_by(Expense.expense_date.desc()).all()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Expenses"
    
    # Styles
    header_fill = PatternFill(start_color="D00000", end_color="D00000", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=16, color="D00000")
    bold_font = Font(bold=True)
    
    # Title
    ws.merge_cells('A1:M1')
    ws['A1'] = 'MOHI INDUSTRIES - EXPENSES REPORT'
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Filter info
    row = 2
    if start_date or end_date or category:
        filter_text = 'Filters: '
        if start_date:
            filter_text += f'From {start_date} '
        if end_date:
            filter_text += f'To {end_date} '
        if category:
            filter_text += f'Category: {category}'
        ws.merge_cells(f'A{row}:M{row}')
        ws[f'A{row}'] = filter_text
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        row += 1
    
    # Headers
    row += 1
    headers = ['Expense #', 'Date', 'Vendor', 'Category', 'Account', 'Description', 'Remarks', 'Amount', 'CGST', 'SGST', 'IGST', 'Total GST', 'Total Amount']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    row += 1
    total_amount = 0
    total_gst = 0
    total_with_gst = 0
    
    for expense in expenses:
        ws.cell(row=row, column=1).value = expense.expense_number
        ws.cell(row=row, column=2).value = expense.expense_date.strftime('%d-%m-%Y')
        ws.cell(row=row, column=3).value = expense.vendor_name or '-'
        ws.cell(row=row, column=4).value = expense.expense_category or 'General'
        ws.cell(row=row, column=5).value = expense.account.name
        ws.cell(row=row, column=6).value = (expense.description or '').strip() or '-'
        ws.cell(row=row, column=7).value = (expense.remarks or '').strip() or '-'
        ws.cell(row=row, column=8).value = float(expense.amount)
        ws.cell(row=row, column=9).value = float(expense.cgst_amount or 0)
        ws.cell(row=row, column=10).value = float(expense.sgst_amount or 0)
        ws.cell(row=row, column=11).value = float(expense.igst_amount or 0)
        ws.cell(row=row, column=12).value = float(expense.total_gst)
        ws.cell(row=row, column=13).value = float(expense.total_amount)
        
        total_amount += expense.amount
        total_gst += expense.total_gst
        total_with_gst += expense.total_amount
        
        row += 1
    
    # Totals
    row += 1
    ws.cell(row=row, column=5).value = 'TOTALS:'
    ws.cell(row=row, column=5).font = bold_font
    ws.cell(row=row, column=8).value = float(total_amount)
    ws.cell(row=row, column=8).font = Font(bold=True, color="D00000")
    ws.cell(row=row, column=12).value = float(total_gst)
    ws.cell(row=row, column=12).font = Font(bold=True, color="D00000")
    ws.cell(row=row, column=13).value = float(total_with_gst)
    ws.cell(row=row, column=13).font = Font(bold=True, color="D00000")
    
    # Column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 35
    ws.column_dimensions['G'].width = 30
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 10
    ws.column_dimensions['J'].width = 10
    ws.column_dimensions['K'].width = 10
    ws.column_dimensions['L'].width = 12
    ws.column_dimensions['M'].width = 15
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate filename with timestamp
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'Expenses_Report_{timestamp}.xlsx'
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@bp.route('/expenses/print-report')
@login_required
def print_expenses_report():
    """Print expenses report"""
    # Get filter parameters
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    category = request.args.get('category')
    
    query = Expense.query
    
    if start_date:
        query = query.filter(Expense.expense_date >= datetime.strptime(start_date, '%Y-%m-%d').date())
    if end_date:
        query = query.filter(Expense.expense_date <= datetime.strptime(end_date, '%Y-%m-%d').date())
    if category:
        query = query.filter(Expense.expense_category == category)
    
    expenses = query.order_by(Expense.expense_date.desc()).all()
    
    # Calculate totals
    total_amount = sum(e.amount for e in expenses)
    total_gst = sum(e.total_gst for e in expenses)
    total_with_gst = sum(e.total_amount for e in expenses)
    
    return render_template(
        'accounting/expenses_report_print.html',
        expenses=expenses,
        total_amount=total_amount,
        total_gst=total_gst,
        total_with_gst=total_with_gst,
        start_date=start_date,
        end_date=end_date,
        category=category,
        company_name=_company_name(),
        logo_data_uri=_logo_data_uri(),
    )


@bp.route('/expenses/<int:id>/print')
@login_required
def print_expense_voucher(id):
    """Print individual expense voucher"""
    expense = Expense.query.get_or_404(id)
    
    return render_template(
        'accounting/expense_voucher_print.html',
        expense=expense,
        company_name=_company_name(),
        logo_data_uri=_logo_data_uri(),
    )


# ==================== LEDGER ====================

@bp.route('/ledger')
@login_required
def ledger():
    """View ledger for all accounts"""
    accounts = ChartOfAccounts.query.filter_by(is_active=True).order_by(
        ChartOfAccounts.name
    ).all()
    
    return render_template('accounting/ledger.html', accounts=accounts)


@bp.route('/ledger/<int:account_id>')
@login_required
def account_ledger(account_id):
    """View ledger for specific account"""
    account = ChartOfAccounts.query.get_or_404(account_id)

    def to_decimal(value) -> Decimal:
        if value is None:
            return Decimal('0')
        if isinstance(value, Decimal):
            return value
        return Decimal(str(value))
    
    # Get date range
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    query = AccountingEntry.query.filter_by(account_id=account_id)
    
    if start_date:
        query = query.filter(AccountingEntry.entry_date >= datetime.strptime(start_date, '%Y-%m-%d').date())
    if end_date:
        query = query.filter(AccountingEntry.entry_date <= datetime.strptime(end_date, '%Y-%m-%d').date())
    
    # Opening balance is the sum of all entries before the range start
    opening_balance = Decimal('0')
    if start_date:
        start_dt = datetime.strptime(start_date, '%Y-%m-%d').date()
        opening_balance = to_decimal(db.session.query(
            func.sum(AccountingEntry.debit - AccountingEntry.credit)
        ).filter(
            AccountingEntry.account_id == account_id,
            AccountingEntry.entry_date < start_dt,
        ).scalar())

    entries = query.order_by(AccountingEntry.entry_date, AccountingEntry.id).all()

    # Calculate running balance
    running_balance = to_decimal(opening_balance)
    entries_with_balance = []
    
    for entry in entries:
        running_balance += to_decimal(entry.debit) - to_decimal(entry.credit)
        entries_with_balance.append({
            'entry': entry,
            'balance': float(running_balance)
        })
    
    return render_template('accounting/account_ledger.html', 
                         account=account,
                         entries_with_balance=entries_with_balance,
                         opening_balance=float(opening_balance),
                         closing_balance=float(running_balance))


@bp.route('/ledger/<int:account_id>/export-excel')
@login_required
def export_ledger_excel(account_id):
    """Export ledger to Excel"""
    from flask import make_response
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO
    
    account = ChartOfAccounts.query.get_or_404(account_id)
    
    def to_decimal(value) -> Decimal:
        if value is None:
            return Decimal('0')
        if isinstance(value, Decimal):
            return value
        return Decimal(str(value))
    
    # Get date range
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    query = AccountingEntry.query.filter_by(account_id=account_id)
    
    if start_date:
        query = query.filter(AccountingEntry.entry_date >= datetime.strptime(start_date, '%Y-%m-%d').date())
    if end_date:
        query = query.filter(AccountingEntry.entry_date <= datetime.strptime(end_date, '%Y-%m-%d').date())
    
    # Opening balance
    opening_balance = Decimal('0')
    if start_date:
        start_dt = datetime.strptime(start_date, '%Y-%m-%d').date()
        opening_balance = to_decimal(db.session.query(
            func.sum(AccountingEntry.debit - AccountingEntry.credit)
        ).filter(
            AccountingEntry.account_id == account_id,
            AccountingEntry.entry_date < start_dt,
        ).scalar())

    entries = query.order_by(AccountingEntry.entry_date, AccountingEntry.id).all()
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ledger"
    
    # Styles
    header_fill = PatternFill(start_color="D00000", end_color="D00000", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=16, color="D00000")
    bold_font = Font(bold=True)
    
    # Title
    ws.merge_cells('A1:G1')
    ws['A1'] = 'MOHI INDUSTRIES - LEDGER'
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Account details
    ws.merge_cells('A2:G2')
    ws['A2'] = f'Account: {account.name} ({account.account_number})'
    ws['A2'].font = bold_font
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Date range
    row = 3
    if start_date and end_date:
        ws.merge_cells(f'A{row}:G{row}')
        ws[f'A{row}'] = f'Period: {start_date} to {end_date}'
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        row += 1
    
    # Opening balance
    row += 1
    ws[f'A{row}'] = 'Opening Balance:'
    ws[f'A{row}'].font = bold_font
    ws[f'G{row}'] = float(opening_balance)
    ws[f'G{row}'].font = bold_font
    
    # Headers
    row += 2
    headers = ['Date', 'Reference', 'Description', 'Debit', 'Credit', 'Balance']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    running_balance = float(opening_balance)
    row += 1
    for entry in entries:
        running_balance += float(entry.debit) - float(entry.credit)
        
        ws.cell(row=row, column=1).value = entry.entry_date.strftime('%d-%m-%Y')
        ws.cell(row=row, column=2).value = f"{entry.reference_type}-{entry.reference_id}"
        ws.cell(row=row, column=3).value = entry.description
        ws.cell(row=row, column=4).value = float(entry.debit) if entry.debit > 0 else ''
        ws.cell(row=row, column=5).value = float(entry.credit) if entry.credit > 0 else ''
        ws.cell(row=row, column=6).value = running_balance
        
        row += 1
    
    # Closing balance
    row += 1
    ws.cell(row=row, column=1).value = 'Closing Balance:'
    ws.cell(row=row, column=1).font = bold_font
    ws.cell(row=row, column=6).value = running_balance
    ws.cell(row=row, column=6).font = Font(bold=True, color="D00000")
    
    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    
    # Save
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = make_response(output.read())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=Ledger_{account.name.replace(" ", "_")}.xlsx'
    
    return response


@bp.route('/ledger/<int:account_id>/print')
@login_required
def print_ledger(account_id):
    """Print ledger"""
    account = ChartOfAccounts.query.get_or_404(account_id)
    
    def to_decimal(value) -> Decimal:
        if value is None:
            return Decimal('0')
        if isinstance(value, Decimal):
            return value
        return Decimal(str(value))
    
    # Get date range
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    query = AccountingEntry.query.filter_by(account_id=account_id)
    
    if start_date:
        query = query.filter(AccountingEntry.entry_date >= datetime.strptime(start_date, '%Y-%m-%d').date())
    if end_date:
        query = query.filter(AccountingEntry.entry_date <= datetime.strptime(end_date, '%Y-%m-%d').date())
    
    # Opening balance
    opening_balance = Decimal('0')
    if start_date:
        start_dt = datetime.strptime(start_date, '%Y-%m-%d').date()
        opening_balance = to_decimal(db.session.query(
            func.sum(AccountingEntry.debit - AccountingEntry.credit)
        ).filter(
            AccountingEntry.account_id == account_id,
            AccountingEntry.entry_date < start_dt,
        ).scalar())

    entries = query.order_by(AccountingEntry.entry_date, AccountingEntry.id).all()

    # Calculate running balance
    running_balance = to_decimal(opening_balance)
    entries_with_balance = []
    
    for entry in entries:
        running_balance += to_decimal(entry.debit) - to_decimal(entry.credit)
        entries_with_balance.append({
            'entry': entry,
            'balance': float(running_balance)
        })
    
    return render_template(
        'accounting/ledger_print.html',
        account=account,
        entries_with_balance=entries_with_balance,
        opening_balance=float(opening_balance),
        closing_balance=float(running_balance),
        start_date=start_date,
        end_date=end_date,
        company_name=_company_name(),
        logo_data_uri=_logo_data_uri(),
    )



# ==================== TRIAL BALANCE ====================

@bp.route('/reports/trial-balance')
@login_required
def trial_balance():
    """Trial Balance"""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    if not start_date or not end_date:
        today = date.today()
        if today.month >= 4:
            start_date = date(today.year, 4, 1)
            end_date = date(today.year + 1, 3, 31)
        else:
            start_date = date(today.year - 1, 4, 1)
            end_date = date(today.year, 3, 31)
    else:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()

    accounts = ChartOfAccounts.query.filter_by(is_active=True).order_by(
        ChartOfAccounts.root_type, ChartOfAccounts.name
    ).all()

    # Opening balances from all entries before start_date
    opening_rows = db.session.query(
        AccountingEntry.account_id,
        func.sum(AccountingEntry.debit - AccountingEntry.credit).label('signed')
    ).filter(
        AccountingEntry.account_id.isnot(None),
        AccountingEntry.entry_date < start_date,
    ).group_by(AccountingEntry.account_id).all()
    opening_signed_map = {r.account_id: float(r.signed or 0.0) for r in opening_rows}

    # Period movement within range
    period_rows = db.session.query(
        AccountingEntry.account_id,
        func.sum(AccountingEntry.debit).label('debit'),
        func.sum(AccountingEntry.credit).label('credit'),
    ).filter(
        AccountingEntry.account_id.isnot(None),
        AccountingEntry.entry_date >= start_date,
        AccountingEntry.entry_date <= end_date,
    ).group_by(AccountingEntry.account_id).all()
    debit_map = {r.account_id: float(r.debit or 0.0) for r in period_rows}
    credit_map = {r.account_id: float(r.credit or 0.0) for r in period_rows}

    rows = []
    total_debit = 0.0
    total_credit = 0.0
    for account in accounts:
        opening = float(opening_signed_map.get(account.id, 0.0))
        debits = debit_map.get(account.id, 0.0)
        credits = credit_map.get(account.id, 0.0)
        closing = opening + debits - credits

        debit_balance = closing if closing >= 0 else 0.0
        credit_balance = abs(closing) if closing < 0 else 0.0

        total_debit += debit_balance
        total_credit += credit_balance

        rows.append({
            'account': account,
            'opening': opening,
            'debit': debits,
            'credit': credits,
            'closing': closing,
            'debit_balance': debit_balance,
            'credit_balance': credit_balance
        })

    return render_template('accounting/trial_balance.html',
                         rows=rows,
                         total_debit=total_debit,
                         total_credit=total_credit,
                         start_date=start_date,
                         end_date=end_date)


@bp.route('/reports/trial-balance/export-excel')
@login_required
def export_trial_balance_excel():
    """Export Trial Balance to Excel"""
    from flask import make_response
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from io import BytesIO
    
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    if not start_date or not end_date:
        today = date.today()
        if today.month >= 4:
            start_date = date(today.year, 4, 1)
            end_date = date(today.year + 1, 3, 31)
        else:
            start_date = date(today.year - 1, 4, 1)
            end_date = date(today.year, 3, 31)
    else:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()

    accounts = ChartOfAccounts.query.filter_by(is_active=True).order_by(
        ChartOfAccounts.root_type, ChartOfAccounts.name
    ).all()

    opening_rows = db.session.query(
        AccountingEntry.account_id,
        func.sum(AccountingEntry.debit - AccountingEntry.credit).label('signed')
    ).filter(
        AccountingEntry.account_id.isnot(None),
        AccountingEntry.entry_date < start_date,
    ).group_by(AccountingEntry.account_id).all()
    opening_signed_map = {r.account_id: float(r.signed or 0.0) for r in opening_rows}

    period_rows = db.session.query(
        AccountingEntry.account_id,
        func.sum(AccountingEntry.debit).label('debit'),
        func.sum(AccountingEntry.credit).label('credit'),
    ).filter(
        AccountingEntry.account_id.isnot(None),
        AccountingEntry.entry_date >= start_date,
        AccountingEntry.entry_date <= end_date,
    ).group_by(AccountingEntry.account_id).all()
    debit_map = {r.account_id: float(r.debit or 0.0) for r in period_rows}
    credit_map = {r.account_id: float(r.credit or 0.0) for r in period_rows}

    rows = []
    total_debit = 0.0
    total_credit = 0.0
    for account in accounts:
        opening = float(opening_signed_map.get(account.id, 0.0))
        debits = debit_map.get(account.id, 0.0)
        credits = credit_map.get(account.id, 0.0)
        closing = opening + debits - credits

        debit_balance = closing if closing >= 0 else 0.0
        credit_balance = abs(closing) if closing < 0 else 0.0

        total_debit += debit_balance
        total_credit += credit_balance

        rows.append({
            'account': account,
            'debit_balance': debit_balance,
            'credit_balance': credit_balance
        })
    
    # Create Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Trial Balance"
    
    header_fill = PatternFill(start_color="D00000", end_color="D00000", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=16, color="D00000")
    bold_font = Font(bold=True)
    
    ws.merge_cells('A1:D1')
    ws['A1'] = 'MOHI INDUSTRIES - TRIAL BALANCE'
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:D2')
    ws['A2'] = f'Period: {start_date.strftime("%d-%m-%Y")} to {end_date.strftime("%d-%m-%Y")}'
    ws['A2'].alignment = Alignment(horizontal='center')
    
    row = 4
    headers = ['Account Code', 'Account Name', 'Debit', 'Credit']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    row += 1
    for r in rows:
        ws.cell(row=row, column=1).value = r['account'].account_number
        ws.cell(row=row, column=2).value = r['account'].name
        ws.cell(row=row, column=3).value = r['debit_balance'] if r['debit_balance'] > 0 else ''
        ws.cell(row=row, column=4).value = r['credit_balance'] if r['credit_balance'] > 0 else ''
        row += 1
    
    row += 1
    ws.cell(row=row, column=2).value = 'TOTAL:'
    ws.cell(row=row, column=2).font = bold_font
    ws.cell(row=row, column=3).value = total_debit
    ws.cell(row=row, column=3).font = Font(bold=True, color="D00000")
    ws.cell(row=row, column=4).value = total_credit
    ws.cell(row=row, column=4).font = Font(bold=True, color="D00000")
    
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = make_response(output.read())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=Trial_Balance_{start_date.strftime("%Y%m%d")}.xlsx'
    
    return response


@bp.route('/reports/trial-balance/print')
@login_required
def print_trial_balance():
    """Print Trial Balance"""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    if not start_date or not end_date:
        today = date.today()
        if today.month >= 4:
            start_date = date(today.year, 4, 1)
            end_date = date(today.year + 1, 3, 31)
        else:
            start_date = date(today.year - 1, 4, 1)
            end_date = date(today.year, 3, 31)
    else:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()

    accounts = ChartOfAccounts.query.filter_by(is_active=True).order_by(
        ChartOfAccounts.root_type, ChartOfAccounts.name
    ).all()

    opening_rows = db.session.query(
        AccountingEntry.account_id,
        func.sum(AccountingEntry.debit - AccountingEntry.credit).label('signed')
    ).filter(
        AccountingEntry.account_id.isnot(None),
        AccountingEntry.entry_date < start_date,
    ).group_by(AccountingEntry.account_id).all()
    opening_signed_map = {r.account_id: float(r.signed or 0.0) for r in opening_rows}

    period_rows = db.session.query(
        AccountingEntry.account_id,
        func.sum(AccountingEntry.debit).label('debit'),
        func.sum(AccountingEntry.credit).label('credit'),
    ).filter(
        AccountingEntry.account_id.isnot(None),
        AccountingEntry.entry_date >= start_date,
        AccountingEntry.entry_date <= end_date,
    ).group_by(AccountingEntry.account_id).all()
    debit_map = {r.account_id: float(r.debit or 0.0) for r in period_rows}
    credit_map = {r.account_id: float(r.credit or 0.0) for r in period_rows}

    rows = []
    total_debit = 0.0
    total_credit = 0.0
    for account in accounts:
        opening = float(opening_signed_map.get(account.id, 0.0))
        debits = debit_map.get(account.id, 0.0)
        credits = credit_map.get(account.id, 0.0)
        closing = opening + debits - credits

        debit_balance = closing if closing >= 0 else 0.0
        credit_balance = abs(closing) if closing < 0 else 0.0

        total_debit += debit_balance
        total_credit += credit_balance

        rows.append({
            'account': account,
            'debit_balance': debit_balance,
            'credit_balance': credit_balance
        })

    return render_template(
        'accounting/trial_balance_print.html',
        rows=rows,
        total_debit=total_debit,
        total_credit=total_credit,
        start_date=start_date,
        end_date=end_date,
        company_name=_company_name(),
        logo_data_uri=_logo_data_uri(),
    )


@bp.route('/reports/trial-balance/email', methods=['POST'])
@login_required
def email_trial_balance():
    """Email Trial Balance as PDF (falls back to HTML if PDF fails)."""
    email_to = _report_recipient_email()
    if not email_to:
        flash('Please provide a recipient email (or set COMPANY_EMAIL).', 'error')
        return redirect(request.referrer or url_for('accounting.trial_balance'))

    start_date = request.form.get('start_date')
    end_date = request.form.get('end_date')
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()

    # Reuse the same logic as trial_balance()
    if not start_date or not end_date:
        today = date.today()
        if today.month >= 4:
            start_date = date(today.year, 4, 1)
            end_date = date(today.year + 1, 3, 31)
        else:
            start_date = date(today.year - 1, 4, 1)
            end_date = date(today.year, 3, 31)

    accounts = ChartOfAccounts.query.filter_by(is_active=True).order_by(
        ChartOfAccounts.root_type, ChartOfAccounts.name
    ).all()

    opening_rows = db.session.query(
        AccountingEntry.account_id,
        func.sum(AccountingEntry.debit - AccountingEntry.credit).label('signed')
    ).filter(
        AccountingEntry.account_id.isnot(None),
        AccountingEntry.entry_date < start_date,
    ).group_by(AccountingEntry.account_id).all()
    opening_signed_map = {r.account_id: float(r.signed or 0.0) for r in opening_rows}

    period_rows = db.session.query(
        AccountingEntry.account_id,
        func.sum(AccountingEntry.debit).label('debit'),
        func.sum(AccountingEntry.credit).label('credit'),
    ).filter(
        AccountingEntry.account_id.isnot(None),
        AccountingEntry.entry_date >= start_date,
        AccountingEntry.entry_date <= end_date,
    ).group_by(AccountingEntry.account_id).all()
    debit_map = {r.account_id: float(r.debit or 0.0) for r in period_rows}
    credit_map = {r.account_id: float(r.credit or 0.0) for r in period_rows}

    rows = []
    total_debit = 0.0
    total_credit = 0.0
    for account in accounts:
        opening = float(opening_signed_map.get(account.id, 0.0))
        debits = debit_map.get(account.id, 0.0)
        credits = credit_map.get(account.id, 0.0)
        closing = opening + debits - credits

        debit_balance = closing if closing >= 0 else 0.0
        credit_balance = abs(closing) if closing < 0 else 0.0

        total_debit += debit_balance
        total_credit += credit_balance

        rows.append({
            'account': account,
            'debit_balance': debit_balance,
            'credit_balance': credit_balance,
        })

    html = render_template(
        'accounting/trial_balance_print.html',
        rows=rows,
        total_debit=total_debit,
        total_credit=total_credit,
        start_date=start_date,
        end_date=end_date,
        company_name=_company_name(),
        logo_data_uri=_logo_data_uri(),
    )

    subject = f"Trial Balance ({start_date.strftime('%d-%m-%Y')} to {end_date.strftime('%d-%m-%Y')})"
    filename = f"Trial_Balance_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.pdf"

    pdf_bytes = None
    try:
        from weasyprint import HTML

        pdf_bytes = HTML(string=html).write_pdf()
    except Exception as e:
        current_app.logger.warning(f"Trial Balance PDF generation failed: {e}")

    attachments = []
    if pdf_bytes:
        attachments = [{'filename': filename, 'content_type': 'application/pdf', 'data': pdf_bytes}]

    ok = EmailService.send_html_email(email_to, subject, html, attachments=attachments)
    flash('Trial Balance emailed successfully.' if ok else 'Failed to send Trial Balance email.', 'success' if ok else 'error')
    return redirect(url_for('accounting.trial_balance', start_date=start_date, end_date=end_date))



# ==================== AR/AP AGING ====================

@bp.route('/reports/ar-aging')
@login_required
def ar_aging():
    """Accounts Receivable aging"""
    today = date.today()
    open_orders = Order.query.filter(Order.payment_status.in_(['pending', 'partial'])).all()

    buckets = {
        '0-30': 0.0,
        '31-60': 0.0,
        '61-90': 0.0,
        '90+': 0.0
    }
    per_customer = {}

    for order in open_orders:
        outstanding = (order.total_amount or 0.0) - (order.paid_amount or 0.0)
        if outstanding <= 0:
            continue
        age = (today - order.order_date).days

        if age <= 30:
            bucket = '0-30'
        elif age <= 60:
            bucket = '31-60'
        elif age <= 90:
            bucket = '61-90'
        else:
            bucket = '90+'

        buckets[bucket] += outstanding

        key = order.distributor_id
        if key not in per_customer:
            per_customer[key] = {
                'distributor': order.distributor,
                'total': 0.0,
                '0-30': 0.0,
                '31-60': 0.0,
                '61-90': 0.0,
                '90+': 0.0
            }
        per_customer[key][bucket] += outstanding
        per_customer[key]['total'] += outstanding

    rows = sorted(per_customer.values(), key=lambda x: x['total'], reverse=True)

    return render_template('accounting/ar_aging.html',
                         buckets=buckets,
                         rows=rows,
                         as_on=today)


@bp.route('/reports/ar-aging/export-excel')
@login_required
def export_ar_aging_excel():
    """Export AR Aging to Excel"""
    from flask import make_response
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from io import BytesIO
    
    today = date.today()
    open_orders = Order.query.filter(Order.payment_status.in_(['pending', 'partial'])).all()

    buckets = {'0-30': 0.0, '31-60': 0.0, '61-90': 0.0, '90+': 0.0}
    per_customer = {}

    for order in open_orders:
        outstanding = (order.total_amount or 0.0) - (order.paid_amount or 0.0)
        if outstanding <= 0:
            continue
        age = (today - order.order_date).days

        if age <= 30:
            bucket = '0-30'
        elif age <= 60:
            bucket = '31-60'
        elif age <= 90:
            bucket = '61-90'
        else:
            bucket = '90+'

        buckets[bucket] += outstanding

        key = order.distributor_id
        if key not in per_customer:
            per_customer[key] = {
                'distributor': order.distributor,
                'total': 0.0,
                '0-30': 0.0,
                '31-60': 0.0,
                '61-90': 0.0,
                '90+': 0.0
            }
        per_customer[key][bucket] += outstanding
        per_customer[key]['total'] += outstanding

    rows = sorted(per_customer.values(), key=lambda x: x['total'], reverse=True)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "AR Aging"
    
    header_fill = PatternFill(start_color="D00000", end_color="D00000", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=16, color="D00000")
    bold_font = Font(bold=True)
    
    ws.merge_cells('A1:F1')
    ws['A1'] = 'MOHI INDUSTRIES - ACCOUNTS RECEIVABLE AGING'
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:F2')
    ws['A2'] = f'As on: {today.strftime("%d-%m-%Y")}'
    ws['A2'].alignment = Alignment(horizontal='center')
    
    row = 4
    headers = ['Customer', '0-30 Days', '31-60 Days', '61-90 Days', '90+ Days', 'Total']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    row += 1
    for r in rows:
        ws.cell(row=row, column=1).value = r['distributor'].business_name
        ws.cell(row=row, column=2).value = r['0-30']
        ws.cell(row=row, column=3).value = r['31-60']
        ws.cell(row=row, column=4).value = r['61-90']
        ws.cell(row=row, column=5).value = r['90+']
        ws.cell(row=row, column=6).value = r['total']
        row += 1
    
    row += 1
    ws.cell(row=row, column=1).value = 'TOTAL:'
    ws.cell(row=row, column=1).font = bold_font
    ws.cell(row=row, column=2).value = buckets['0-30']
    ws.cell(row=row, column=2).font = Font(bold=True, color="D00000")
    ws.cell(row=row, column=3).value = buckets['31-60']
    ws.cell(row=row, column=3).font = Font(bold=True, color="D00000")
    ws.cell(row=row, column=4).value = buckets['61-90']
    ws.cell(row=row, column=4).font = Font(bold=True, color="D00000")
    ws.cell(row=row, column=5).value = buckets['90+']
    ws.cell(row=row, column=5).font = Font(bold=True, color="D00000")
    ws.cell(row=row, column=6).value = sum(buckets.values())
    ws.cell(row=row, column=6).font = Font(bold=True, size=12, color="D00000")
    
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = make_response(output.read())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=AR_Aging_{today.strftime("%Y%m%d")}.xlsx'
    
    return response


@bp.route('/reports/ar-aging/print')
@login_required
def print_ar_aging():
    """Print AR Aging"""
    today = date.today()
    open_orders = Order.query.filter(Order.payment_status.in_(['pending', 'partial'])).all()

    buckets = {'0-30': 0.0, '31-60': 0.0, '61-90': 0.0, '90+': 0.0}
    per_customer = {}

    for order in open_orders:
        outstanding = (order.total_amount or 0.0) - (order.paid_amount or 0.0)
        if outstanding <= 0:
            continue
        age = (today - order.order_date).days

        if age <= 30:
            bucket = '0-30'
        elif age <= 60:
            bucket = '31-60'
        elif age <= 90:
            bucket = '61-90'
        else:
            bucket = '90+'

        buckets[bucket] += outstanding

        key = order.distributor_id
        if key not in per_customer:
            per_customer[key] = {
                'distributor': order.distributor,
                'total': 0.0,
                '0-30': 0.0,
                '31-60': 0.0,
                '61-90': 0.0,
                '90+': 0.0
            }
        per_customer[key][bucket] += outstanding
        per_customer[key]['total'] += outstanding

    rows = sorted(per_customer.values(), key=lambda x: x['total'], reverse=True)

    return render_template(
        'accounting/ar_aging_print.html',
        buckets=buckets,
        rows=rows,
        as_on=today,
        company_name=_company_name(),
        logo_data_uri=_logo_data_uri(),
    )



@bp.route('/reports/ap-aging')
@login_required
def ap_aging():
    """Accounts Payable aging"""
    today = date.today()
    open_bills = VendorBill.query.filter(VendorBill.status.in_(['pending', 'partial'])).all()

    buckets = {
        '0-30': 0.0,
        '31-60': 0.0,
        '61-90': 0.0,
        '90+': 0.0
    }
    per_vendor = {}

    for bill in open_bills:
        outstanding = (bill.total_amount or 0.0) - (bill.paid_amount or 0.0)
        if outstanding <= 0:
            continue
        age = (today - bill.bill_date).days

        if age <= 30:
            bucket = '0-30'
        elif age <= 60:
            bucket = '31-60'
        elif age <= 90:
            bucket = '61-90'
        else:
            bucket = '90+'

        buckets[bucket] += outstanding

        key = bill.vendor_id
        if key not in per_vendor:
            per_vendor[key] = {
                'vendor': bill.vendor,
                'total': 0.0,
                '0-30': 0.0,
                '31-60': 0.0,
                '61-90': 0.0,
                '90+': 0.0
            }
        per_vendor[key][bucket] += outstanding
        per_vendor[key]['total'] += outstanding

    rows = sorted(per_vendor.values(), key=lambda x: x['total'], reverse=True)

    return render_template('accounting/ap_aging.html',
                         buckets=buckets,
                         rows=rows,
                         as_on=today)


@bp.route('/reports/ap-aging/export-excel')
@login_required
def export_ap_aging_excel():
    """Export AP Aging to Excel"""
    from flask import make_response
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from io import BytesIO
    
    today = date.today()
    open_bills = VendorBill.query.filter(VendorBill.status.in_(['pending', 'partial'])).all()

    buckets = {'0-30': 0.0, '31-60': 0.0, '61-90': 0.0, '90+': 0.0}
    per_vendor = {}

    for bill in open_bills:
        outstanding = (bill.total_amount or 0.0) - (bill.paid_amount or 0.0)
        if outstanding <= 0:
            continue
        age = (today - bill.bill_date).days

        if age <= 30:
            bucket = '0-30'
        elif age <= 60:
            bucket = '31-60'
        elif age <= 90:
            bucket = '61-90'
        else:
            bucket = '90+'

        buckets[bucket] += outstanding

        key = bill.vendor_id
        if key not in per_vendor:
            per_vendor[key] = {
                'vendor': bill.vendor,
                'total': 0.0,
                '0-30': 0.0,
                '31-60': 0.0,
                '61-90': 0.0,
                '90+': 0.0
            }
        per_vendor[key][bucket] += outstanding
        per_vendor[key]['total'] += outstanding

    rows = sorted(per_vendor.values(), key=lambda x: x['total'], reverse=True)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "AP Aging"
    
    header_fill = PatternFill(start_color="D00000", end_color="D00000", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=16, color="D00000")
    bold_font = Font(bold=True)
    
    ws.merge_cells('A1:F1')
    ws['A1'] = 'MOHI INDUSTRIES - ACCOUNTS PAYABLE AGING'
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:F2')
    ws['A2'] = f'As on: {today.strftime("%d-%m-%Y")}'
    ws['A2'].alignment = Alignment(horizontal='center')
    
    row = 4
    headers = ['Vendor', '0-30 Days', '31-60 Days', '61-90 Days', '90+ Days', 'Total']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    row += 1
    for r in rows:
        ws.cell(row=row, column=1).value = r['vendor'].business_name
        ws.cell(row=row, column=2).value = r['0-30']
        ws.cell(row=row, column=3).value = r['31-60']
        ws.cell(row=row, column=4).value = r['61-90']
        ws.cell(row=row, column=5).value = r['90+']
        ws.cell(row=row, column=6).value = r['total']
        row += 1
    
    row += 1
    ws.cell(row=row, column=1).value = 'TOTAL:'
    ws.cell(row=row, column=1).font = bold_font
    ws.cell(row=row, column=2).value = buckets['0-30']
    ws.cell(row=row, column=2).font = Font(bold=True, color="D00000")
    ws.cell(row=row, column=3).value = buckets['31-60']
    ws.cell(row=row, column=3).font = Font(bold=True, color="D00000")
    ws.cell(row=row, column=4).value = buckets['61-90']
    ws.cell(row=row, column=4).font = Font(bold=True, color="D00000")
    ws.cell(row=row, column=5).value = buckets['90+']
    ws.cell(row=row, column=5).font = Font(bold=True, color="D00000")
    ws.cell(row=row, column=6).value = sum(buckets.values())
    ws.cell(row=row, column=6).font = Font(bold=True, size=12, color="D00000")
    
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = make_response(output.read())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=AP_Aging_{today.strftime("%Y%m%d")}.xlsx'
    
    return response


@bp.route('/reports/ap-aging/print')
@login_required
def print_ap_aging():
    """Print AP Aging"""
    today = date.today()
    open_bills = VendorBill.query.filter(VendorBill.status.in_(['pending', 'partial'])).all()

    buckets = {'0-30': 0.0, '31-60': 0.0, '61-90': 0.0, '90+': 0.0}
    per_vendor = {}

    for bill in open_bills:
        outstanding = (bill.total_amount or 0.0) - (bill.paid_amount or 0.0)
        if outstanding <= 0:
            continue
        age = (today - bill.bill_date).days

        if age <= 30:
            bucket = '0-30'
        elif age <= 60:
            bucket = '31-60'
        elif age <= 90:
            bucket = '61-90'
        else:
            bucket = '90+'

        buckets[bucket] += outstanding

        key = bill.vendor_id
        if key not in per_vendor:
            per_vendor[key] = {
                'vendor': bill.vendor,
                'total': 0.0,
                '0-30': 0.0,
                '31-60': 0.0,
                '61-90': 0.0,
                '90+': 0.0
            }
        per_vendor[key][bucket] += outstanding
        per_vendor[key]['total'] += outstanding

    rows = sorted(per_vendor.values(), key=lambda x: x['total'], reverse=True)

    return render_template(
        'accounting/ap_aging_print.html',
        buckets=buckets,
        rows=rows,
        as_on=today,
        company_name=_company_name(),
        logo_data_uri=_logo_data_uri(),
    )



# ==================== REPORTS ====================

@bp.route('/reports/profit-loss')
@login_required
def profit_loss():
    """Profit & Loss Statement"""
    # Get date range
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    if not start_date or not end_date:
        # Default to current financial year (Apr 1 to Mar 31)
        today = date.today()
        if today.month >= 4:
            start_date = date(today.year, 4, 1)
            end_date = date(today.year + 1, 3, 31)
        else:
            start_date = date(today.year - 1, 4, 1)
            end_date = date(today.year, 3, 31)
    else:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    # Calculate Income
    income_entries = db.session.query(
        AccountingEntry.account_head,
        func.sum(AccountingEntry.credit - AccountingEntry.debit).label('amount')
    ).join(ChartOfAccounts, AccountingEntry.account_id == ChartOfAccounts.id
    ).filter(
        ChartOfAccounts.root_type == 'Income',
        AccountingEntry.entry_date >= start_date,
        AccountingEntry.entry_date <= end_date
    ).group_by(AccountingEntry.account_head).all()
    
    total_income = sum(entry.amount for entry in income_entries)
    
    # Calculate Expenses
    expense_entries = db.session.query(
        AccountingEntry.account_head,
        func.sum(AccountingEntry.debit - AccountingEntry.credit).label('amount')
    ).join(ChartOfAccounts, AccountingEntry.account_id == ChartOfAccounts.id
    ).filter(
        ChartOfAccounts.root_type == 'Expense',
        AccountingEntry.entry_date >= start_date,
        AccountingEntry.entry_date <= end_date
    ).group_by(AccountingEntry.account_head).all()
    
    total_expenses = sum(entry.amount for entry in expense_entries)
    
    # Calculate Profit/Loss
    net_profit = total_income - total_expenses
    
    return render_template('accounting/profit_loss.html',
                         income_entries=income_entries,
                         expense_entries=expense_entries,
                         total_income=total_income,
                         total_expenses=total_expenses,
                         net_profit=net_profit,
                         start_date=start_date,
                         end_date=end_date)


@bp.route('/reports/profit-loss/export-excel')
@login_required
def export_profit_loss_excel():
    """Export Profit & Loss to Excel"""
    from flask import make_response
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from io import BytesIO
    
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    if not start_date or not end_date:
        today = date.today()
        if today.month >= 4:
            start_date = date(today.year, 4, 1)
            end_date = date(today.year + 1, 3, 31)
        else:
            start_date = date(today.year - 1, 4, 1)
            end_date = date(today.year, 3, 31)
    else:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    income_entries = db.session.query(
        AccountingEntry.account_head,
        func.sum(AccountingEntry.credit - AccountingEntry.debit).label('amount')
    ).join(ChartOfAccounts, AccountingEntry.account_id == ChartOfAccounts.id
    ).filter(
        ChartOfAccounts.root_type == 'Income',
        AccountingEntry.entry_date >= start_date,
        AccountingEntry.entry_date <= end_date
    ).group_by(AccountingEntry.account_head).all()
    
    total_income = sum(entry.amount for entry in income_entries)
    
    expense_entries = db.session.query(
        AccountingEntry.account_head,
        func.sum(AccountingEntry.debit - AccountingEntry.credit).label('amount')
    ).join(ChartOfAccounts, AccountingEntry.account_id == ChartOfAccounts.id
    ).filter(
        ChartOfAccounts.root_type == 'Expense',
        AccountingEntry.entry_date >= start_date,
        AccountingEntry.entry_date <= end_date
    ).group_by(AccountingEntry.account_head).all()
    
    total_expenses = sum(entry.amount for entry in expense_entries)
    net_profit = total_income - total_expenses
    
    # Create Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Profit & Loss"
    
    header_fill = PatternFill(start_color="D00000", end_color="D00000", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=16, color="D00000")
    bold_font = Font(bold=True)
    
    ws.merge_cells('A1:C1')
    ws['A1'] = 'MOHI INDUSTRIES - PROFIT & LOSS STATEMENT'
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:C2')
    ws['A2'] = f'Period: {start_date.strftime("%d-%m-%Y")} to {end_date.strftime("%d-%m-%Y")}'
    ws['A2'].alignment = Alignment(horizontal='center')
    
    row = 4
    ws[f'A{row}'] = 'INCOME'
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    
    row += 1
    for entry in income_entries:
        ws.cell(row=row, column=1).value = entry.account_head
        ws.cell(row=row, column=3).value = float(entry.amount)
        row += 1
    
    ws.cell(row=row, column=1).value = 'Total Income:'
    ws.cell(row=row, column=1).font = bold_font
    ws.cell(row=row, column=3).value = total_income
    ws.cell(row=row, column=3).font = Font(bold=True, color="D00000")
    
    row += 2
    ws[f'A{row}'] = 'EXPENSES'
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    
    row += 1
    for entry in expense_entries:
        ws.cell(row=row, column=1).value = entry.account_head
        ws.cell(row=row, column=3).value = float(entry.amount)
        row += 1
    
    ws.cell(row=row, column=1).value = 'Total Expenses:'
    ws.cell(row=row, column=1).font = bold_font
    ws.cell(row=row, column=3).value = total_expenses
    ws.cell(row=row, column=3).font = Font(bold=True, color="D00000")
    
    row += 2
    ws.cell(row=row, column=1).value = 'NET PROFIT/LOSS:'
    ws.cell(row=row, column=1).font = Font(bold=True, size=14, color="D00000")
    ws.cell(row=row, column=3).value = net_profit
    ws.cell(row=row, column=3).font = Font(bold=True, size=14, color="D00000")
    
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = make_response(output.read())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=Profit_Loss_{start_date.strftime("%Y%m%d")}.xlsx'
    
    return response


@bp.route('/reports/profit-loss/print')
@login_required
def print_profit_loss():
    """Print Profit & Loss"""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    if not start_date or not end_date:
        today = date.today()
        if today.month >= 4:
            start_date = date(today.year, 4, 1)
            end_date = date(today.year + 1, 3, 31)
        else:
            start_date = date(today.year - 1, 4, 1)
            end_date = date(today.year, 3, 31)
    else:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    income_entries = db.session.query(
        AccountingEntry.account_head,
        func.sum(AccountingEntry.credit - AccountingEntry.debit).label('amount')
    ).join(ChartOfAccounts, AccountingEntry.account_id == ChartOfAccounts.id
    ).filter(
        ChartOfAccounts.root_type == 'Income',
        AccountingEntry.entry_date >= start_date,
        AccountingEntry.entry_date <= end_date
    ).group_by(AccountingEntry.account_head).all()
    
    total_income = sum(entry.amount for entry in income_entries)
    
    expense_entries = db.session.query(
        AccountingEntry.account_head,
        func.sum(AccountingEntry.debit - AccountingEntry.credit).label('amount')
    ).join(ChartOfAccounts, AccountingEntry.account_id == ChartOfAccounts.id
    ).filter(
        ChartOfAccounts.root_type == 'Expense',
        AccountingEntry.entry_date >= start_date,
        AccountingEntry.entry_date <= end_date
    ).group_by(AccountingEntry.account_head).all()
    
    total_expenses = sum(entry.amount for entry in expense_entries)
    net_profit = total_income - total_expenses
    
    return render_template(
        'accounting/profit_loss_print.html',
        income_entries=income_entries,
        expense_entries=expense_entries,
        total_income=total_income,
        total_expenses=total_expenses,
        net_profit=net_profit,
        start_date=start_date,
        end_date=end_date,
        company_name=_company_name(),
        logo_data_uri=_logo_data_uri(),
    )


@bp.route('/reports/profit-loss/email', methods=['POST'])
@login_required
def email_profit_loss():
    """Email Profit & Loss as PDF (falls back to HTML if PDF fails)."""
    email_to = _report_recipient_email()
    if not email_to:
        flash('Please provide a recipient email (or set COMPANY_EMAIL).', 'error')
        return redirect(request.referrer or url_for('accounting.profit_loss'))

    start_date = request.form.get('start_date')
    end_date = request.form.get('end_date')
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()

    if not start_date or not end_date:
        today = date.today()
        if today.month >= 4:
            start_date = date(today.year, 4, 1)
            end_date = date(today.year + 1, 3, 31)
        else:
            start_date = date(today.year - 1, 4, 1)
            end_date = date(today.year, 3, 31)

    income_entries = db.session.query(
        AccountingEntry.account_head,
        func.sum(AccountingEntry.credit - AccountingEntry.debit).label('amount')
    ).join(ChartOfAccounts, AccountingEntry.account_id == ChartOfAccounts.id
    ).filter(
        ChartOfAccounts.root_type == 'Income',
        AccountingEntry.entry_date >= start_date,
        AccountingEntry.entry_date <= end_date
    ).group_by(AccountingEntry.account_head).all()
    total_income = sum(entry.amount for entry in income_entries)

    expense_entries = db.session.query(
        AccountingEntry.account_head,
        func.sum(AccountingEntry.debit - AccountingEntry.credit).label('amount')
    ).join(ChartOfAccounts, AccountingEntry.account_id == ChartOfAccounts.id
    ).filter(
        ChartOfAccounts.root_type == 'Expense',
        AccountingEntry.entry_date >= start_date,
        AccountingEntry.entry_date <= end_date
    ).group_by(AccountingEntry.account_head).all()
    total_expenses = sum(entry.amount for entry in expense_entries)
    net_profit = total_income - total_expenses

    html = render_template(
        'accounting/profit_loss_print.html',
        income_entries=income_entries,
        expense_entries=expense_entries,
        total_income=total_income,
        total_expenses=total_expenses,
        net_profit=net_profit,
        start_date=start_date,
        end_date=end_date,
        company_name=_company_name(),
        logo_data_uri=_logo_data_uri(),
    )

    subject = f"Profit & Loss ({start_date.strftime('%d-%m-%Y')} to {end_date.strftime('%d-%m-%Y')})"
    filename = f"Profit_Loss_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.pdf"

    pdf_bytes = None
    try:
        from weasyprint import HTML

        pdf_bytes = HTML(string=html).write_pdf()
    except Exception as e:
        current_app.logger.warning(f"Profit & Loss PDF generation failed: {e}")

    attachments = []
    if pdf_bytes:
        attachments = [{'filename': filename, 'content_type': 'application/pdf', 'data': pdf_bytes}]

    ok = EmailService.send_html_email(email_to, subject, html, attachments=attachments)
    flash('Profit & Loss emailed successfully.' if ok else 'Failed to send Profit & Loss email.', 'success' if ok else 'error')
    return redirect(url_for('accounting.profit_loss', start_date=start_date, end_date=end_date))



@bp.route('/reports/balance-sheet')
@login_required
def balance_sheet():
    """Balance Sheet"""
    as_on_date = request.args.get('as_on_date')
    
    if not as_on_date:
        as_on_date = date.today()
    else:
        as_on_date = datetime.strptime(as_on_date, '%Y-%m-%d').date()
    
    # Calculate Assets
    asset_accounts = ChartOfAccounts.query.filter_by(
        root_type='Asset',
        is_active=True
    ).all()
    
    assets = []
    total_assets = Decimal('0')
    
    for account in asset_accounts:
        balance = Decimal('0')
        entries = AccountingEntry.query.filter(
            AccountingEntry.account_id == account.id,
            AccountingEntry.entry_date <= as_on_date
        ).all()
        
        for entry in entries:
            balance += _to_decimal(entry.debit) - _to_decimal(entry.credit)
        
        assets.append({'account': account, 'balance': balance})
        total_assets += balance
    
    # Calculate Liabilities
    liability_accounts = ChartOfAccounts.query.filter_by(
        root_type='Liability',
        is_active=True
    ).all()
    
    liabilities = []
    total_liabilities = Decimal('0')
    
    for account in liability_accounts:
        balance = Decimal('0')
        entries = AccountingEntry.query.filter(
            AccountingEntry.account_id == account.id,
            AccountingEntry.entry_date <= as_on_date
        ).all()
        
        for entry in entries:
            balance += _to_decimal(entry.credit) - _to_decimal(entry.debit)
        
        liabilities.append({'account': account, 'balance': balance})
        total_liabilities += balance
    
    # Calculate Equity
    equity_accounts = ChartOfAccounts.query.filter_by(
        root_type='Equity',
        is_active=True
    ).all()
    
    equity = []
    total_equity = Decimal('0')
    
    for account in equity_accounts:
        balance = Decimal('0')
        entries = AccountingEntry.query.filter(
            AccountingEntry.account_id == account.id,
            AccountingEntry.entry_date <= as_on_date
        ).all()
        
        for entry in entries:
            balance += _to_decimal(entry.credit) - _to_decimal(entry.debit)
        
        equity.append({'account': account, 'balance': balance})
        total_equity += balance
    
    return render_template('accounting/balance_sheet.html',
                         assets=assets,
                         liabilities=liabilities,
                         equity=equity,
                         total_assets=total_assets,
                         total_liabilities=total_liabilities,
                         total_equity=total_equity,
                         as_on_date=as_on_date)


@bp.route('/reports/balance-sheet/export-excel')
@login_required  
def export_balance_sheet_excel():
    """Export Balance Sheet to Excel"""
    from flask import make_response
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from io import BytesIO
    
    as_on_date = request.args.get('as_on_date')
    if not as_on_date:
        as_on_date = date.today()
    else:
        as_on_date = datetime.strptime(as_on_date, '%Y-%m-%d').date()
    
    asset_accounts = ChartOfAccounts.query.filter_by(root_type='Asset', is_active=True).all()
    assets = []
    total_assets = Decimal('0')
    for account in asset_accounts:
        entries = AccountingEntry.query.filter(
            AccountingEntry.account_id == account.id,
            AccountingEntry.entry_date <= as_on_date,
        ).all()
        balance = Decimal('0')
        for e in entries:
            balance += _to_decimal(e.debit) - _to_decimal(e.credit)
        assets.append({'account': account, 'balance': balance})
        total_assets += balance
    
    liability_accounts = ChartOfAccounts.query.filter_by(root_type='Liability', is_active=True).all()
    liabilities = []
    total_liabilities = Decimal('0')
    for account in liability_accounts:
        entries = AccountingEntry.query.filter(
            AccountingEntry.account_id == account.id,
            AccountingEntry.entry_date <= as_on_date,
        ).all()
        balance = Decimal('0')
        for e in entries:
            balance += _to_decimal(e.credit) - _to_decimal(e.debit)
        liabilities.append({'account': account, 'balance': balance})
        total_liabilities += balance
    
    equity_accounts = ChartOfAccounts.query.filter_by(root_type='Equity', is_active=True).all()
    equity = []
    total_equity = Decimal('0')
    for account in equity_accounts:
        entries = AccountingEntry.query.filter(
            AccountingEntry.account_id == account.id,
            AccountingEntry.entry_date <= as_on_date,
        ).all()
        balance = Decimal('0')
        for e in entries:
            balance += _to_decimal(e.credit) - _to_decimal(e.debit)
        equity.append({'account': account, 'balance': balance})
        total_equity += balance
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Balance Sheet"
    
    header_fill = PatternFill(start_color="D00000", end_color="D00000", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=16, color="D00000")
    bold_font = Font(bold=True)
    
    ws.merge_cells('A1:D1')
    ws['A1'] = 'MOHI INDUSTRIES - BALANCE SHEET'
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:D2')
    ws['A2'] = f'As on: {as_on_date.strftime("%d-%m-%Y")}'
    ws['A2'].alignment = Alignment(horizontal='center')
    
    row = 4
    ws[f'A{row}'] = 'ASSETS'
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws[f'C{row}'] = 'LIABILITIES & EQUITY'
    ws[f'C{row}'].font = header_font
    ws[f'C{row}'].fill = header_fill
    
    row += 1
    max_rows = max(len(assets), len(liabilities) + len(equity) + 2)
    
    for i in range(max_rows):
        if i < len(assets):
            ws.cell(row=row+i, column=1).value = assets[i]['account'].name
            ws.cell(row=row+i, column=2).value = float(assets[i]['balance'])
        
        if i < len(liabilities):
            ws.cell(row=row+i, column=3).value = liabilities[i]['account'].name
            ws.cell(row=row+i, column=4).value = float(liabilities[i]['balance'])
        elif i == len(liabilities) + 1:
            ws.cell(row=row+i, column=3).value = 'EQUITY'
            ws.cell(row=row+i, column=3).font = bold_font
        elif i > len(liabilities) + 1 and i < len(liabilities) + len(equity) + 2:
            eq_idx = i - len(liabilities) - 2
            if eq_idx < len(equity):
                ws.cell(row=row+i, column=3).value = equity[eq_idx]['account'].name
                ws.cell(row=row+i, column=4).value = float(equity[eq_idx]['balance'])
    
    row += max_rows + 1
    ws.cell(row=row, column=1).value = 'TOTAL ASSETS:'
    ws.cell(row=row, column=1).font = Font(bold=True, size=12, color="D00000")
    ws.cell(row=row, column=2).value = float(total_assets)
    ws.cell(row=row, column=2).font = Font(bold=True, size=12, color="D00000")
    
    ws.cell(row=row, column=3).value = 'TOTAL LIABILITIES & EQUITY:'
    ws.cell(row=row, column=3).font = Font(bold=True, size=12, color="D00000")
    ws.cell(row=row, column=4).value = float(total_liabilities + total_equity)
    ws.cell(row=row, column=4).font = Font(bold=True, size=12, color="D00000")
    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 20
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = make_response(output.read())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=Balance_Sheet_{as_on_date.strftime("%Y%m%d")}.xlsx'
    
    return response


@bp.route('/reports/balance-sheet/print')
@login_required
def print_balance_sheet():
    """Print Balance Sheet"""
    as_on_date = request.args.get('as_on_date')
    if not as_on_date:
        as_on_date = date.today()
    else:
        as_on_date = datetime.strptime(as_on_date, '%Y-%m-%d').date()
    
    asset_accounts = ChartOfAccounts.query.filter_by(root_type='Asset', is_active=True).all()
    assets = []
    total_assets = Decimal('0')
    for account in asset_accounts:
        entries = AccountingEntry.query.filter(
            AccountingEntry.account_id == account.id,
            AccountingEntry.entry_date <= as_on_date,
        ).all()
        balance = Decimal('0')
        for e in entries:
            balance += _to_decimal(e.debit) - _to_decimal(e.credit)
        assets.append({'account': account, 'balance': balance})
        total_assets += balance
    
    liability_accounts = ChartOfAccounts.query.filter_by(root_type='Liability', is_active=True).all()
    liabilities = []
    total_liabilities = Decimal('0')
    for account in liability_accounts:
        entries = AccountingEntry.query.filter(
            AccountingEntry.account_id == account.id,
            AccountingEntry.entry_date <= as_on_date,
        ).all()
        balance = Decimal('0')
        for e in entries:
            balance += _to_decimal(e.credit) - _to_decimal(e.debit)
        liabilities.append({'account': account, 'balance': balance})
        total_liabilities += balance
    
    equity_accounts = ChartOfAccounts.query.filter_by(root_type='Equity', is_active=True).all()
    equity = []
    total_equity = Decimal('0')
    for account in equity_accounts:
        entries = AccountingEntry.query.filter(
            AccountingEntry.account_id == account.id,
            AccountingEntry.entry_date <= as_on_date,
        ).all()
        balance = Decimal('0')
        for e in entries:
            balance += _to_decimal(e.credit) - _to_decimal(e.debit)
        equity.append({'account': account, 'balance': balance})
        total_equity += balance
    
    return render_template(
        'accounting/balance_sheet_print.html',
        assets=assets,
        liabilities=liabilities,
        equity=equity,
        total_assets=total_assets,
        total_liabilities=total_liabilities,
        total_equity=total_equity,
        as_on_date=as_on_date,
        company_name=_company_name(),
        logo_data_uri=_logo_data_uri(),
    )


@bp.route('/reports/balance-sheet/email', methods=['POST'])
@login_required
def email_balance_sheet():
    """Email Balance Sheet as PDF (falls back to HTML if PDF fails)."""
    email_to = _report_recipient_email()
    if not email_to:
        flash('Please provide a recipient email (or set COMPANY_EMAIL).', 'error')
        return redirect(request.referrer or url_for('accounting.balance_sheet'))

    as_on_date = request.form.get('as_on_date')
    if not as_on_date:
        as_on_date = date.today()
    else:
        as_on_date = datetime.strptime(as_on_date, '%Y-%m-%d').date()

    asset_accounts = ChartOfAccounts.query.filter_by(root_type='Asset', is_active=True).all()
    assets = []
    total_assets = Decimal('0')
    for account in asset_accounts:
        entries = AccountingEntry.query.filter(
            AccountingEntry.account_id == account.id,
            AccountingEntry.entry_date <= as_on_date,
        ).all()
        balance = Decimal('0')
        for e in entries:
            balance += _to_decimal(e.debit) - _to_decimal(e.credit)
        assets.append({'account': account, 'balance': balance})
        total_assets += balance

    liability_accounts = ChartOfAccounts.query.filter_by(root_type='Liability', is_active=True).all()
    liabilities = []
    total_liabilities = Decimal('0')
    for account in liability_accounts:
        entries = AccountingEntry.query.filter(
            AccountingEntry.account_id == account.id,
            AccountingEntry.entry_date <= as_on_date,
        ).all()
        balance = Decimal('0')
        for e in entries:
            balance += _to_decimal(e.credit) - _to_decimal(e.debit)
        liabilities.append({'account': account, 'balance': balance})
        total_liabilities += balance

    equity_accounts = ChartOfAccounts.query.filter_by(root_type='Equity', is_active=True).all()
    equity = []
    total_equity = Decimal('0')
    for account in equity_accounts:
        entries = AccountingEntry.query.filter(
            AccountingEntry.account_id == account.id,
            AccountingEntry.entry_date <= as_on_date,
        ).all()
        balance = Decimal('0')
        for e in entries:
            balance += _to_decimal(e.credit) - _to_decimal(e.debit)
        equity.append({'account': account, 'balance': balance})
        total_equity += balance

    html = render_template(
        'accounting/balance_sheet_print.html',
        assets=assets,
        liabilities=liabilities,
        equity=equity,
        total_assets=total_assets,
        total_liabilities=total_liabilities,
        total_equity=total_equity,
        as_on_date=as_on_date,
        company_name=_company_name(),
        logo_data_uri=_logo_data_uri(),
    )

    subject = f"Balance Sheet (As on {as_on_date.strftime('%d-%m-%Y')})"
    filename = f"Balance_Sheet_{as_on_date.strftime('%Y%m%d')}.pdf"

    pdf_bytes = None
    try:
        from weasyprint import HTML

        pdf_bytes = HTML(string=html).write_pdf()
    except Exception as e:
        current_app.logger.warning(f"Balance Sheet PDF generation failed: {e}")

    attachments = []
    if pdf_bytes:
        attachments = [{'filename': filename, 'content_type': 'application/pdf', 'data': pdf_bytes}]

    ok = EmailService.send_html_email(email_to, subject, html, attachments=attachments)
    flash('Balance Sheet emailed successfully.' if ok else 'Failed to send Balance Sheet email.', 'success' if ok else 'error')
    return redirect(url_for('accounting.balance_sheet', as_on_date=as_on_date))



@bp.route('/reports/day-book')
@login_required
def day_book():
    """Day Book - All transactions for a day"""
    selected_date = request.args.get('date')
    
    if not selected_date:
        selected_date = date.today()
    else:
        selected_date = datetime.strptime(selected_date, '%Y-%m-%d').date()
    
    # Get all entries for the day
    entries = AccountingEntry.query.filter_by(
        entry_date=selected_date
    ).order_by(AccountingEntry.id).all()
    
    # Calculate totals
    total_debit = sum(e.debit for e in entries)
    total_credit = sum(e.credit for e in entries)
    
    return render_template('accounting/day_book.html',
                         entries=entries,
                         selected_date=selected_date,
                         total_debit=total_debit,
                         total_credit=total_credit)


@bp.route('/reports/day-book/export-excel')
@login_required
def export_day_book_excel():
    """Export Day Book to Excel"""
    from flask import make_response
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from io import BytesIO
    
    selected_date = request.args.get('date')
    if not selected_date:
        selected_date = date.today()
    else:
        selected_date = datetime.strptime(selected_date, '%Y-%m-%d').date()
    
    entries = AccountingEntry.query.filter_by(entry_date=selected_date).order_by(AccountingEntry.id).all()
    total_debit = sum(e.debit for e in entries)
    total_credit = sum(e.credit for e in entries)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Day Book"
    
    header_fill = PatternFill(start_color="D00000", end_color="D00000", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=16, color="D00000")
    bold_font = Font(bold=True)
    
    ws.merge_cells('A1:F1')
    ws['A1'] = 'MOHI INDUSTRIES - DAY BOOK'
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:F2')
    ws['A2'] = f'Date: {selected_date.strftime("%d-%m-%Y")}'
    ws['A2'].font = bold_font
    ws['A2'].alignment = Alignment(horizontal='center')
    
    row = 4
    headers = ['Entry ID', 'Account', 'Reference', 'Description', 'Debit', 'Credit']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    row += 1
    for entry in entries:
        ws.cell(row=row, column=1).value = entry.id
        ws.cell(row=row, column=2).value = entry.account_head
        ws.cell(row=row, column=3).value = f"{entry.reference_type}-{entry.reference_id}"
        ws.cell(row=row, column=4).value = entry.description
        ws.cell(row=row, column=5).value = float(entry.debit) if entry.debit > 0 else ''
        ws.cell(row=row, column=6).value = float(entry.credit) if entry.credit > 0 else ''
        row += 1
    
    row += 1
    ws.cell(row=row, column=4).value = 'TOTAL:'
    ws.cell(row=row, column=4).font = bold_font
    ws.cell(row=row, column=5).value = total_debit
    ws.cell(row=row, column=5).font = Font(bold=True, color="D00000")
    ws.cell(row=row, column=6).value = total_credit
    ws.cell(row=row, column=6).font = Font(bold=True, color="D00000")
    
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = make_response(output.read())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=Day_Book_{selected_date.strftime("%Y%m%d")}.xlsx'
    
    return response


@bp.route('/reports/day-book/print')
@login_required
def print_day_book():
    """Print Day Book"""
    selected_date = request.args.get('date')
    if not selected_date:
        selected_date = date.today()
    else:
        selected_date = datetime.strptime(selected_date, '%Y-%m-%d').date()
    
    entries = AccountingEntry.query.filter_by(entry_date=selected_date).order_by(AccountingEntry.id).all()
    total_debit = sum(e.debit for e in entries)
    total_credit = sum(e.credit for e in entries)
    
    return render_template(
        'accounting/day_book_print.html',
        entries=entries,
        selected_date=selected_date,
        total_debit=total_debit,
        total_credit=total_credit,
        company_name=_company_name(),
        logo_data_uri=_logo_data_uri(),
    )



@bp.route('/reports/cash-flow')
@login_required
def cash_flow():
    """Cash Flow Statement"""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    if not start_date or not end_date:
        end_date = date.today()
        start_date = end_date - timedelta(days=30)
    else:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    # Cash Inflows (Payments received)
    cash_inflows = Payment.query.filter(
        Payment.payment_date >= start_date,
        Payment.payment_date <= end_date,
        Payment.status == 'cleared'
    ).all()
    
    total_inflow = sum(p.amount for p in cash_inflows)
    
    # Cash Outflows (Expenses)
    cash_outflows = Expense.query.filter(
        Expense.expense_date >= start_date,
        Expense.expense_date <= end_date,
        Expense.payment_status == 'paid'
    ).all()
    
    total_outflow = sum(e.total_amount for e in cash_outflows)
    
    # Net Cash Flow
    net_cash_flow = total_inflow - total_outflow
    
    return render_template('accounting/cash_flow.html',
                         cash_inflows=cash_inflows,
                         cash_outflows=cash_outflows,
                         total_inflow=total_inflow,
                         total_outflow=total_outflow,
                         net_cash_flow=net_cash_flow,
                         start_date=start_date,
                         end_date=end_date)


@bp.route('/reports/cash-flow/export-excel')
@login_required
def export_cash_flow_excel():
    """Export Cash Flow to Excel"""
    from flask import make_response
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from io import BytesIO
    
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    if not start_date or not end_date:
        end_date = date.today()
        start_date = end_date - timedelta(days=30)
    else:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    cash_inflows = Payment.query.filter(
        Payment.payment_date >= start_date,
        Payment.payment_date <= end_date,
        Payment.status == 'cleared'
    ).all()
    total_inflow = sum(p.amount for p in cash_inflows)
    
    cash_outflows = Expense.query.filter(
        Expense.expense_date >= start_date,
        Expense.expense_date <= end_date,
        Expense.payment_status == 'paid'
    ).all()
    total_outflow = sum(e.total_amount for e in cash_outflows)
    net_cash_flow = total_inflow - total_outflow
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cash Flow"
    
    header_fill = PatternFill(start_color="D00000", end_color="D00000", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=16, color="D00000")
    bold_font = Font(bold=True)
    
    ws.merge_cells('A1:D1')
    ws['A1'] = 'MOHI INDUSTRIES - CASH FLOW STATEMENT'
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:D2')
    ws['A2'] = f'Period: {start_date.strftime("%d-%m-%Y")} to {end_date.strftime("%d-%m-%Y")}'
    ws['A2'].alignment = Alignment(horizontal='center')
    
    row = 4
    ws[f'A{row}'] = 'CASH INFLOWS (Receipts)'
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    
    row += 1
    ws.cell(row=row, column=1).value = 'Date'
    ws.cell(row=row, column=2).value = 'Payment #'
    ws.cell(row=row, column=3).value = 'Customer'
    ws.cell(row=row, column=4).value = 'Amount'
    for col in range(1, 5):
        ws.cell(row=row, column=col).font = bold_font
    
    row += 1
    for payment in cash_inflows:
        ws.cell(row=row, column=1).value = payment.payment_date.strftime('%d-%m-%Y')
        ws.cell(row=row, column=2).value = payment.payment_number
        ws.cell(row=row, column=3).value = payment.order.distributor.business_name if payment.order else ''
        ws.cell(row=row, column=4).value = float(payment.amount)
        row += 1
    
    ws.cell(row=row, column=3).value = 'Total Inflows:'
    ws.cell(row=row, column=3).font = bold_font
    ws.cell(row=row, column=4).value = total_inflow
    ws.cell(row=row, column=4).font = Font(bold=True, color="D00000")
    
    row += 2
    ws[f'A{row}'] = 'CASH OUTFLOWS (Expenses)'
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    
    row += 1
    ws.cell(row=row, column=1).value = 'Date'
    ws.cell(row=row, column=2).value = 'Expense #'
    ws.cell(row=row, column=3).value = 'Description'
    ws.cell(row=row, column=4).value = 'Amount'
    for col in range(1, 5):
        ws.cell(row=row, column=col).font = bold_font
    
    row += 1
    for expense in cash_outflows:
        ws.cell(row=row, column=1).value = expense.expense_date.strftime('%d-%m-%Y')
        ws.cell(row=row, column=2).value = expense.expense_number
        ws.cell(row=row, column=3).value = expense.description or expense.vendor_name
        ws.cell(row=row, column=4).value = float(expense.total_amount)
        row += 1
    
    ws.cell(row=row, column=3).value = 'Total Outflows:'
    ws.cell(row=row, column=3).font = bold_font
    ws.cell(row=row, column=4).value = total_outflow
    ws.cell(row=row, column=4).font = Font(bold=True, color="D00000")
    
    row += 2
    ws.cell(row=row, column=3).value = 'NET CASH FLOW:'
    ws.cell(row=row, column=3).font = Font(bold=True, size=12, color="D00000")
    ws.cell(row=row, column=4).value = net_cash_flow
    ws.cell(row=row, column=4).font = Font(bold=True, size=12, color="D00000")
    
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 15
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = make_response(output.read())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=Cash_Flow_{start_date.strftime("%Y%m%d")}.xlsx'
    
    return response


@bp.route('/reports/cash-flow/print')
@login_required
def print_cash_flow():
    """Print Cash Flow"""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    if not start_date or not end_date:
        end_date = date.today()
        start_date = end_date - timedelta(days=30)
    else:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    cash_inflows = Payment.query.filter(
        Payment.payment_date >= start_date,
        Payment.payment_date <= end_date,
        Payment.status == 'cleared'
    ).all()
    total_inflow = sum(p.amount for p in cash_inflows)
    
    cash_outflows = Expense.query.filter(
        Expense.expense_date >= start_date,
        Expense.expense_date <= end_date,
        Expense.payment_status == 'paid'
    ).all()
    total_outflow = sum(e.total_amount for e in cash_outflows)
    net_cash_flow = total_inflow - total_outflow
    
    return render_template(
        'accounting/cash_flow_print.html',
        cash_inflows=cash_inflows,
        cash_outflows=cash_outflows,
        total_inflow=total_inflow,
        total_outflow=total_outflow,
        net_cash_flow=net_cash_flow,
        start_date=start_date,
        end_date=end_date,
        company_name=_company_name(),
        logo_data_uri=_logo_data_uri(),
    )



# ==================== DASHBOARD ====================

@bp.route('/dashboard')
@login_required
def dashboard():
    """Accounting Dashboard"""
    today = date.today()
    
    # Today's transactions count
    today_entries = AccountingEntry.query.filter(
        AccountingEntry.entry_date == today
    ).count()
    
    # This month's income
    month_start = date(today.year, today.month, 1)
    month_income = db.session.query(
        func.sum(AccountingEntry.credit - AccountingEntry.debit)
    ).join(ChartOfAccounts, AccountingEntry.account_id == ChartOfAccounts.id
    ).filter(
        ChartOfAccounts.root_type == 'Income',
        AccountingEntry.entry_date >= month_start
    ).scalar() or 0.0
    
    # This month's expenses
    month_expenses = db.session.query(
        func.sum(AccountingEntry.debit - AccountingEntry.credit)
    ).join(ChartOfAccounts, AccountingEntry.account_id == ChartOfAccounts.id
    ).filter(
        ChartOfAccounts.root_type == 'Expense',
        AccountingEntry.entry_date >= month_start
    ).scalar() or 0.0
    
    # Pending expenses
    pending_expenses = Expense.query.filter_by(approval_status='pending').count()
    
    # Cash balance
    cash_account = ChartOfAccounts.query.filter_by(account_type='Cash').first()
    cash_balance = cash_account.get_balance() if cash_account else 0.0
    
    # Bank balance
    bank_account = ChartOfAccounts.query.filter_by(account_type='Bank').first()
    bank_balance = bank_account.get_balance() if bank_account else 0.0
    
    return render_template('accounting/dashboard.html',
                         today_entries=today_entries,
                         month_income=month_income,
                         month_expenses=month_expenses,
                         pending_expenses=pending_expenses,
                         cash_balance=cash_balance,
                         bank_balance=bank_balance)
