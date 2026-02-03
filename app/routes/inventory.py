"""
Inventory Management Routes - Complete CRUD
"""
from flask import Blueprint, current_app, render_template, request, redirect, url_for, flash, send_file
from flask_login import login_required
from app import db
from app.models import Product, ProductCategory, Inventory, Batch, Warehouse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO
from datetime import datetime

bp = Blueprint('inventory', __name__, url_prefix='/inventory')

@bp.route('/')
@login_required
def list_inventory():
    """List inventory across all warehouses"""
    inventory = Inventory.query.join(Product).join(Warehouse).all()
    return render_template('inventory/list.html', inventory=inventory)

@bp.route('/export-excel')
@login_required
def export_inventory_excel():
    """Export inventory to Excel"""
    inventory = Inventory.query.join(Product).join(Warehouse).order_by(Product.sku).all()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"
    
    # Header styling
    header_fill = PatternFill(start_color="D00000", end_color="D00000", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = [
        'SKU', 'Product Name', 'Warehouse', 'Location',
        'Total Quantity', 'Reserved', 'Available',
        'Min Stock', 'Reorder Level', 'Status'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Data rows
    for row_idx, inv in enumerate(inventory, 2):
        ws.cell(row=row_idx, column=1, value=inv.product.sku)
        ws.cell(row=row_idx, column=2, value=inv.product.name)
        ws.cell(row=row_idx, column=3, value=inv.warehouse.name)
        ws.cell(row=row_idx, column=4, value=inv.warehouse.location)
        ws.cell(row=row_idx, column=5, value=inv.quantity)
        ws.cell(row=row_idx, column=6, value=inv.reserved_quantity)
        ws.cell(row=row_idx, column=7, value=inv.available_quantity)
        ws.cell(row=row_idx, column=8, value=inv.product.min_stock_level)
        ws.cell(row=row_idx, column=9, value=inv.product.reorder_level)
        
        # Status based on stock level
        if inv.available_quantity <= 0:
            status = 'Out of Stock'
        elif inv.available_quantity <= inv.product.min_stock_level:
            status = 'Low Stock'
        elif inv.available_quantity <= inv.product.reorder_level:
            status = 'Reorder'
        else:
            status = 'In Stock'
        ws.cell(row=row_idx, column=10, value=status)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 15
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f'inventory_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@bp.route('/products')
@login_required
def list_products():
    """List all products"""
    products = Product.query.filter_by(is_active=True).all()
    return render_template('inventory/products.html', products=products)

@bp.route('/products/export-excel')
@login_required
def export_products_excel():
    """Export products to Excel"""
    products = Product.query.filter_by(is_active=True).order_by(Product.sku).all()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"
    
    # Header styling
    header_fill = PatternFill(start_color="D00000", end_color="D00000", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = [
        'SKU', 'Product Name', 'Category', 'Unit', 'Pack Size',
        'HSN Code', 'GST %', 'MRP', 'Base Price', 'Cost Price',
        'Min Stock', 'Reorder Level', 'Shelf Life (Days)', 'Status'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Data rows
    for row_idx, prod in enumerate(products, 2):
        ws.cell(row=row_idx, column=1, value=prod.sku)
        ws.cell(row=row_idx, column=2, value=prod.name)
        ws.cell(row=row_idx, column=3, value=prod.category.name if prod.category else '')
        ws.cell(row=row_idx, column=4, value=prod.unit)
        ws.cell(row=row_idx, column=5, value=prod.pack_size)
        ws.cell(row=row_idx, column=6, value=prod.hsn_code)
        ws.cell(row=row_idx, column=7, value=prod.gst_rate)
        ws.cell(row=row_idx, column=8, value=prod.mrp)
        ws.cell(row=row_idx, column=9, value=prod.base_price)
        ws.cell(row=row_idx, column=10, value=prod.cost_price)
        ws.cell(row=row_idx, column=11, value=prod.min_stock_level)
        ws.cell(row=row_idx, column=12, value=prod.reorder_level)
        ws.cell(row=row_idx, column=13, value=prod.shelf_life_days)
        ws.cell(row=row_idx, column=14, value='Active' if prod.is_active else 'Inactive')
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 12
    ws.column_dimensions['K'].width = 12
    ws.column_dimensions['L'].width = 12
    ws.column_dimensions['M'].width = 15
    ws.column_dimensions['N'].width = 10
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f'products_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@bp.route('/products/add', methods=['GET', 'POST'])
@login_required
def add_product():
    """Add new product"""
    if request.method == 'POST':
        try:
            product = Product(
                sku=request.form.get('sku'),
                name=request.form.get('name'),
                category_id=request.form.get('category_id'),
                description=request.form.get('description'),
                unit=request.form.get('unit', 'pcs'),
                pack_size=request.form.get('pack_size'),
                hsn_code=request.form.get('hsn_code'),
                gst_rate=float(request.form.get('gst_rate', 0)),
                shelf_life_days=int(request.form.get('shelf_life_days', 0)),
                mrp=float(request.form.get('mrp')),
                base_price=float(request.form.get('base_price')),
                cost_price=float(request.form.get('cost_price', 0)),
                min_stock_level=int(request.form.get('min_stock_level', 0)),
                reorder_level=int(request.form.get('reorder_level', 0)),
                requires_batch_tracking=request.form.get('requires_batch_tracking') == 'on',
                is_active=True
            )
            
            db.session.add(product)
            db.session.commit()
            
            flash(f'Product {product.sku} - {product.name} added successfully!', 'success')
            return redirect(url_for('inventory.list_products'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error adding product: {str(e)}', 'error')
    
    categories = ProductCategory.query.all()
    return render_template('inventory/product_add.html', categories=categories)

@bp.route('/products/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def edit_product(id):
    """Edit existing product"""
    product = Product.query.get_or_404(id)
    
    if request.method == 'POST':
        try:
            product.sku = request.form.get('sku')
            product.name = request.form.get('name')
            product.category_id = request.form.get('category_id')
            product.description = request.form.get('description')
            product.unit = request.form.get('unit')
            product.pack_size = request.form.get('pack_size')
            product.hsn_code = request.form.get('hsn_code')
            product.gst_rate = float(request.form.get('gst_rate', 0))
            product.shelf_life_days = int(request.form.get('shelf_life_days', 0))
            product.mrp = float(request.form.get('mrp'))
            product.base_price = float(request.form.get('base_price'))
            product.cost_price = float(request.form.get('cost_price', 0))
            product.min_stock_level = int(request.form.get('min_stock_level', 0))
            product.reorder_level = int(request.form.get('reorder_level', 0))
            product.is_active = request.form.get('is_active') == 'on'
            
            db.session.commit()
            flash(f'Product {product.sku} updated successfully!', 'success')
            return redirect(url_for('inventory.list_products'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating product: {str(e)}', 'error')
    
    categories = ProductCategory.query.all()
    return render_template('inventory/product_edit.html', product=product, categories=categories)

@bp.route('/products/<int:id>/delete', methods=['POST'])
@login_required
def delete_product(id):
    """Delete product (soft delete - set inactive)"""
    product = Product.query.get_or_404(id)
    
    try:
        # Soft delete - just set inactive
        product.is_active = False
        db.session.commit()
        flash(f'Product {product.sku} deactivated successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deactivating product: {str(e)}', 'error')
    
    return redirect(url_for('inventory.list_products'))

@bp.route('/products/<int:id>/print-label')
@login_required
def print_product_label(id):
    """Print professional retail product label with barcode"""
    product = Product.query.get_or_404(id)
    
    # Get latest batch if available
    batch = Batch.query.filter_by(product_id=id).order_by(Batch.manufacturing_date.desc()).first()
    
    label_image = None
    label_error = None
    barcode_lib_available = False

    # Try to import barcode generator
    try:
        from app.utils.barcode_generator import BarcodeGenerator
        from io import BytesIO
        import base64

        barcode_lib_available = True

        # Generate professional label
        label_img = BarcodeGenerator.generate_product_label(product, batch, label_size=(400, 600))

        # Convert to base64 for embedding in HTML
        img_io = BytesIO()
        label_img.save(img_io, 'PNG')
        img_io.seek(0)
        label_image = base64.b64encode(img_io.read()).decode()
    except ImportError:
        label_error = 'Barcode libraries are not installed on the server.'
    except Exception as e:
        label_error = f'Error generating label: {str(e)}'

    default_company_prefix = str(current_app.config.get('BARCODE_COMPANY_PREFIX', '890123456'))
    return_url = url_for('inventory.print_product_label', id=product.id)

    return render_template(
        'inventory/product_label_print.html',
        product=product,
        batch=batch,
        label_image=label_image,
        label_error=label_error,
        barcode_missing=not bool(product.ean_barcode),
        barcode_lib_available=barcode_lib_available,
        default_company_prefix=default_company_prefix,
        return_url=return_url,
    )

@bp.route('/batches')
@login_required
def list_batches():
    """List all batches"""
    batches = Batch.query.join(Product).order_by(Batch.expiry_date).all()
    return render_template('inventory/batches.html', batches=batches)

@bp.route('/batches/export-excel')
@login_required
def export_batches_excel():
    """Export all batches to Excel"""
    batches = Batch.query.join(Product).order_by(Batch.batch_number).all()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Batches"
    
    # Header styling
    header_fill = PatternFill(start_color="D00000", end_color="D00000", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = [
        'Batch Number', 'Product', 'Mfg Date', 'Expiry Date',
        'Qty Produced', 'Qty Available', 'Warehouse', 'QC Status'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Data rows
    for row_idx, batch in enumerate(batches, 2):
        ws.cell(row=row_idx, column=1, value=batch.batch_number)
        ws.cell(row=row_idx, column=2, value=batch.product.name)
        ws.cell(row=row_idx, column=3, value=batch.manufacturing_date.strftime('%d-%m-%Y'))
        ws.cell(row=row_idx, column=4, value=batch.expiry_date.strftime('%d-%m-%Y'))
        ws.cell(row=row_idx, column=5, value=batch.quantity_produced)
        ws.cell(row=row_idx, column=6, value=batch.quantity_available)
        ws.cell(row=row_idx, column=7, value=batch.warehouse.name if batch.warehouse else '')
        ws.cell(row=row_idx, column=8, value=batch.qc_status)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 12
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f'batches_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@bp.route('/batches/expiring')
@login_required
def expiring_batches():
    """Show batches expiring in next 30 days"""
    from datetime import date, timedelta
    expiry_threshold = date.today() + timedelta(days=30)
    
    batches = Batch.query.filter(
        Batch.expiry_date <= expiry_threshold,
        Batch.quantity_available > 0
    ).order_by(Batch.expiry_date).all()
    
    return render_template('inventory/expiring.html', batches=batches)

# ============ INVENTORY CRUD ============

@bp.route('/stock/add', methods=['GET', 'POST'])
@login_required
def add_inventory():
    """Add inventory stock entry"""
    if request.method == 'POST':
        try:
            product_id = request.form.get('product_id')
            warehouse_id = request.form.get('warehouse_id')
            
            # Check if inventory entry already exists
            existing = Inventory.query.filter_by(
                product_id=product_id,
                warehouse_id=warehouse_id
            ).first()
            
            if existing:
                flash('Inventory entry already exists for this product-warehouse combination!', 'error')
                return redirect(url_for('inventory.add_inventory'))
            
            inventory = Inventory(
                product_id=product_id,
                warehouse_id=warehouse_id,
                quantity=int(request.form.get('quantity', 0)),
                reserved_quantity=int(request.form.get('reserved_quantity', 0)),
                available_quantity=int(request.form.get('quantity', 0)) - int(request.form.get('reserved_quantity', 0))
            )
            
            db.session.add(inventory)
            db.session.commit()
            
            flash('Inventory entry added successfully!', 'success')
            return redirect(url_for('inventory.list_inventory'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error adding inventory: {str(e)}', 'error')
    
    products = Product.query.filter_by(is_active=True).all()
    warehouses = Warehouse.query.filter_by(is_active=True).all()
    return render_template('inventory/stock_add.html', products=products, warehouses=warehouses)

@bp.route('/stock/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def edit_inventory(id):
    """Edit inventory stock entry"""
    inventory = Inventory.query.get_or_404(id)
    
    if request.method == 'POST':
        try:
            inventory.quantity = int(request.form.get('quantity', 0))
            inventory.reserved_quantity = int(request.form.get('reserved_quantity', 0))
            inventory.available_quantity = inventory.quantity - inventory.reserved_quantity
            
            db.session.commit()
            flash('Inventory updated successfully!', 'success')
            return redirect(url_for('inventory.list_inventory'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating inventory: {str(e)}', 'error')
    
    return render_template('inventory/stock_edit.html', inventory=inventory)

@bp.route('/stock/<int:id>/delete', methods=['POST'])
@login_required
def delete_inventory(id):
    """Delete inventory entry"""
    inventory = Inventory.query.get_or_404(id)
    
    try:
        db.session.delete(inventory)
        db.session.commit()
        flash('Inventory entry deleted successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting inventory: {str(e)}', 'error')
    
    return redirect(url_for('inventory.list_inventory'))

# ============ BATCH CRUD ============

@bp.route('/batches/add', methods=['GET', 'POST'])
@login_required
def add_batch():
    """Add new batch"""
    if request.method == 'POST':
        try:
            from datetime import datetime
            
            batch = Batch(
                batch_number=request.form.get('batch_number'),
                product_id=request.form.get('product_id'),
                manufacturing_date=datetime.strptime(request.form.get('manufacturing_date'), '%Y-%m-%d').date(),
                expiry_date=datetime.strptime(request.form.get('expiry_date'), '%Y-%m-%d').date(),
                quantity_produced=int(request.form.get('quantity_produced')),
                quantity_available=int(request.form.get('quantity_produced')),
                warehouse_id=request.form.get('warehouse_id'),
                qc_status=request.form.get('qc_status', 'pending')
            )
            
            db.session.add(batch)
            db.session.commit()
            
            flash(f'Batch {batch.batch_number} added successfully!', 'success')
            return redirect(url_for('inventory.list_batches'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error adding batch: {str(e)}', 'error')
    
    products = Product.query.filter_by(is_active=True, requires_batch_tracking=True).all()
    warehouses = Warehouse.query.filter_by(is_active=True).all()
    from datetime import date
    today = date.today().strftime('%Y-%m-%d')
    return render_template('inventory/batch_add.html', products=products, warehouses=warehouses, today=today)

@bp.route('/batches/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def edit_batch(id):
    """Edit batch"""
    batch = Batch.query.get_or_404(id)
    
    if request.method == 'POST':
        try:
            from datetime import datetime
            
            batch.batch_number = request.form.get('batch_number')
            batch.manufacturing_date = datetime.strptime(request.form.get('manufacturing_date'), '%Y-%m-%d').date()
            batch.expiry_date = datetime.strptime(request.form.get('expiry_date'), '%Y-%m-%d').date()
            batch.quantity_available = int(request.form.get('quantity_available'))
            batch.qc_status = request.form.get('qc_status')
            batch.qc_remarks = request.form.get('qc_remarks')
            
            db.session.commit()
            flash(f'Batch {batch.batch_number} updated successfully!', 'success')
            return redirect(url_for('inventory.list_batches'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating batch: {str(e)}', 'error')
    
    warehouses = Warehouse.query.filter_by(is_active=True).all()
    return render_template('inventory/batch_edit.html', batch=batch, warehouses=warehouses)

@bp.route('/batches/<int:id>/delete', methods=['POST'])
@login_required
def delete_batch(id):
    """Delete batch"""
    batch = Batch.query.get_or_404(id)
    
    try:
        batch_number = batch.batch_number
        db.session.delete(batch)
        db.session.commit()
        flash(f'Batch {batch_number} deleted successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting batch: {str(e)}', 'error')
    
    return redirect(url_for('inventory.list_batches'))

# ============ WAREHOUSE CRUD ============

@bp.route('/warehouses')
@login_required
def list_warehouses():
    """List all warehouses"""
    warehouses = Warehouse.query.all()
    return render_template('inventory/warehouses.html', warehouses=warehouses)

@bp.route('/warehouses/export-excel')
@login_required
def export_warehouses_excel():
    """Export all warehouses to Excel"""
    warehouses = Warehouse.query.order_by(Warehouse.code).all()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Warehouses"
    
    # Header styling
    header_fill = PatternFill(start_color="D00000", end_color="D00000", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = ['Code', 'Name', 'Location', 'City', 'State', 'Status']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Data rows
    for row_idx, wh in enumerate(warehouses, 2):
        ws.cell(row=row_idx, column=1, value=wh.code)
        ws.cell(row=row_idx, column=2, value=wh.name)
        ws.cell(row=row_idx, column=3, value=wh.location)
        ws.cell(row=row_idx, column=4, value=wh.city)
        ws.cell(row=row_idx, column=5, value=wh.state)
        ws.cell(row=row_idx, column=6, value='Active' if wh.is_active else 'Inactive')
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 10
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f'warehouses_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@bp.route('/warehouses/add', methods=['GET', 'POST'])
@login_required
def add_warehouse():
    """Add new warehouse"""
    if request.method == 'POST':
        try:
            warehouse = Warehouse(
                code=request.form.get('code'),
                name=request.form.get('name'),
                location=request.form.get('location'),
                city=request.form.get('city'),
                state=request.form.get('state'),
                is_active=True
            )
            
            db.session.add(warehouse)
            db.session.commit()
            
            flash(f'Warehouse {warehouse.code} - {warehouse.name} added successfully!', 'success')
            return redirect(url_for('inventory.list_warehouses'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error adding warehouse: {str(e)}', 'error')
    
    return render_template('inventory/warehouse_add.html')

@bp.route('/warehouses/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def edit_warehouse(id):
    """Edit warehouse"""
    warehouse = Warehouse.query.get_or_404(id)
    
    if request.method == 'POST':
        try:
            warehouse.code = request.form.get('code')
            warehouse.name = request.form.get('name')
            warehouse.location = request.form.get('location')
            warehouse.city = request.form.get('city')
            warehouse.state = request.form.get('state')
            warehouse.is_active = request.form.get('is_active') == 'on'
            
            db.session.commit()
            flash(f'Warehouse {warehouse.code} updated successfully!', 'success')
            return redirect(url_for('inventory.list_warehouses'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating warehouse: {str(e)}', 'error')
    
    return render_template('inventory/warehouse_edit.html', warehouse=warehouse)

@bp.route('/warehouses/<int:id>/delete', methods=['POST'])
@login_required
def delete_warehouse(id):
    """Delete warehouse"""
    warehouse = Warehouse.query.get_or_404(id)
    
    try:
        # Check if warehouse has inventory
        if warehouse.inventory.count() > 0:
            flash('Cannot delete warehouse with existing inventory!', 'error')
            return redirect(url_for('inventory.list_warehouses'))
        
        warehouse_code = warehouse.code
        db.session.delete(warehouse)
        db.session.commit()
        flash(f'Warehouse {warehouse_code} deleted successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting warehouse: {str(e)}', 'error')
    
    return redirect(url_for('inventory.list_warehouses'))
