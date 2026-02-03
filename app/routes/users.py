"""
User Management Routes - Admin Only
"""
from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file
from flask_login import login_required, current_user
from functools import wraps
from app import db
from app.models import User
from datetime import datetime
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO

bp = Blueprint('users', __name__, url_prefix='/users')

def validate_password(password):
    if not password or len(password) < 8:
        return False, 'Password must be at least 8 characters long.'
    if not re.search(r'[a-z]', password):
        return False, 'Password must include a lowercase letter.'
    if not re.search(r'[A-Z]', password):
        return False, 'Password must include an uppercase letter.'
    if not re.search(r'\d', password):
        return False, 'Password must include a number.'
    if not re.search(r'[^A-Za-z0-9]', password):
        return False, 'Password must include a special character.'
    return True, ''

def admin_required(f):
    """Decorator to require admin role"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or current_user.role != 'admin':
            flash('You need administrator privileges to access this page.', 'error')
            return redirect(url_for('main.dashboard'))
        return f(*args, **kwargs)
    return decorated_function

@bp.route('/')
@login_required
@admin_required
def list_users():
    """List all users"""
    users = User.query.order_by(User.created_at.desc()).all()
    return render_template('users/list.html', users=users)

@bp.route('/export-excel')
@login_required
@admin_required
def export_users_excel():
    """Export all users to Excel"""
    users = User.query.order_by(User.username).all()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Users"
    
    # Header styling
    header_fill = PatternFill(start_color="D00000", end_color="D00000", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = ['Username', 'Email', 'Full Name', 'Role', 'Status', 'Created At', 'Last Login']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Data rows
    for row_idx, user in enumerate(users, 2):
        ws.cell(row=row_idx, column=1, value=user.username)
        ws.cell(row=row_idx, column=2, value=user.email or '')
        ws.cell(row=row_idx, column=3, value=user.full_name or '')
        ws.cell(row=row_idx, column=4, value=user.role)
        ws.cell(row=row_idx, column=5, value='Active' if user.is_active else 'Inactive')
        ws.cell(row=row_idx, column=6, value=user.created_at.strftime('%d-%m-%Y %H:%M') if user.created_at else '')
        ws.cell(row=row_idx, column=7, value=user.last_login.strftime('%d-%m-%Y %H:%M') if user.last_login else '')
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 18
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f'users_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@bp.route('/add', methods=['GET', 'POST'])
@login_required
@admin_required
def add_user():
    """Add new user"""
    if request.method == 'POST':
        try:
            # Check if username already exists
            existing_user = User.query.filter_by(username=request.form.get('username')).first()
            if existing_user:
                flash('Username already exists!', 'error')
                return redirect(url_for('users.add_user'))
            
            # Check if email already exists
            existing_email = User.query.filter_by(email=request.form.get('email')).first()
            if existing_email:
                flash('Email already exists!', 'error')
                return redirect(url_for('users.add_user'))
            
            user = User(
                username=request.form.get('username'),
                email=request.form.get('email'),
                full_name=request.form.get('full_name'),
                role=request.form.get('role', 'user'),
                is_active=True
            )
            
            password = request.form.get('password')
            is_valid, message = validate_password(password)
            if not is_valid:
                flash(message, 'error')
                return redirect(url_for('users.add_user'))
            
            user.set_password(password)
            
            db.session.add(user)
            db.session.commit()
            
            flash(f'User {user.username} created successfully!', 'success')
            return redirect(url_for('users.list_users'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error creating user: {str(e)}', 'error')
    
    return render_template('users/add.html')

@bp.route('/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_user(id):
    """Edit user"""
    user = User.query.get_or_404(id)
    
    if request.method == 'POST':
        try:
            # Check if username is taken by another user
            existing_user = User.query.filter(
                User.username == request.form.get('username'),
                User.id != id
            ).first()
            if existing_user:
                flash('Username already exists!', 'error')
                return redirect(url_for('users.edit_user', id=id))
            
            # Check if email is taken by another user
            existing_email = User.query.filter(
                User.email == request.form.get('email'),
                User.id != id
            ).first()
            if existing_email:
                flash('Email already exists!', 'error')
                return redirect(url_for('users.edit_user', id=id))
            
            user.username = request.form.get('username')
            user.email = request.form.get('email')
            user.full_name = request.form.get('full_name')
            user.role = request.form.get('role')
            user.is_active = request.form.get('is_active') == 'on'
            
            # Only update password if provided
            new_password = request.form.get('password')
            if new_password:
                is_valid, message = validate_password(new_password)
                if not is_valid:
                    flash(message, 'error')
                    return redirect(url_for('users.edit_user', id=id))
                user.set_password(new_password)
            
            db.session.commit()
            flash(f'User {user.username} updated successfully!', 'success')
            return redirect(url_for('users.list_users'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating user: {str(e)}', 'error')
    
    return render_template('users/edit.html', user=user)

@bp.route('/<int:id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_user(id):
    """Delete user"""
    user = User.query.get_or_404(id)
    
    # Prevent deleting yourself
    if user.id == current_user.id:
        flash('You cannot delete your own account!', 'error')
        return redirect(url_for('users.list_users'))
    
    # Prevent deleting the last admin
    if user.role == 'admin':
        admin_count = User.query.filter_by(role='admin', is_active=True).count()
        if admin_count <= 1:
            flash('Cannot delete the last admin user!', 'error')
            return redirect(url_for('users.list_users'))
    
    try:
        username = user.username
        db.session.delete(user)
        db.session.commit()
        flash(f'User {username} deleted successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting user: {str(e)}', 'error')
    
    return redirect(url_for('users.list_users'))

@bp.route('/profile', methods=['GET', 'POST'])
@login_required
def profile():
    """User profile - change own password"""
    if request.method == 'POST':
        try:
            current_password = request.form.get('current_password')
            new_password = request.form.get('new_password')
            confirm_password = request.form.get('confirm_password')
            
            # Verify current password
            if not current_user.check_password(current_password):
                flash('Current password is incorrect!', 'error')
                return redirect(url_for('users.profile'))
            
            # Check if new passwords match
            if new_password != confirm_password:
                flash('New passwords do not match!', 'error')
                return redirect(url_for('users.profile'))
            
            is_valid, message = validate_password(new_password)
            if not is_valid:
                flash(message, 'error')
                return redirect(url_for('users.profile'))
            
            # Update password
            current_user.set_password(new_password)
            db.session.commit()
            
            flash('Password changed successfully!', 'success')
            return redirect(url_for('main.dashboard'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error changing password: {str(e)}', 'error')
    
    return render_template('users/profile.html')
