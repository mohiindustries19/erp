"""
Goods Inventory Routes - Complete CRUD with Export, Email, WhatsApp
"""
from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file, jsonify
from flask_login import login_required, current_user
from app import db
from app.models import Goods
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO
from datetime import datetime, date

bp = Blueprint('goods', __name__, url_prefix='/goods')

@bp.route('/')
@login_required
def list_goods():
    """List all goods"""
    goods = Goods.query.filter_by(is_active=True).order_by(Goods.item_number).all()
    return render_template('goods/list.html', goods=goods)

@bp.route('/add', methods=['GET', 'POST'])
@login_required
def add_goods():
    """Add new goods item"""
    if request.method == 'POST':
        try:
            # Parse purchase date
            purchase_date_str = request.form.get('purchase_date')
            purchase_date = None
            if purchase_date_str:
                purchase_date = datetime.strptime(purchase_date_str, '%Y-%m-%d').date()
            
            goods = Goods(
                item_number=int(request.form.get('item_number')),
                name=request.form.get('name'),
                category=request.form.get('category'),
                quantity=int(request.form.get('quantity', 0)),
                unit=request.form.get('unit', 'pcs'),
                location=request.form.get('location'),
                condition=request.form.get('condition', 'good'),
                purchase_date=purchase_date,
                purchase_price=float(request.form.get('purchase_price', 0)) if request.form.get('purchase_price') else None,
                supplier=request.form.get('supplier'),
                notes=request.form.get('notes'),
                is_active=True
            )
            
            db.session.add(goods)
            db.session.commit()
            
            flash(f'✅ Goods item #{goods.item_number} - {goods.name} added successfully!', 'success')
            return redirect(url_for('goods.list_goods'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'❌ Error adding goods: {str(e)}', 'error')
    
    return render_template('goods/add.html')

@bp.route('/<int:id>')
@login_required
def view_goods(id):
    """View goods details"""
    goods = Goods.query.get_or_404(id)
    return render_template('goods/view.html', goods=goods)

@bp.route('/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def edit_goods(id):
    """Edit goods item"""
    goods = Goods.query.get_or_404(id)
    
    if request.method == 'POST':
        try:
            # Parse purchase date
            purchase_date_str = request.form.get('purchase_date')
            purchase_date = None
            if purchase_date_str:
                purchase_date = datetime.strptime(purchase_date_str, '%Y-%m-%d').date()
            
            goods.item_number = int(request.form.get('item_number'))
            goods.name = request.form.get('name')
            goods.category = request.form.get('category')
            goods.quantity = int(request.form.get('quantity', 0))
            goods.unit = request.form.get('unit', 'pcs')
            goods.location = request.form.get('location')
            goods.condition = request.form.get('condition', 'good')
            goods.purchase_date = purchase_date
            goods.purchase_price = float(request.form.get('purchase_price', 0)) if request.form.get('purchase_price') else None
            goods.supplier = request.form.get('supplier')
            goods.notes = request.form.get('notes')
            goods.is_active = request.form.get('is_active') == 'on'
            
            db.session.commit()
            flash(f'✅ Goods item #{goods.item_number} updated successfully!', 'success')
            return redirect(url_for('goods.list_goods'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'❌ Error updating goods: {str(e)}', 'error')
    
    return render_template('goods/edit.html', goods=goods)

@bp.route('/<int:id>/delete', methods=['POST'])
@login_required
def delete_goods(id):
    """Delete goods item (soft delete)"""
    goods = Goods.query.get_or_404(id)
    
    try:
        goods.is_active = False
        db.session.commit()
        flash(f'✅ Goods item #{goods.item_number} - {goods.name} deleted successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'❌ Error deleting goods: {str(e)}', 'error')
    
    return redirect(url_for('goods.list_goods'))

@bp.route('/export-excel')
@login_required
def export_excel():
    """Export goods inventory to Excel"""
    goods_list = Goods.query.filter_by(is_active=True).order_by(Goods.item_number).all()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Goods Inventory"
    
    # Header styling
    header_fill = PatternFill(start_color="D00000", end_color="D00000", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = [
        'Item #', 'Name', 'Category', 'Quantity', 'Unit',
        'Location', 'Condition', 'Purchase Date', 'Purchase Price', 'Supplier', 'Notes'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Data rows
    for row_idx, goods in enumerate(goods_list, 2):
        ws.cell(row=row_idx, column=1, value=goods.item_number)
        ws.cell(row=row_idx, column=2, value=goods.name)
        ws.cell(row=row_idx, column=3, value=goods.category)
        ws.cell(row=row_idx, column=4, value=goods.quantity)
        ws.cell(row=row_idx, column=5, value=goods.unit)
        ws.cell(row=row_idx, column=6, value=goods.location)
        ws.cell(row=row_idx, column=7, value=goods.condition)
        ws.cell(row=row_idx, column=8, value=goods.purchase_date.strftime('%d-%m-%Y') if goods.purchase_date else '')
        ws.cell(row=row_idx, column=9, value=goods.purchase_price)
        ws.cell(row=row_idx, column=10, value=goods.supplier)
        ws.cell(row=row_idx, column=11, value=goods.notes)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 25
    ws.column_dimensions['K'].width = 30
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f'goods_inventory_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@bp.route('/print')
@login_required
def print_goods():
    """Print goods inventory"""
    goods_list = Goods.query.filter_by(is_active=True).order_by(Goods.item_number).all()
    return render_template('goods/print.html', goods=goods_list)

@bp.route('/email', methods=['POST'])
@login_required
def email_goods():
    """Email goods inventory"""
    try:
        from flask_mail import Message
        from app import mail
        
        recipient = request.form.get('email')
        if not recipient:
            return jsonify({'success': False, 'message': 'Email address required'}), 400
        
        # Generate Excel file
        goods_list = Goods.query.filter_by(is_active=True).order_by(Goods.item_number).all()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Goods Inventory"
        
        # Headers
        headers = ['Item #', 'Name', 'Category', 'Quantity', 'Unit', 'Location', 'Condition']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Data
        for row_idx, goods in enumerate(goods_list, 2):
            ws.cell(row=row_idx, column=1, value=goods.item_number)
            ws.cell(row=row_idx, column=2, value=goods.name)
            ws.cell(row=row_idx, column=3, value=goods.category)
            ws.cell(row=row_idx, column=4, value=goods.quantity)
            ws.cell(row=row_idx, column=5, value=goods.unit)
            ws.cell(row=row_idx, column=6, value=goods.location)
            ws.cell(row=row_idx, column=7, value=goods.condition)
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Send email
        msg = Message(
            subject='Goods Inventory Report - Mohi Industries',
            recipients=[recipient],
            body=f'''Dear Team,

Please find attached the Goods Inventory Report as of {datetime.now().strftime('%d-%m-%Y %H:%M')}.

Total Items: {len(goods_list)}

Best regards,
Mohi Industries ERP System
'''
        )
        
        filename = f'goods_inventory_{datetime.now().strftime("%Y%m%d")}.xlsx'
        msg.attach(filename, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', output.read())
        
        mail.send(msg)
        
        return jsonify({'success': True, 'message': f'✅ Inventory report sent to {recipient}'})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'❌ Error sending email: {str(e)}'}), 500

@bp.route('/whatsapp', methods=['POST'])
@login_required
def whatsapp_goods():
    """Send goods inventory via WhatsApp"""
    try:
        phone = request.form.get('phone')
        if not phone:
            return jsonify({'success': False, 'message': 'Phone number required'}), 400
        
        # Format phone number
        if not phone.startswith('+'):
            phone = f'+91{phone}'
        
        # Generate inventory summary
        goods_list = Goods.query.filter_by(is_active=True).order_by(Goods.item_number).all()
        
        message = f'''*Goods Inventory Report*
_Mohi Industries_
Date: {datetime.now().strftime('%d-%m-%Y %H:%M')}

Total Items: {len(goods_list)}

'''
        
        # Add top 20 items
        for goods in goods_list[:20]:
            message += f"#{goods.item_number} {goods.name}\n"
            message += f"   Qty: {goods.quantity} {goods.unit} | {goods.condition}\n\n"
        
        if len(goods_list) > 20:
            message += f"\n... and {len(goods_list) - 20} more items\n"
        
        message += "\n_Generated by Mohi Industries ERP_"
        
        # Send via WhatsApp (using Twilio)
        import os
        whatsapp_enabled = os.getenv('WHATSAPP_ENABLED', 'false').lower() == 'true'
        
        if not whatsapp_enabled:
            return jsonify({'success': False, 'message': '⚠️ WhatsApp integration not enabled'}), 400
        
        from twilio.rest import Client
        
        account_sid = os.getenv('TWILIO_ACCOUNT_SID')
        auth_token = os.getenv('TWILIO_AUTH_TOKEN')
        whatsapp_number = os.getenv('TWILIO_WHATSAPP_NUMBER', 'whatsapp:+14155238886')
        
        client = Client(account_sid, auth_token)
        
        message_obj = client.messages.create(
            body=message,
            from_=whatsapp_number,
            to=f'whatsapp:{phone}'
        )
        
        return jsonify({'success': True, 'message': f'✅ Inventory report sent to {phone} via WhatsApp'})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'❌ Error sending WhatsApp: {str(e)}'}), 500

@bp.route('/bulk-import', methods=['POST'])
@login_required
def bulk_import():
    """Bulk import goods from Excel"""
    try:
        if 'file' not in request.files:
            flash('❌ No file uploaded', 'error')
            return redirect(url_for('goods.list_goods'))
        
        file = request.files['file']
        if file.filename == '':
            flash('❌ No file selected', 'error')
            return redirect(url_for('goods.list_goods'))
        
        # Read Excel file
        from openpyxl import load_workbook
        wb = load_workbook(file)
        ws = wb.active
        
        success_count = 0
        error_count = 0
        
        # Skip header row
        for row in ws.iter_rows(min_row=2, values_only=True):
            try:
                if not row[0] or not row[1]:  # Skip if no item number or name
                    continue
                
                # Check if item already exists
                existing = Goods.query.filter_by(item_number=int(row[0])).first()
                if existing:
                    error_count += 1
                    continue
                
                goods = Goods(
                    item_number=int(row[0]),
                    name=str(row[1]),
                    category=str(row[2]) if row[2] else None,
                    quantity=int(row[3]) if row[3] else 0,
                    unit=str(row[4]) if row[4] else 'pcs',
                    location=str(row[5]) if row[5] else None,
                    condition=str(row[6]) if row[6] else 'good',
                    is_active=True
                )
                
                db.session.add(goods)
                success_count += 1
                
            except Exception as e:
                error_count += 1
                continue
        
        db.session.commit()
        
        flash(f'✅ Imported {success_count} items successfully! {error_count} errors.', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'❌ Error importing file: {str(e)}', 'error')
    
    return redirect(url_for('goods.list_goods'))
