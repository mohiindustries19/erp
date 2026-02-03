"""
Purchasing Routes - Vendors, Purchase Orders, Vendor Bills
"""
from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file
from flask_login import login_required, current_user
from app import db
from app.models import Product
from app.models.purchasing import Vendor, PurchaseOrder, PurchaseOrderItem, VendorBill, VendorBillItem, VendorPayment
from app.models.accounting import AccountingEntry, ChartOfAccounts
from datetime import datetime, date
from app.services.permissions import role_required
from sqlalchemy import func
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO

bp = Blueprint('purchasing', __name__, url_prefix='/purchasing')

COMPANY_STATE_CODE = '27'


def _gst_split_for_vendor(vendor_gstin, tax_amount):
    if not tax_amount or tax_amount <= 0:
        return 0.0, 0.0, 0.0
    if vendor_gstin and len(vendor_gstin) >= 2:
        vendor_state_code = vendor_gstin[:2]
        if vendor_state_code == COMPANY_STATE_CODE:
            cgst = round(tax_amount / 2, 2)
            sgst = round(tax_amount / 2, 2)
            return cgst, sgst, 0.0
        return 0.0, 0.0, round(tax_amount, 2)
    cgst = round(tax_amount / 2, 2)
    sgst = round(tax_amount / 2, 2)
    return cgst, sgst, 0.0


def _normalize_code(value, prefix='ACC'):
    safe = ''.join(c for c in value.upper() if c.isalnum())
    return f'{prefix}-{safe[:10]}' if safe else prefix


def _get_or_create_account(name, account_type, account_group, code_prefix):
    account = ChartOfAccounts.query.filter(func.lower(ChartOfAccounts.name) == name.lower()).first()
    if account:
        return account

    base_code = _normalize_code(name, code_prefix)
    code = base_code
    counter = 1
    while ChartOfAccounts.query.filter_by(account_number=code).first():
        counter += 1
        code = f'{base_code}-{counter}'

    root_map = {
        'asset': 'Asset',
        'liability': 'Liability',
        'income': 'Income',
        'expense': 'Expense',
        'equity': 'Equity',
    }
    root_type = root_map.get((account_type or '').lower(), 'Asset')
    account = ChartOfAccounts(
        account_number=code,
        name=name,
        root_type=root_type,
        account_type=account_type,
        description=account_group,
        is_active=True,
    )
    db.session.add(account)
    db.session.flush()
    return account


def _ensure_vendor_ap_account(vendor):
    if vendor.ap_account_id:
        account = ChartOfAccounts.query.get(vendor.ap_account_id)
        if account:
            account.name = f'Accounts Payable - {vendor.business_name}'
            return account
    account_name = f'Accounts Payable - {vendor.business_name}'
    account_code = f'AP-{vendor.code}'
    account = ChartOfAccounts.query.filter_by(account_number=account_code).first()
    if not account:
        account = ChartOfAccounts(
            account_number=account_code,
            name=account_name,
            root_type='Liability',
            account_type='Payable',
            description='Current Liabilities',
            is_active=True,
        )
        db.session.add(account)
        db.session.flush()
    vendor.ap_account_id = account.id
    return account


def create_vendor_bill_entries(bill):
    """Create accounting entries for vendor bill"""
    if bill.approval_status != 'approved':
        return
    existing = AccountingEntry.query.filter_by(
        reference_type='vendor_bill',
        reference_id=bill.id
    ).count()
    if existing:
        return

    purchases_account = _get_or_create_account('Purchases', 'expense', 'Direct Expenses', 'PUR')
    input_cgst = _get_or_create_account('Input CGST', 'asset', 'Current Assets', 'CGST')
    input_sgst = _get_or_create_account('Input SGST', 'asset', 'Current Assets', 'SGST')
    input_igst = _get_or_create_account('Input IGST', 'asset', 'Current Assets', 'IGST')
    ap_account = _ensure_vendor_ap_account(bill.vendor)

    cgst, sgst, igst = _gst_split_for_vendor(bill.vendor.gstin, bill.tax_amount)

    entries = [
        AccountingEntry(
            entry_date=bill.bill_date,
            reference_type='vendor_bill',
            reference_id=bill.id,
            account_id=purchases_account.id,
            account_head='Purchases',
            debit=bill.subtotal,
            credit=0.0,
            description=f'Vendor bill {bill.bill_number} - Purchases'
        )
    ]

    if cgst > 0:
        entries.append(AccountingEntry(
            entry_date=bill.bill_date,
            reference_type='vendor_bill',
            reference_id=bill.id,
            account_id=input_cgst.id,
            account_head='Input CGST',
            debit=cgst,
            credit=0.0,
            description=f'Vendor bill {bill.bill_number} - Input CGST'
        ))
    if sgst > 0:
        entries.append(AccountingEntry(
            entry_date=bill.bill_date,
            reference_type='vendor_bill',
            reference_id=bill.id,
            account_id=input_sgst.id,
            account_head='Input SGST',
            debit=sgst,
            credit=0.0,
            description=f'Vendor bill {bill.bill_number} - Input SGST'
        ))
    if igst > 0:
        entries.append(AccountingEntry(
            entry_date=bill.bill_date,
            reference_type='vendor_bill',
            reference_id=bill.id,
            account_id=input_igst.id,
            account_head='Input IGST',
            debit=igst,
            credit=0.0,
            description=f'Vendor bill {bill.bill_number} - Input IGST'
        ))

    entries.append(AccountingEntry(
        entry_date=bill.bill_date,
        reference_type='vendor_bill',
        reference_id=bill.id,
        account_id=ap_account.id,
        account_head='Accounts Payable',
        debit=0.0,
        credit=bill.total_amount,
        description=f'Vendor bill {bill.bill_number} - Payable'
    ))

    for entry in entries:
        db.session.add(entry)


def create_vendor_payment_entries(bill, payment):
    """Create accounting entries when vendor bill is paid"""
    existing = AccountingEntry.query.filter_by(
        reference_type='vendor_payment',
        reference_id=payment.id
    ).count()
    if existing:
        return

    amount = payment.amount
    if amount <= 0:
        return

    ap_account = _ensure_vendor_ap_account(bill.vendor)
    cash_account = _get_or_create_account('Cash', 'asset', 'Current Assets', 'CASH')
    bank_account = _get_or_create_account('Bank', 'asset', 'Current Assets', 'BANK')
    is_cash = payment.payment_mode == 'cash'
    credit_account = cash_account if is_cash else bank_account
    credit_head = 'Cash' if is_cash else 'Bank'

    db.session.add(AccountingEntry(
        entry_date=payment.payment_date,
        reference_type='vendor_payment',
        reference_id=payment.id,
        account_id=ap_account.id,
        account_head='Accounts Payable',
        debit=amount,
        credit=0.0,
        description=f'Vendor bill {bill.bill_number} - Payment'
    ))
    db.session.add(AccountingEntry(
        entry_date=payment.payment_date,
        reference_type='vendor_payment',
        reference_id=payment.id,
        account_id=credit_account.id,
        account_head=credit_head,
        debit=0.0,
        credit=amount,
        description=f'Vendor bill {bill.bill_number} - Payment'
    ))

# ==================== VENDORS ====================

@bp.route('/vendors')
@login_required
def list_vendors():
    vendors = Vendor.query.order_by(Vendor.business_name.asc()).all()
    return render_template('purchasing/vendors.html', vendors=vendors)

@bp.route('/vendors/export-excel')
@login_required
def export_vendors_excel():
    """Export all vendors to Excel"""
    vendors = Vendor.query.order_by(Vendor.code).all()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Vendors"
    
    # Header styling
    header_fill = PatternFill(start_color="D00000", end_color="D00000", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = [
        'Code', 'Business Name', 'Contact Person', 'Phone', 'Email',
        'GSTIN', 'Address', 'City', 'State', 'Payment Terms', 'Status'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Data rows
    for row_idx, vendor in enumerate(vendors, 2):
        ws.cell(row=row_idx, column=1, value=vendor.code)
        ws.cell(row=row_idx, column=2, value=vendor.business_name)
        ws.cell(row=row_idx, column=3, value=vendor.contact_person)
        ws.cell(row=row_idx, column=4, value=vendor.phone)
        ws.cell(row=row_idx, column=5, value=vendor.email)
        ws.cell(row=row_idx, column=6, value=vendor.gstin)
        ws.cell(row=row_idx, column=7, value=vendor.address)
        ws.cell(row=row_idx, column=8, value=vendor.city)
        ws.cell(row=row_idx, column=9, value=vendor.state)
        ws.cell(row=row_idx, column=10, value=vendor.payment_terms)
        ws.cell(row=row_idx, column=11, value=vendor.status)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 30
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 10
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f'vendors_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@bp.route('/vendors/add', methods=['GET', 'POST'])
@login_required
def add_vendor():
    if request.method == 'POST':
        try:
            code = request.form.get('code')
            existing = Vendor.query.filter_by(code=code).first()
            if existing:
                flash('Vendor code already exists!', 'error')
                return redirect(url_for('purchasing.add_vendor'))

            vendor = Vendor(
                code=code,
                business_name=request.form.get('business_name'),
                contact_person=request.form.get('contact_person'),
                phone=request.form.get('phone'),
                email=request.form.get('email'),
                gstin=request.form.get('gstin'),
                address=request.form.get('address'),
                city=request.form.get('city'),
                state=request.form.get('state'),
                payment_terms=request.form.get('payment_terms'),
                status=request.form.get('status', 'active')
            )

            db.session.add(vendor)
            db.session.flush()
            _ensure_vendor_ap_account(vendor)
            db.session.commit()

            flash(f'Vendor {vendor.business_name} created successfully!', 'success')
            return redirect(url_for('purchasing.list_vendors'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error creating vendor: {str(e)}', 'error')

    return render_template('purchasing/vendor_add.html')


@bp.route('/vendors/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def edit_vendor(id):
    vendor = Vendor.query.get_or_404(id)

    if request.method == 'POST':
        try:
            vendor.code = request.form.get('code')
            vendor.business_name = request.form.get('business_name')
            vendor.contact_person = request.form.get('contact_person')
            vendor.phone = request.form.get('phone')
            vendor.email = request.form.get('email')
            vendor.gstin = request.form.get('gstin')
            vendor.address = request.form.get('address')
            vendor.city = request.form.get('city')
            vendor.state = request.form.get('state')
            vendor.payment_terms = request.form.get('payment_terms')
            vendor.status = request.form.get('status', 'active')

            _ensure_vendor_ap_account(vendor)

            db.session.commit()
            flash(f'Vendor {vendor.business_name} updated successfully!', 'success')
            return redirect(url_for('purchasing.list_vendors'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating vendor: {str(e)}', 'error')

    return render_template('purchasing/vendor_edit.html', vendor=vendor)


@bp.route('/vendors/<int:id>/delete', methods=['POST'])
@login_required
def delete_vendor(id):
    vendor = Vendor.query.get_or_404(id)
    try:
        vendor.status = 'inactive'
        db.session.commit()
        flash(f'Vendor {vendor.business_name} deactivated successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting vendor: {str(e)}', 'error')

    return redirect(url_for('purchasing.list_vendors'))


# ==================== PURCHASE ORDERS ====================

@bp.route('/purchase-orders')
@login_required
def list_purchase_orders():
    orders = PurchaseOrder.query.order_by(PurchaseOrder.created_at.desc()).all()
    return render_template('purchasing/purchase_orders.html', orders=orders)

@bp.route('/purchase-orders/export-excel')
@login_required
def export_purchase_orders_excel():
    """Export all purchase orders to Excel"""
    orders = PurchaseOrder.query.order_by(PurchaseOrder.po_number).all()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Purchase Orders"
    
    # Header styling
    header_fill = PatternFill(start_color="D00000", end_color="D00000", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = [
        'PO Number', 'Date', 'Vendor', 'Subtotal', 'Tax', 'Total',
        'Status', 'Approval Status', 'Created By'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Data rows
    for row_idx, po in enumerate(orders, 2):
        ws.cell(row=row_idx, column=1, value=po.po_number)
        ws.cell(row=row_idx, column=2, value=po.po_date.strftime('%d-%m-%Y'))
        ws.cell(row=row_idx, column=3, value=po.vendor.business_name)
        ws.cell(row=row_idx, column=4, value=po.subtotal)
        ws.cell(row=row_idx, column=5, value=po.tax_amount)
        ws.cell(row=row_idx, column=6, value=po.total_amount)
        ws.cell(row=row_idx, column=7, value=po.status)
        ws.cell(row=row_idx, column=8, value=po.approval_status)
        ws.cell(row=row_idx, column=9, value=po.creator.username if po.creator else '')
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f'purchase_orders_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@bp.route('/purchase-orders/add', methods=['GET', 'POST'])
@login_required
def add_purchase_order():
    if request.method == 'POST':
        try:
            items_data = request.form.get('items_json')
            if not items_data:
                flash('Please add at least one item to the purchase order!', 'error')
                return redirect(url_for('purchasing.add_purchase_order'))

            import json
            items_data = json.loads(items_data)

            today = date.today()
            order_prefix = f'PO{today.strftime("%Y%m%d")}'
            last_order = PurchaseOrder.query.filter(
                PurchaseOrder.order_number.like(f'{order_prefix}%')
            ).order_by(PurchaseOrder.id.desc()).first()

            if last_order:
                last_num = int(last_order.order_number[-4:])
                order_number = f'{order_prefix}{last_num+1:04d}'
            else:
                order_number = f'{order_prefix}0001'

            order = PurchaseOrder(
                order_number=order_number,
                vendor_id=request.form.get('vendor_id'),
                order_date=datetime.strptime(request.form.get('order_date'), '%Y-%m-%d').date(),
                expected_date=datetime.strptime(request.form.get('expected_date'), '%Y-%m-%d').date() if request.form.get('expected_date') else None,
                status=request.form.get('status', 'draft'),
                remarks=request.form.get('remarks')
            )

            db.session.add(order)
            db.session.flush()

            added_items = 0
            for item in items_data:
                product = Product.query.get(item.get('product_id'))
                if not product:
                    continue

                quantity = int(item.get('quantity', 0))
                unit_cost = float(item.get('unit_cost', 0))

                if quantity <= 0 or unit_cost <= 0:
                    flash('Quantity and unit cost must be greater than zero.', 'error')
                    db.session.rollback()
                    return redirect(url_for('purchasing.add_purchase_order'))

                po_item = PurchaseOrderItem(
                    purchase_order_id=order.id,
                    product_id=product.id,
                    quantity=quantity,
                    unit_cost=unit_cost,
                    gst_rate=product.gst_rate or 0
                )
                po_item.calculate_line_total()
                db.session.add(po_item)
                added_items += 1

            if added_items == 0:
                flash('Please add at least one valid item to the purchase order!', 'error')
                db.session.rollback()
                return redirect(url_for('purchasing.add_purchase_order'))

            order.calculate_totals()
            db.session.commit()

            flash(f'Purchase order {order_number} created successfully!', 'success')
            return redirect(url_for('purchasing.view_purchase_order', id=order.id))
        except Exception as e:
            db.session.rollback()
            flash(f'Error creating purchase order: {str(e)}', 'error')

    vendors = Vendor.query.filter_by(status='active').all()
    products = Product.query.filter_by(is_active=True).all()
    today = date.today().strftime('%Y-%m-%d')
    return render_template('purchasing/purchase_order_add.html', vendors=vendors, products=products, today=today)


@bp.route('/purchase-orders/<int:id>')
@login_required
def view_purchase_order(id):
    order = PurchaseOrder.query.get_or_404(id)
    return render_template('purchasing/purchase_order_view.html', order=order)


@bp.route('/purchase-orders/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def edit_purchase_order(id):
    order = PurchaseOrder.query.get_or_404(id)

    if request.method == 'POST':
        try:
            order.expected_date = datetime.strptime(request.form.get('expected_date'), '%Y-%m-%d').date() if request.form.get('expected_date') else None
            order.status = request.form.get('status')
            order.remarks = request.form.get('remarks')
            db.session.commit()
            flash(f'Purchase order {order.order_number} updated successfully!', 'success')
            return redirect(url_for('purchasing.view_purchase_order', id=order.id))
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating purchase order: {str(e)}', 'error')

    return render_template('purchasing/purchase_order_edit.html', order=order)


@bp.route('/purchase-orders/<int:id>/print')
@login_required
def print_purchase_order(id):
    """Print purchase order"""
    order = PurchaseOrder.query.get_or_404(id)
    return render_template('purchasing/purchase_order_print.html', order=order)


@bp.route('/purchase-orders/<int:id>/delete', methods=['POST'])
@login_required
def delete_purchase_order(id):
    order = PurchaseOrder.query.get_or_404(id)

    if order.status not in ('draft', 'cancelled'):
        flash('Only draft/cancelled purchase orders can be deleted.', 'error')
        return redirect(url_for('purchasing.view_purchase_order', id=order.id))

    linked_bill = VendorBill.query.filter_by(purchase_order_id=order.id).first()
    if linked_bill:
        flash('Cannot delete this purchase order because a vendor bill is linked to it.', 'error')
        return redirect(url_for('purchasing.view_purchase_order', id=order.id))

    try:
        order_number = order.order_number
        db.session.delete(order)
        db.session.commit()
        flash(f'Purchase order {order_number} deleted successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting purchase order: {str(e)}', 'error')

    return redirect(url_for('purchasing.list_purchase_orders'))


# ==================== VENDOR BILLS ====================

@bp.route('/vendor-bills')
@login_required
def list_vendor_bills():
    bills = VendorBill.query.order_by(VendorBill.created_at.desc()).all()
    return render_template('purchasing/vendor_bills.html', bills=bills)

@bp.route('/vendor-bills/export-excel')
@login_required
def export_vendor_bills_excel():
    """Export all vendor bills to Excel"""
    bills = VendorBill.query.order_by(VendorBill.bill_number).all()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Vendor Bills"
    
    # Header styling
    header_fill = PatternFill(start_color="D00000", end_color="D00000", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = [
        'Bill Number', 'Bill Date', 'Vendor', 'Subtotal', 'Tax', 'Total',
        'Paid Amount', 'Balance', 'Status', 'Approval Status'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Data rows
    for row_idx, bill in enumerate(bills, 2):
        ws.cell(row=row_idx, column=1, value=bill.bill_number)
        ws.cell(row=row_idx, column=2, value=bill.bill_date.strftime('%d-%m-%Y'))
        ws.cell(row=row_idx, column=3, value=bill.vendor.business_name)
        ws.cell(row=row_idx, column=4, value=bill.subtotal)
        ws.cell(row=row_idx, column=5, value=bill.tax_amount)
        ws.cell(row=row_idx, column=6, value=bill.total_amount)
        ws.cell(row=row_idx, column=7, value=bill.paid_amount)
        ws.cell(row=row_idx, column=8, value=bill.total_amount - bill.paid_amount)
        ws.cell(row=row_idx, column=9, value=bill.payment_status)
        ws.cell(row=row_idx, column=10, value=bill.approval_status)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 15
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f'vendor_bills_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@bp.route('/vendor-bills/add', methods=['GET', 'POST'])
@login_required
def add_vendor_bill():
    if request.method == 'POST':
        try:
            items_data = request.form.get('items_json')
            if not items_data:
                flash('Please add at least one item to the vendor bill!', 'error')
                return redirect(url_for('purchasing.add_vendor_bill'))

            import json
            items_data = json.loads(items_data)

            today = date.today()
            bill_prefix = f'VB{today.strftime("%Y%m%d")}'
            last_bill = VendorBill.query.filter(
                VendorBill.bill_number.like(f'{bill_prefix}%')
            ).order_by(VendorBill.id.desc()).first()

            if last_bill:
                last_num = int(last_bill.bill_number[-4:])
                bill_number = f'{bill_prefix}{last_num+1:04d}'
            else:
                bill_number = f'{bill_prefix}0001'

            approval_status = request.form.get('approval_status', 'pending')
            bill = VendorBill(
                bill_number=bill_number,
                vendor_id=request.form.get('vendor_id'),
                purchase_order_id=request.form.get('purchase_order_id') or None,
                bill_date=datetime.strptime(request.form.get('bill_date'), '%Y-%m-%d').date(),
                due_date=datetime.strptime(request.form.get('due_date'), '%Y-%m-%d').date() if request.form.get('due_date') else None,
                status=request.form.get('status', 'pending'),
                approval_status=approval_status,
                remarks=request.form.get('remarks')
            )

            db.session.add(bill)
            db.session.flush()

            added_items = 0
            for item in items_data:
                product = Product.query.get(item.get('product_id'))
                if not product:
                    continue

                quantity = int(item.get('quantity', 0))
                unit_cost = float(item.get('unit_cost', 0))

                if quantity <= 0 or unit_cost <= 0:
                    flash('Quantity and unit cost must be greater than zero.', 'error')
                    db.session.rollback()
                    return redirect(url_for('purchasing.add_vendor_bill'))

                bill_item = VendorBillItem(
                    vendor_bill_id=bill.id,
                    product_id=product.id,
                    quantity=quantity,
                    unit_cost=unit_cost,
                    gst_rate=product.gst_rate or 0
                )
                bill_item.calculate_line_total()
                db.session.add(bill_item)
                added_items += 1

            if added_items == 0:
                flash('Please add at least one valid item to the vendor bill!', 'error')
                db.session.rollback()
                return redirect(url_for('purchasing.add_vendor_bill'))

            bill.calculate_totals()
            if bill.status == 'paid' and bill.approval_status != 'approved':
                flash('Bill must be approved before marking as paid.', 'error')
                db.session.rollback()
                return redirect(url_for('purchasing.add_vendor_bill'))

            if bill.approval_status == 'approved':
                bill.approved_by = current_user.id
                bill.approved_at = datetime.utcnow()
                create_vendor_bill_entries(bill)
            db.session.commit()

            flash(f'Vendor bill {bill_number} created successfully!', 'success')
            return redirect(url_for('purchasing.view_vendor_bill', id=bill.id))
        except Exception as e:
            db.session.rollback()
            flash(f'Error creating vendor bill: {str(e)}', 'error')

    vendors = Vendor.query.filter_by(status='active').all()
    purchase_orders = PurchaseOrder.query.order_by(PurchaseOrder.created_at.desc()).all()
    products = Product.query.filter_by(is_active=True).all()
    today = date.today().strftime('%Y-%m-%d')
    return render_template('purchasing/vendor_bill_add.html', vendors=vendors, purchase_orders=purchase_orders, products=products, today=today)


@bp.route('/vendor-bills/<int:id>')
@login_required
def view_vendor_bill(id):
    bill = VendorBill.query.get_or_404(id)
    return render_template('purchasing/vendor_bill_view.html', bill=bill)


@bp.route('/vendor-bills/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@role_required(['admin', 'manager'])
def edit_vendor_bill(id):
    bill = VendorBill.query.get_or_404(id)

    if request.method == 'POST':
        try:
            bill.due_date = datetime.strptime(request.form.get('due_date'), '%Y-%m-%d').date() if request.form.get('due_date') else None
            new_status = request.form.get('status')
            new_approval = request.form.get('approval_status', bill.approval_status)
            bill.status = new_status
            bill.approval_status = new_approval
            bill.remarks = request.form.get('remarks')
            if new_status == 'paid' and bill.approval_status != 'approved':
                flash('Bill must be approved before marking as paid.', 'error')
                return redirect(url_for('purchasing.edit_vendor_bill', id=bill.id))

            if new_approval == 'approved':
                bill.approved_by = current_user.id
                bill.approved_at = datetime.utcnow()
                create_vendor_bill_entries(bill)
            elif new_approval in ['pending', 'rejected']:
                AccountingEntry.query.filter_by(
                    reference_type='vendor_bill',
                    reference_id=bill.id
                ).delete()

            db.session.commit()
            flash(f'Vendor bill {bill.bill_number} updated successfully!', 'success')
            return redirect(url_for('purchasing.view_vendor_bill', id=bill.id))
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating vendor bill: {str(e)}', 'error')

    return render_template('purchasing/vendor_bill_edit.html', bill=bill)


@bp.route('/vendor-bills/<int:id>/pay', methods=['GET', 'POST'])
@login_required
@role_required(['admin', 'manager'])
def pay_vendor_bill(id):
    bill = VendorBill.query.get_or_404(id)
    outstanding = bill.total_amount - bill.paid_amount

    if request.method == 'POST':
        try:
            if bill.approval_status != 'approved':
                flash('Bill must be approved before payment.', 'error')
                return redirect(url_for('purchasing.pay_vendor_bill', id=bill.id))

            amount = float(request.form.get('amount'))
            if amount <= 0 or amount > outstanding:
                flash('Payment amount must be positive and not exceed outstanding.', 'error')
                return redirect(url_for('purchasing.pay_vendor_bill', id=bill.id))

            payment = VendorPayment(
                vendor_bill_id=bill.id,
                payment_date=datetime.strptime(request.form.get('payment_date'), '%Y-%m-%d').date(),
                amount=amount,
                payment_mode=request.form.get('payment_mode'),
                reference_number=request.form.get('reference_number'),
                bank_name=request.form.get('bank_name'),
                status=request.form.get('status', 'cleared'),
                recorded_by=current_user.id
            )

            db.session.add(payment)

            bill.paid_amount = bill.paid_amount + amount
            if bill.paid_amount >= bill.total_amount:
                bill.status = 'paid'
            elif bill.paid_amount > 0:
                bill.status = 'partial'
            else:
                bill.status = 'pending'

            create_vendor_payment_entries(bill, payment)
            db.session.commit()

            flash('Vendor payment recorded successfully!', 'success')
            return redirect(url_for('purchasing.view_vendor_bill', id=bill.id))
        except Exception as e:
            db.session.rollback()
            flash(f'Error recording payment: {str(e)}', 'error')

    today = date.today().strftime('%Y-%m-%d')
    return render_template('purchasing/vendor_bill_pay.html', bill=bill, outstanding=outstanding, today=today)


@bp.route('/vendor-bills/<int:id>/print')
@login_required
def print_vendor_bill(id):
    """Print vendor bill"""
    bill = VendorBill.query.get_or_404(id)
    return render_template('purchasing/vendor_bill_print.html', bill=bill)
